import base64
import csv
import hashlib
import hmac
import html
import io
import json
import os
import re
import secrets
import sqlite3
import time
import urllib.error
import urllib.request
import uuid
import zipfile
from datetime import date, datetime, time as datetime_time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional, Union
from urllib.parse import urlencode

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response


APP_DIR = Path(__file__).resolve().parent


def load_dotenv(path: Path) -> None:
    """轻量读取 .env 到 os.environ，无第三方依赖。已存在的环境变量优先。"""
    if not path.is_file():
        return
    for line in path.read_text("utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


load_dotenv(APP_DIR / ".env")

DB_PATH = Path(os.environ.get("CLAIM_POOL_DB", APP_DIR / "claim_pool.db"))
UPLOAD_DIR = Path(os.environ.get("CLAIM_UPLOAD_DIR", APP_DIR / "uploads"))
MAX_ATTACHMENT_BYTES = 20 * 1024 * 1024
ALLOWED_ATTACHMENT_EXTENSIONS = {".png", ".jpg", ".jpeg", ".pdf", ".csv", ".xls", ".xlsx"}
BROAD_TERMS = {"公司", "有限公司", "集团", "科技", "教育", "转账", "付款", "收入"}

# 部门 → 中心/小组 → 项目 三级分类，来源：智库产品分类.xlsx
# 更新分类时直接替换 catalog.json 即可，无需改代码
CATALOG_PATH = Path(os.environ.get("CLAIM_CATALOG", APP_DIR / "catalog.json"))
BASE_CATALOG: dict[str, dict[str, list[str]]] = (
    json.loads(CATALOG_PATH.read_text("utf-8")) if CATALOG_PATH.is_file() else {}
)
CATALOG: dict[str, dict[str, list[str]]] = {
    department: {team: list(projects) for team, projects in teams.items()}
    for department, teams in BASE_CATALOG.items()
}
DEPARTMENTS = list(BASE_CATALOG)

# ── 飞书登录配置 ──────────────────────────────────────────────
# 这些值从 .env 读取（见 .env.example）。未配置时飞书登录入口自动隐藏，
# 系统仍可用 URL 参数身份（本地开发/演示）访问。
FEISHU_APP_ID = os.environ.get("FEISHU_APP_ID", "")
FEISHU_APP_SECRET = os.environ.get("FEISHU_APP_SECRET", "")
FEISHU_REDIRECT_URI = os.environ.get("FEISHU_REDIRECT_URI", "")
FEISHU_SCOPE = os.environ.get("FEISHU_SCOPE", "contact:user.base:readonly")
# 超级管理员 open_id 白名单（根权限，逗号分隔）。这一级只能改 .env，
# 普通管理员由超管在后台勾选、存数据库。兼容旧变量名 FEISHU_ADMIN_OPEN_IDS。
FEISHU_SUPERADMIN_OPEN_IDS = {
    x.strip()
    for x in (
        os.environ.get("FEISHU_SUPERADMIN_OPEN_IDS")
        or os.environ.get("FEISHU_ADMIN_OPEN_IDS", "")
    ).split(",")
    if x.strip()
}
# 给会话 cookie 签名用，必须保密；未设置则随机生成（重启后旧会话失效）
SESSION_SECRET = os.environ.get("SESSION_SECRET", secrets.token_hex(32)).encode()
SESSION_COOKIE = "claim_session"
SESSION_MAX_AGE = 7 * 24 * 3600  # 7 天


def env_bool(name: str, default: bool) -> bool:
    raw = os.environ.get(name)
    if raw is None or not raw.strip():
        return default
    return raw.strip().lower() not in {"0", "false", "no"}


# 未显式配置时，按回调地址协议决定 cookie 是否加 Secure。
# HTTP 临时部署可登录；HTTPS 正式部署仍默认走安全 cookie。
COOKIE_SECURE = env_bool("COOKIE_SECURE", FEISHU_REDIRECT_URI.lower().startswith("https://"))
HSTS_ENABLED = os.environ.get("HSTS_ENABLED", "true").lower() not in {"0", "false", "no"}
REQUIRE_LOGIN_FOR_CLAIM = os.environ.get("REQUIRE_LOGIN_FOR_CLAIM", "true").lower() not in {"0", "false", "no"}
RATE_LIMIT_WINDOW = int(os.environ.get("RATE_LIMIT_WINDOW", "60"))
RATE_LIMITS = {
    "search": int(os.environ.get("RATE_LIMIT_SEARCH", "60")),
    "login": int(os.environ.get("RATE_LIMIT_LOGIN", "20")),
    "import": int(os.environ.get("RATE_LIMIT_IMPORT", "10")),
    "export": int(os.environ.get("RATE_LIMIT_EXPORT", "20")),
}
RATE_BUCKETS: dict[tuple[str, str], list[float]] = {}

FEISHU_AUTHORIZE_URL = "https://accounts.feishu.cn/open-apis/authen/v1/authorize"
FEISHU_TOKEN_URL = "https://open.feishu.cn/open-apis/authen/v2/oauth/token"
FEISHU_USERINFO_URL = "https://open.feishu.cn/open-apis/authen/v1/user_info"
FEISHU_TENANT_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
FEISHU_SEND_MSG_URL = "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=open_id"
FEISHU_SEND_CHAT_MSG_URL = "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id"
FEISHU_NOTIFY_CHAT_ID = os.environ.get("FEISHU_NOTIFY_CHAT_ID", "").strip()
# 应用对外根地址（用于消息里的查看链接），从回调地址推导
APP_BASE_URL = os.environ.get("APP_BASE_URL") or FEISHU_REDIRECT_URI.replace("/oauth/callback", "")

# OCR 配置：none / tencent。腾讯云 OCR 用于图片型银行回单 PDF。
OCR_PROVIDER = os.environ.get("OCR_PROVIDER", "none").strip().lower()
TENCENTCLOUD_SECRET_ID = os.environ.get("TENCENTCLOUD_SECRET_ID", "")
TENCENTCLOUD_SECRET_KEY = os.environ.get("TENCENTCLOUD_SECRET_KEY", "")
TENCENTCLOUD_REGION = os.environ.get("TENCENTCLOUD_REGION", "ap-guangzhou")
TENCENT_OCR_HOST = "ocr.tencentcloudapi.com"
TENCENT_OCR_SERVICE = "ocr"
TENCENT_OCR_VERSION = "2018-11-19"


def feishu_enabled() -> bool:
    return bool(FEISHU_APP_ID and FEISHU_APP_SECRET and FEISHU_REDIRECT_URI)


MANAGED_ROLES = {"claimant", "admin", "general_manager"}


def normalize_managed_role(role: str) -> str:
    role = (role or "").strip()
    return role if role in MANAGED_ROLES else "claimant"


def db_user_role(open_id: str) -> tuple[str, bool]:
    """返回数据库托管身份；兼容旧 is_admin 字段。"""
    with get_conn() as conn:
        try:
            row = conn.execute(
                "SELECT managed_role, is_admin FROM app_users WHERE open_id = ?",
                (open_id,),
            ).fetchone()
        except sqlite3.OperationalError:
            row = conn.execute(
                "SELECT is_admin FROM app_users WHERE open_id = ?",
                (open_id,),
            ).fetchone()
            return "claimant", bool(row and row["is_admin"])
    if not row:
        return "claimant", False
    return normalize_managed_role(row["managed_role"]), bool(row["is_admin"])


def compute_role(open_id: str) -> str:
    """根据超管白名单 + 数据库托管身份实时判定角色。"""
    if open_id in FEISHU_SUPERADMIN_OPEN_IDS:
        return "superadmin"
    managed_role, legacy_is_admin = db_user_role(open_id)
    if managed_role in {"admin", "general_manager"}:
        return managed_role
    if legacy_is_admin:
        return "admin"
    return "claimant"


app = FastAPI(title="飞书到款认领系统 MVP")


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "same-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:; "
        "form-action 'self'; "
        "frame-ancestors 'none'; "
        "base-uri 'self'",
    )
    if HSTS_ENABLED and request.url.scheme == "https":
        response.headers.setdefault("Strict-Transport-Security", "max-age=15552000; includeSubDomains")
    return response


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def safe_next_url(value: str) -> str:
    return value if value.startswith("/") and not value.startswith("//") else "/search"


def rate_limit(request: Request, bucket: str) -> None:
    limit = RATE_LIMITS.get(bucket, 60)
    if limit <= 0:
        return
    now = time.time()
    key = (bucket, client_ip(request))
    recent = [t for t in RATE_BUCKETS.get(key, []) if now - t < RATE_LIMIT_WINDOW]
    if len(recent) >= limit:
        raise HTTPException(status_code=429, detail="请求过于频繁，请稍后再试")
    recent.append(now)
    RATE_BUCKETS[key] = recent


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def refresh_catalog(conn: Optional[sqlite3.Connection] = None) -> None:
    CATALOG.clear()
    CATALOG.update(
        {
            department: {team: list(projects) for team, projects in teams.items()}
            for department, teams in BASE_CATALOG.items()
        }
    )
    owns_conn = conn is None
    catalog_conn = conn or get_conn()
    try:
        rows = catalog_conn.execute(
            "SELECT change_type, department, team, project, active FROM catalog_project_changes ORDER BY id"
        ).fetchall()
        for row in rows:
            if not row["active"]:
                continue
            if row["change_type"] == "department":
                CATALOG.setdefault(row["department"], {})
            elif row["change_type"] == "team" and row["department"] in CATALOG:
                CATALOG[row["department"]].setdefault(row["team"], [])
        DEPARTMENTS[:] = list(CATALOG)
        for row in rows:
            if row["change_type"] != "project":
                continue
            projects = CATALOG.get(row["department"], {}).get(row["team"])
            if projects is None:
                continue
            if row["active"] and row["project"] not in projects:
                projects.append(row["project"])
            elif not row["active"] and row["project"] in projects:
                projects.remove(row["project"])
    finally:
        if owns_conn:
            catalog_conn.close()


def init_db() -> None:
    UPLOAD_DIR.mkdir(exist_ok=True)
    with get_conn() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS import_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_name TEXT NOT NULL,
                created_at TEXT NOT NULL,
                created_by TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'draft',
                raw_count INTEGER NOT NULL DEFAULT 0,
                imported_count INTEGER NOT NULL DEFAULT 0,
                skipped_count INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER,
                imported_at TEXT NOT NULL,
                confirmed_at TEXT,
                received_date TEXT,
                received_time TEXT,
                payer_name TEXT,
                amount_cents INTEGER NOT NULL DEFAULT 0,
                bank_note TEXT,
                receiver_company TEXT,
                receiver_account TEXT,
                serial_no TEXT,
                source_ref TEXT,
                confidence REAL NOT NULL DEFAULT 1,
                status TEXT NOT NULL DEFAULT 'draft',
                claimed_department TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                contract_invoice TEXT,
                claim_note TEXT,
                finance_note TEXT,
                raw_json TEXT
            );

            CREATE TABLE IF NOT EXISTS claims (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payment_id INTEGER NOT NULL,
                department TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                customer_project TEXT,
                contract_invoice TEXT,
                note TEXT,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );

            CREATE TABLE IF NOT EXISTS user_profiles (
                open_id TEXT PRIMARY KEY,
                name TEXT,
                department TEXT,
                team TEXT,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS user_scopes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                open_id TEXT NOT NULL,
                department TEXT NOT NULL,
                team TEXT,
                label TEXT,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                created_by TEXT NOT NULL,
                created_by_name TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS app_users (
                open_id TEXT PRIMARY KEY,
                name TEXT,
                managed_role TEXT NOT NULL DEFAULT 'claimant',
                is_admin INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                last_login TEXT
            );

            CREATE TABLE IF NOT EXISTS catalog_project_changes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                change_type TEXT NOT NULL DEFAULT 'project',
                department TEXT NOT NULL,
                team TEXT NOT NULL,
                project TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT NOT NULL,
                updated_by TEXT NOT NULL,
                updated_by_name TEXT NOT NULL,
                UNIQUE(department, team, project)
            );

            CREATE INDEX IF NOT EXISTS idx_payments_status ON payments(status);
            CREATE INDEX IF NOT EXISTS idx_payments_date ON payments(received_date);
            CREATE INDEX IF NOT EXISTS idx_claims_payment ON claims(payment_id);
            CREATE INDEX IF NOT EXISTS idx_audit_payment ON audit_logs(payment_id);
            """
        )
        ensure_column(conn, "payments", "claimed_team", "claimed_team TEXT")
        ensure_column(conn, "payments", "closed_at", "closed_at TEXT")
        ensure_column(conn, "payments", "receiver_company", "receiver_company TEXT")
        ensure_column(conn, "claims", "team", "team TEXT")
        ensure_column(conn, "claims", "amount_cents", "amount_cents INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "app_users", "managed_role", "managed_role TEXT NOT NULL DEFAULT 'claimant'")
        ensure_column(
            conn,
            "catalog_project_changes",
            "change_type",
            "change_type TEXT NOT NULL DEFAULT 'project'",
        )
        repair_compact_payment_dates(conn)
        conn.execute(
            "UPDATE payments SET closed_at = ? WHERE status = 'closed' AND COALESCE(closed_at, '') = ''",
            (now_text(),),
        )
        repair_rejected_payments_to_pending(conn)
        repair_pending_claims_to_accepted(conn)
        refresh_catalog(conn)


def ensure_column(conn: sqlite3.Connection, table: str, column: str, ddl: str) -> None:
    cols = [row[1] for row in conn.execute(f"PRAGMA table_info({table})")]
    if column not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")


def repair_compact_payment_dates(conn: sqlite3.Connection) -> int:
    repaired = 0
    rows = conn.execute(
        "SELECT id, received_date FROM payments WHERE LENGTH(COALESCE(received_date, '')) = 8"
    ).fetchall()
    for row in rows:
        raw_date = str(row["received_date"] or "").strip()
        if not re.fullmatch(r"\d{8}", raw_date):
            continue
        normalized = parse_date(raw_date)
        if normalized == raw_date:
            continue
        conn.execute("UPDATE payments SET received_date = ? WHERE id = ?", (normalized, row["id"]))
        repaired += 1
    return repaired


@app.on_event("startup")
def startup() -> None:
    init_db()


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def url(path: str, **params: str) -> str:
    return f"{path}?{urlencode(params)}"


def money(cents: Optional[int]) -> str:
    return f"{(cents or 0) / 100:,.2f}"


def receiver_company_label(value: Any) -> str:
    return str(value or "").strip() or "未填写"


def row_value(row: Any, key: str, default: Any = "") -> Any:
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        if key in row.keys():
            value = row[key]
            return default if value is None else value
    except (AttributeError, KeyError, IndexError):
        pass
    return default


def build_batch_confirm_message() -> str:
    return "\n".join(
        [
            "【今日到款已入池】",
            "新一批到款流水已完成入池，请相关同事进入财务到款认领系统查看并认领。",
        ]
    )


def build_claim_reject_message(
    payment: Union[sqlite3.Row, dict[str, Any]],
    claim: Union[sqlite3.Row, dict[str, Any]],
) -> str:
    return "\n".join(
        [
            "【部分到款认领被驳回】",
            "你提交的部分到款认领已被财务驳回。",
            f"付款方：{payment['payer_name']}",
            f"到款日期：{payment['received_date']}",
            f"认领金额：¥ {money(claim['amount_cents'])}",
            f"认领归属：{claim['department']} / {claim['team']} · {claim['customer_project']}",
            "驳回原因：未填写",
        ]
    )


def build_payment_reject_message(payment: Union[sqlite3.Row, dict[str, Any]], reason: str) -> str:
    return "\n".join(
        [
            "【到款认领被驳回】",
            "你提交的到款认领已被财务驳回。",
            f"付款方：{payment['payer_name']}",
            f"到款日期：{payment['received_date']}",
            f"认领归属：{payment['claimed_department'] or ''} / {payment['claimed_team'] or ''} · {payment['customer_project'] or ''}",
            f"驳回原因：{reason.strip() or '未填写'}",
        ]
    )


def build_claim_cancel_admin_message(
    payment: Union[sqlite3.Row, dict[str, Any]],
    claim: Union[sqlite3.Row, dict[str, Any]],
    actor: dict[str, str],
    payment_status: str,
) -> str:
    status_labels = {
        "pending": "待认领",
        "partial_claiming": "部分认领中",
        "claimed": "已认领",
        "pending_confirm": "已认领",
        "rejected": "已驳回",
        "closed": "已关闭",
    }
    department = row_value(claim, "department") or row_value(payment, "claimed_department")
    team = row_value(claim, "team") or row_value(payment, "claimed_team")
    project = row_value(claim, "customer_project") or row_value(payment, "customer_project")
    owner = " / ".join(part for part in [department, team] if part)
    if project:
        owner = f"{owner} · {project}" if owner else project
    return "\n".join(
        [
            "【认领已取消】",
            f"{actor.get('name') or '有同事'}取消了一笔到款认领。",
            f"付款方：{row_value(payment, 'payer_name') or '未填写付款方'}",
            f"到款日期：{row_value(payment, 'received_date') or '未填写日期'}",
            f"到款金额：¥ {money(row_value(payment, 'amount_cents', 0))}",
            f"取消认领金额：¥ {money(row_value(claim, 'amount_cents', 0))}",
            f"原认领归属：{owner or '未填写'}",
            f"当前状态：{status_labels.get(payment_status, payment_status or '未知')}",
        ]
    )


def admin_notification_open_ids(conn: sqlite3.Connection) -> list[str]:
    try:
        rows = conn.execute(
            """
            SELECT open_id
            FROM app_users
            WHERE managed_role = 'admin' OR is_admin = 1
            ORDER BY last_login DESC
            """
        ).fetchall()
    except sqlite3.OperationalError:
        return []
    open_ids: list[str] = []
    for row in rows:
        open_id = str(row["open_id"] or "").strip()
        if open_id and open_id not in FEISHU_SUPERADMIN_OPEN_IDS and open_id not in open_ids:
            open_ids.append(open_id)
    return open_ids


def notify_admins_claim_canceled(
    conn: sqlite3.Connection,
    payment: Union[sqlite3.Row, dict[str, Any]],
    claim: Union[sqlite3.Row, dict[str, Any]],
    actor: dict[str, str],
    payment_status: str,
) -> dict[str, Any]:
    admin_open_ids = admin_notification_open_ids(conn)
    message = build_claim_cancel_admin_message(payment, claim, actor, payment_status)
    sent_count = 0
    for open_id in admin_open_ids:
        if feishu_send_text(open_id, message):
            sent_count += 1
    return {
        "admin_notify_count": len(admin_open_ids),
        "admin_notified_count": sent_count,
        "admin_notified": sent_count > 0,
    }


def claim_note_summary(row: sqlite3.Row, claim_rows: list[sqlite3.Row]) -> str:
    notes: list[str] = []
    primary_note = str(row["claim_note"] or "").strip()
    if primary_note:
        notes.append(primary_note)
    for claim in claim_rows:
        note = str(claim["note"] or "").strip()
        if note and note not in notes:
            notes.append(note)
    return "；".join(notes)


def admin_claim_details_html(row: sqlite3.Row, claim_rows: list[sqlite3.Row]) -> str:
    active_claims = [claim for claim in claim_rows if claim["status"] in {"pending", "accepted"}]
    if len(active_claims) <= 1:
        return ""
    lines: list[str] = []
    for claim in active_claims:
        department = str(claim["department"] or "").strip() or "未填写部门"
        team = str(claim["team"] or "").strip() or "未填写中心"
        project = str(claim["customer_project"] or "").strip() or "未填写项目"
        actor_name = str(claim["actor_name"] or "").strip() or "未填写认领人"
        lines.append(
            f"{department} · {team} · {project} · {actor_name} · ¥ {money(claim['amount_cents'])}"
        )
    return '<div class="muted">认领明细：<br>' + "<br>".join(esc(line) for line in lines) + "</div>"


def parse_amount(value: Any) -> int:
    text = str(value or "").strip()
    text = text.replace(",", "").replace("￥", "").replace("¥", "").replace("元", "")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text:
        return 0
    try:
        return int((Decimal(text) * 100).quantize(Decimal("1")))
    except (InvalidOperation, ValueError):
        return 0


def parse_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    compact_match = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", text)
    if compact_match:
        y, m, d = compact_match.groups()
        return f"{y}-{int(m):02d}-{int(d):02d}"
    text = text.replace("/", "-").replace(".", "-")
    match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text)
    if match:
        y, m, d = match.groups()
        return f"{y}-{int(m):02d}-{int(d):02d}"
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        y, m, d = match.groups()
        return f"{y}-{int(m):02d}-{int(d):02d}"
    match = re.search(r"(\d{1,2})-(\d{1,2})", text)
    if match:
        m, d = match.groups()
        return f"{datetime.now().year}-{int(m):02d}-{int(d):02d}"
    return text


def save_attachment(attachment: Optional[UploadFile]) -> str:
    if not attachment or not attachment.filename:
        return ""

    original_name = Path(attachment.filename).name
    suffix = Path(original_name).suffix.lower()
    if suffix not in ALLOWED_ATTACHMENT_EXTENSIONS:
        raise HTTPException(status_code=400, detail="附件格式暂不支持")

    content = attachment.file.read()
    if len(content) > MAX_ATTACHMENT_BYTES:
        raise HTTPException(status_code=400, detail="附件不能超过 20MB")

    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(original_name).stem).strip("._")
    if not safe_stem:
        safe_stem = "attachment"
    stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:10]}_{safe_stem}{suffix}"
    target = UPLOAD_DIR / stored_name
    target.write_bytes(content)
    return f"uploads/{stored_name}"


def attachment_link(source_ref: str, actor: dict[str, str]) -> str:
    if not source_ref or not source_ref.startswith("uploads/"):
        return ""
    filename = Path(source_ref).name
    href = url(f"/attachments/{filename}", role=actor["role"], user=actor["id"], name=actor["name"])
    return f'<a href="{esc(href)}" target="_blank">查看附件</a>'


# ── 会话（签名 cookie，无第三方依赖） ───────────────────────────
def _b64e(raw: bytes) -> str:
    return base64.urlsafe_b64encode(raw).decode().rstrip("=")


def _b64d(text: str) -> bytes:
    return base64.urlsafe_b64decode(text + "=" * (-len(text) % 4))


def make_session(data: dict[str, Any]) -> str:
    payload = dict(data)
    payload["exp"] = int(time.time()) + SESSION_MAX_AGE
    body = _b64e(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode())
    sig = _b64e(hmac.new(SESSION_SECRET, body.encode(), hashlib.sha256).digest())
    return f"{body}.{sig}"


def read_session(token: str) -> Optional[dict[str, Any]]:
    if not token or "." not in token:
        return None
    body, _, sig = token.partition(".")
    expected = _b64e(hmac.new(SESSION_SECRET, body.encode(), hashlib.sha256).digest())
    if not hmac.compare_digest(sig, expected):
        return None
    try:
        data = json.loads(_b64d(body))
    except Exception:
        return None
    if int(data.get("exp", 0)) < int(time.time()):
        return None
    return data


# ── 飞书 API 调用（标准库 urllib，自动尊重环境代理） ──────────────
def _feishu_request(
    url: str, method: str = "GET", json_body: Optional[dict] = None, bearer: str = ""
) -> dict[str, Any]:
    headers: dict[str, str] = {}
    data = None
    if json_body is not None:
        data = json.dumps(json_body).encode()
        headers["Content-Type"] = "application/json"
    if bearer:
        headers["Authorization"] = f"Bearer {bearer}"
    req = urllib.request.Request(url, data=data, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as exc:
        try:
            return json.loads(exc.read().decode())
        except Exception:
            return {"error": "http_error", "status": exc.code}
    except Exception as exc:  # noqa: BLE001
        return {"error": "network_error", "message": str(exc)}


def feishu_exchange_token(code: str) -> str:
    resp = _feishu_request(
        FEISHU_TOKEN_URL,
        "POST",
        {
            "grant_type": "authorization_code",
            "client_id": FEISHU_APP_ID,
            "client_secret": FEISHU_APP_SECRET,
            "code": code,
            "redirect_uri": FEISHU_REDIRECT_URI,
        },
    )
    return resp.get("access_token", "")


def feishu_get_userinfo(user_token: str) -> dict[str, Any]:
    resp = _feishu_request(FEISHU_USERINFO_URL, "GET", bearer=user_token)
    return resp.get("data") or {}


def feishu_tenant_token() -> str:
    resp = _feishu_request(
        FEISHU_TENANT_TOKEN_URL,
        "POST",
        {"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET},
    )
    return resp.get("tenant_access_token", "")


def feishu_send_text(open_id: str, text: str) -> bool:
    """以应用身份给用户发飞书单聊文本。容错：失败返回 False，不抛异常。
    需要权限 im:message:send_as_bot + 应用开启机器人能力。"""
    if not (feishu_enabled() and open_id):
        return False
    token = feishu_tenant_token()
    if not token:
        return False
    resp = _feishu_request(
        FEISHU_SEND_MSG_URL,
        "POST",
        {"receive_id": open_id, "msg_type": "text", "content": json.dumps({"text": text}, ensure_ascii=False)},
        bearer=token,
    )
    return resp.get("code") == 0


def feishu_send_chat_text(chat_id: str, text: str) -> bool:
    """以应用机器人身份给群聊发文本。容错：失败返回 False，不影响业务动作。"""
    if not (feishu_enabled() and chat_id):
        return False
    token = feishu_tenant_token()
    if not token:
        return False
    resp = _feishu_request(
        FEISHU_SEND_CHAT_MSG_URL,
        "POST",
        {"receive_id": chat_id, "msg_type": "text", "content": json.dumps({"text": text}, ensure_ascii=False)},
        bearer=token,
    )
    return resp.get("code") == 0


def get_user_profile(open_id: str) -> Optional[sqlite3.Row]:
    if not open_id:
        return None
    with get_conn() as conn:
        return conn.execute(
            "SELECT department, team FROM user_profiles WHERE open_id = ?", (open_id,)
        ).fetchone()


def can_self_set_profile(conn: sqlite3.Connection, open_id: str) -> bool:
    row = conn.execute(
        "SELECT department, team FROM user_profiles WHERE open_id = ?", (open_id,)
    ).fetchone()
    if not row:
        return True
    return not ((row["department"] or "").strip() and (row["team"] or "").strip())


def get_user_scopes(conn: sqlite3.Connection, open_id: str, active_only: bool = True) -> list[sqlite3.Row]:
    if not open_id:
        return []
    where = "WHERE open_id = ?"
    params: list[Any] = [open_id]
    if active_only:
        where += " AND active = 1"
    return conn.execute(
        f"""
        SELECT *
        FROM user_scopes
        {where}
        ORDER BY id DESC
        """,
        params,
    ).fetchall()


def scope_display_label(department: str, team: str = "", label: str = "") -> str:
    label = (label or "").strip()
    if label:
        return label
    team = (team or "").strip()
    return f"{department} / {team}" if team else department


def dashboard_scope_choices(
    actor: dict[str, Any],
    extra_scopes: list[sqlite3.Row],
) -> list[dict[str, Any]]:
    scope_items: list[dict[str, Any]] = []

    def add_scope(key: str, label: str, department: str, team: str = "") -> None:
        department = (department or "").strip()
        team = (team or "").strip()
        if department not in DEPARTMENTS:
            return
        if team and team not in CATALOG.get(department, {}):
            return
        identity = (department, team)
        if any(item["identity"] == identity for item in scope_items):
            return
        scope_items.append(
            {
                "key": key,
                "label": label,
                "identity": identity,
                "scopes": [{"department": department, "team": team}],
            }
        )

    primary_department = actor.get("department", "")
    primary_team = actor.get("team", "")
    add_scope(
        "primary",
        "主身份：" + scope_display_label(primary_department, primary_team),
        primary_department,
        primary_team,
    )
    for row in extra_scopes:
        add_scope(
            f"extra:{row['id']}",
            scope_display_label(row["department"], row["team"] or "", row["label"] or ""),
            row["department"],
            row["team"] or "",
        )

    all_scopes = []
    seen: set[tuple[str, str]] = set()
    for item in scope_items:
        for scope in item["scopes"]:
            identity = (scope["department"], scope.get("team", ""))
            if identity not in seen:
                seen.add(identity)
                all_scopes.append(scope)
    return [{"key": "all", "label": "全部角色", "scopes": all_scopes, "identity": ("", "")}, *scope_items]


def actor_from_request(request: Request) -> dict[str, str]:
    session = read_session(request.cookies.get(SESSION_COOKIE, ""))
    if session:
        open_id = session.get("id", "feishu-user")
        profile = get_user_profile(open_id)
        department = (profile["department"] if profile else "") or ""
        team = (profile["team"] if profile else "") or ""
        return {
            "id": open_id,
            "name": session.get("name", "飞书用户"),
            "role": compute_role(open_id),  # 实时判定，超管增减管理员立即生效
            "department": department or "未设置部门",
            "team": team,
            "authed": "1",
        }
    # 无会话 = 未授权：身份字段仅供 demo 展示，role 一律强制为最低权限 claimant，
    # 绝不从 query/header 读取角色，杜绝 ?role=admin / x-role 之类的越权。
    params = request.query_params
    return {
        "id": params.get("user") or request.headers.get("x-user-id") or "demo-user",
        "name": params.get("name") or request.headers.get("x-user-name") or "演示用户",
        "role": "claimant",
        "department": params.get("department") or request.headers.get("x-department") or "未设置部门",
        "team": params.get("team") or "",
        "authed": "",
    }


def actor_from_form(
    user: str,
    name: str,
    role: str,
    department: str = "",
) -> dict[str, str]:
    return {
        "id": user or "demo-user",
        "name": name or "演示用户",
        "role": role or "claimant",
        "department": department or "未设置部门",
    }


def require_admin(actor: dict[str, str]) -> None:
    if actor["role"] not in {"finance", "admin", "superadmin"}:
        raise HTTPException(status_code=403, detail="只有管理员可以访问这个页面")


def require_superadmin(actor: dict[str, str]) -> None:
    if actor["role"] != "superadmin":
        raise HTTPException(status_code=403, detail="只有超级管理员可以管理管理员")


def audit(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    action: str,
    payment_id: Optional[int] = None,
    detail: Optional[dict[str, Any]] = None,
    request: Optional[Request] = None,
) -> None:
    conn.execute(
        """
        INSERT INTO audit_logs
            (at, actor_id, actor_name, actor_role, action, payment_id, detail_json, ip)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            now_text(),
            actor["id"],
            actor["name"],
            actor["role"],
            action,
            payment_id,
            json.dumps(detail or {}, ensure_ascii=False),
            client_ip(request) if request else "",
        ),
    )


BASE_CSS = """
    :root {
      color-scheme: light;
      --bg:#f3f5f8; --card:#ffffff;
      --line:#e5e8ee; --line-strong:#cfd5df;
      --text:#1a2334; --muted:#5d6b84; --faint:#8d99ad;
      --primary:#2456d6; --primary-dark:#1b44b0; --primary-soft:#e9effc;
      --radius:12px;
      --shadow:0 1px 2px rgba(23,32,51,.05), 0 1px 3px rgba(23,32,51,.06);
    }
    * { box-sizing:border-box; }
    body { margin:0; background:var(--bg); color:var(--text);
      font:14px/1.6 -apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Hiragino Sans GB","Microsoft YaHei",sans-serif; }
    a { color:var(--primary); text-decoration:none; }
    a:hover { text-decoration:underline; }

    header { position:sticky; top:0; z-index:10; background:rgba(255,255,255,.92); backdrop-filter:blur(8px);
      border-bottom:1px solid var(--line); display:flex; align-items:center; justify-content:space-between;
      gap:16px; padding:10px 28px; }
    .brand { display:flex; align-items:center; gap:10px; }
    .brand-mark { width:34px; height:34px; border-radius:9px; background:linear-gradient(135deg,#2456d6,#4f7df7);
      color:#fff; display:flex; align-items:center; justify-content:center; font-size:17px; font-weight:700; }
    .brand-text { display:flex; flex-direction:column; line-height:1.3; }
    .brand-text strong { font-size:15px; }
    .brand-text small { font-size:11px; color:var(--faint); }
    nav { display:flex; gap:6px; }
    nav a { padding:7px 14px; border-radius:8px; color:var(--muted); font-weight:500; }
    nav a:hover { background:#f0f3f8; color:var(--text); text-decoration:none; }
    nav a.active { background:var(--primary-soft); color:var(--primary); }
    .login-btn { display:inline-block; padding:7px 16px; border-radius:8px; background:var(--primary);
      color:#fff; font-weight:500; white-space:nowrap; }
    .login-btn:hover { background:var(--primary-dark); text-decoration:none; }
    .user-area { white-space:nowrap; font-size:13px; }

    main { max-width:1180px; margin:0 auto; padding:28px 28px 64px; }
    .page-head { margin-bottom:20px; }
    h1 { font-size:24px; margin:0; letter-spacing:-.01em; }
    .page-sub { margin:4px 0 0; color:var(--muted); font-size:13px; }
    h2 { font-size:16px; margin:32px 0 12px; display:flex; align-items:center; gap:8px; }
    h2::before { content:""; width:4px; height:16px; border-radius:2px; background:var(--primary); }

    .panel { background:var(--card); border:1px solid var(--line); border-radius:var(--radius);
      padding:20px; margin-bottom:20px; box-shadow:var(--shadow); }
    .muted { color:var(--muted); font-size:12.5px; }
    .hint { color:var(--faint); font-size:12.5px; margin:12px 0 0; }
    .nowrap { white-space:nowrap; }

    .row { display:flex; gap:12px; flex-wrap:wrap; align-items:end; }
    .field { margin-bottom:10px; }
    label { display:block; font-size:12.5px; font-weight:500; color:var(--muted); margin:0 0 6px; }
    input, select, textarea { width:100%; border:1px solid var(--line-strong); border-radius:8px;
      padding:9px 12px; font:inherit; background:#fff; color:var(--text);
      transition:border-color .15s, box-shadow .15s; }
    input:focus, select:focus, textarea:focus { outline:none; border-color:var(--primary);
      box-shadow:0 0 0 3px rgba(36,86,214,.13); }
    textarea { min-height:140px; font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; font-size:12.5px; }
    input[type=file] { padding:7px; background:#fafbfd; }
    button { border:0; border-radius:8px; padding:9px 18px; background:var(--primary); color:#fff;
      font:inherit; font-weight:500; cursor:pointer; transition:background .15s; }
    button:hover { background:var(--primary-dark); }
    button.secondary { background:#fff; color:var(--text); border:1px solid var(--line-strong); }
    button.secondary:hover { background:#f4f6fa; }
    .secondary-link { display:inline-block; padding:8px 14px; border:1px solid var(--line-strong);
      border-radius:8px; background:#fff; color:var(--text); font-weight:500; white-space:nowrap; }
    .secondary-link:hover { background:#f4f6fa; text-decoration:none; }
    button.danger { background:#c2362b; }
    button.success { background:#16a34a; }
    button.success:hover { background:#15803d; }
    button:disabled { opacity:.55; cursor:not-allowed; }
    button:disabled:hover { background:var(--primary); }

    .table-wrap { background:var(--card); border:1px solid var(--line); border-radius:var(--radius);
      box-shadow:var(--shadow); margin-bottom:20px; overflow-x:auto; }
    table { width:100%; border-collapse:collapse; }
    th { padding:10px 16px; background:#f8fafc; border-bottom:1px solid var(--line);
      font-size:12px; font-weight:600; color:var(--muted); text-align:left; white-space:nowrap; }
    td { padding:13px 16px; border-bottom:1px solid var(--line); font-size:13px; vertical-align:top; }
    tbody tr:last-child td { border-bottom:0; }
    tbody tr:hover { background:#fafbfd; }
    td.num { font-variant-numeric:tabular-nums; font-weight:600; white-space:nowrap; }
    td.payment-summary { min-width:280px; white-space:normal; overflow-wrap:anywhere; }
    td.empty { padding:36px; text-align:center; color:var(--faint); }
    .select-cell { width:38px; text-align:center; padding-left:12px; padding-right:8px; }
    .select-cell input { width:auto; margin:0; box-shadow:none; }
    .code { font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; font-size:11.5px;
      color:var(--muted); word-break:break-all; }
    .sort-link { display:inline-flex; align-items:center; gap:5px; color:var(--muted); font-weight:600; }
    .sort-link:hover { color:var(--primary); text-decoration:none; }
    .sort-link.active { color:var(--primary); }
    .sort-arrow { font-size:11px; line-height:1; }

    .status { display:inline-flex; align-items:center; gap:6px; padding:3px 10px; border-radius:999px;
      font-size:12px; font-weight:500; background:#eef1f5; color:#475569; white-space:nowrap; }
    .status::before { content:""; width:6px; height:6px; border-radius:50%; background:currentColor; flex:none; }
    .status.draft { background:#e8edf5; color:#3b5275; }
    .status.pending { background:#fdf2d9; color:#9a6700; }
    .status.partial_claiming { background:#e9effc; color:#2456d6; }
    .status.claimed { background:#dcf5e7; color:#13794c; }
    .status.pending_confirm { background:#fdeaec; color:#b4232e; }
    .status.rejected { background:#efeff2; color:#6b7280; }
    .status.canceled { background:#efeff2; color:#6b7280; }
    .status.closed { background:#e7e9fd; color:#4341c8; }

    .grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(165px,1fr)); gap:14px; margin-bottom:8px; }
    .admin-stat-grid { grid-template-columns:repeat(6,minmax(0,1fr)); }
    .stat { background:var(--card); border:1px solid var(--line); border-radius:var(--radius);
      padding:16px 18px; box-shadow:var(--shadow); }
    .stat-label { font-size:12.5px; color:var(--muted); display:flex; align-items:center; gap:7px; }
    .stat-label::before { content:""; width:8px; height:8px; border-radius:50%; background:var(--dot,#94a3b8); flex:none; }
    .stat strong { display:block; font-size:26px; font-weight:650; margin:8px 0 2px;
      font-variant-numeric:tabular-nums; letter-spacing:-.02em; }
    .stat-amount { font-size:12px; color:var(--faint); font-variant-numeric:tabular-nums; }

    .callout { background:#fff; border:1px solid var(--line); border-left:4px solid var(--faint);
      border-radius:10px; padding:14px 18px; margin-bottom:20px; font-size:13.5px; box-shadow:var(--shadow); }
    .callout.warn { border-left-color:#e3a008; background:#fffdf5; }
    .callout.info { border-left-color:var(--primary); }
    .callout.success { border-left-color:#16a34a; background:#f3fcf7; }

    details.fold { border:1px solid var(--line); border-radius:10px; background:#fafbfd; margin-bottom:8px; }
    details.fold:last-child { margin-bottom:0; }
    details.fold summary { cursor:pointer; padding:8px 12px; font-size:12.5px; font-weight:500;
      color:var(--primary); list-style:none; display:flex; align-items:center; gap:6px; user-select:none; }
    details.fold summary::-webkit-details-marker { display:none; }
    details.fold summary::before { content:"\\25B8"; font-size:11px; transition:transform .15s; }
    details.fold[open] summary::before { transform:rotate(90deg); }
    details.fold .fold-body { padding:10px 12px 12px; border-top:1px dashed var(--line); }
    .actions { min-width:300px; }
    .split-form-cell { padding:0 16px 16px; background:#fff; }
    .split-form-cell details.fold { margin:0; }
    .split-form-cell details.fold .fold-body { padding:16px; }
    .split-form-table { min-width:960px; }
    .split-form-table th:last-child, .split-form-table td:last-child { min-width:260px; }
    .confirm-claim-form { margin-top:10px; }
    .confirm-claim-form button { width:100%; }
    .bulk-bar { display:flex; justify-content:space-between; align-items:center; gap:12px;
      padding:12px 16px; border-bottom:1px solid var(--line); background:#fbfcfe; }
    .bulk-bar form { display:flex; align-items:center; gap:10px; margin:0; }
    .bulk-count { color:var(--muted); font-size:12.5px; }
    .admin-payment-table { table-layout:fixed; min-width:1080px; }
    .admin-payment-table th, .admin-payment-table td { padding:12px 10px; }
    .admin-payment-table .col-select { width:40px; }
    .admin-payment-table .col-id { width:72px; }
    .admin-payment-table .col-date { width:112px; }
    .admin-payment-table .col-amount { width:126px; }
    .admin-payment-table .col-receiver { width:130px; }
    .admin-payment-table .col-status { width:190px; }
    .admin-payment-table .col-actions { width:218px; }
    .admin-payment-table .payment-summary { min-width:0; }
    .admin-payment-table td { overflow-wrap:anywhere; }
    .admin-payment-table .actions { min-width:0; }
    .admin-actions details.fold summary { padding:8px 10px; }
    .admin-actions button { width:100%; padding-left:10px; padding-right:10px; }
    .admin-actions .edit-row { display:grid; grid-template-columns:1fr 1fr; gap:8px; }
    .admin-actions .edit-row > div { width:auto !important; min-width:0; }
    .dash-grid { display:grid; grid-template-columns:1fr; gap:14px; margin-bottom:20px; }
    .dash-toolbar { display:flex; align-items:flex-end; justify-content:space-between; gap:16px;
      flex-wrap:wrap; margin-bottom:12px; }
    .dash-toolbar h2 { margin-bottom:6px; }
    .dash-toolbar .hint { margin:0; }
    .dash-date-filter { display:flex; align-items:flex-end; justify-content:flex-end; gap:8px;
      flex-wrap:wrap; margin-left:auto; }
    .dash-date-filter .date-field { min-width:150px; }
    .dash-date-filter .date-actions { display:flex; align-items:center; gap:8px; }
    .identity-card > div:nth-child(2) { min-width:0; }
    .dash-panel { margin-bottom:0; padding:16px; }
    .dash-head { display:flex; justify-content:space-between; gap:12px; align-items:flex-start; margin-bottom:14px; }
    .dash-amount { font-size:20px; font-weight:650; font-variant-numeric:tabular-nums; white-space:nowrap; }
    .dash-table-wrap { max-height:240px; overflow-y:auto; overflow-x:hidden; border-top:1px solid var(--line); }
    .dash-table-wrap::-webkit-scrollbar { width:8px; }
    .dash-table-wrap::-webkit-scrollbar-track { background:#eef1f5; border-radius:999px; }
    .dash-table-wrap::-webkit-scrollbar-thumb { background:#c7ceda; border-radius:999px; }
    .dash-table-wrap::-webkit-scrollbar-thumb:hover { background:#aeb8c8; }
    .dash-table { width:100%; table-layout:fixed; }
    .dash-table th { position:sticky; top:0; z-index:1; padding:8px 10px 8px 0; background:#fff;
      border-bottom:1px solid var(--line); font-size:11.5px; }
    .dash-table td { padding:8px 10px 8px 0; font-size:12.5px; }
    .dash-table td, .dash-table th { white-space:normal; overflow-wrap:anywhere; }
    .dash-table td:last-child, .dash-table th:last-child { padding-right:18px; text-align:right; }
    .dash-table .num { font-weight:650; white-space:nowrap; }
    .dash-claim-details { margin-top:5px; color:var(--muted); font-size:11.5px; line-height:1.55; }
    .dash-claim-details summary { width:max-content; max-width:100%; color:var(--primary); cursor:pointer; }
    .dash-claim-details div { margin-top:3px; overflow-wrap:anywhere; }
    .diagnostic-actions { display:flex; align-items:center; gap:10px; flex-wrap:wrap; margin-top:10px; }
    .copy-status { color:#16a34a; font-size:12.5px; min-height:18px; }
    .modal-backdrop { position:fixed; inset:0; z-index:100; background:rgba(15,23,42,.42);
      display:flex; align-items:center; justify-content:center; padding:20px; }
    .modal { width:min(560px,100%); background:#fff; border-radius:14px; box-shadow:0 18px 50px rgba(15,23,42,.22);
      padding:24px; border:1px solid var(--line); }
    .modal h2 { margin:0 0 8px; }
    .modal h2::before { display:none; }
    .modal .hint { margin:0 0 18px; }

    @media (max-width: 760px) {
      header { padding:10px 16px; }
      main { padding:20px 16px 48px; }
      .brand-text small { display:none; }
      .admin-stat-grid { grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); }
      table { min-width:680px; }
      .dash-table-wrap { overflow-x:hidden; }
      .dash-table { display:block; min-width:0; }
      .dash-table colgroup, .dash-table thead { display:none; }
      .dash-table tbody { display:block; }
      .dash-table tr { display:grid; grid-template-columns:minmax(0,1fr) auto; gap:5px 12px;
        padding:12px 0; border-bottom:1px solid var(--line); }
      .dash-table tr:last-child { border-bottom:0; }
      .dash-table td { display:block; min-width:0; padding:0; border:0; }
      .dash-table td:first-child { grid-column:1; grid-row:1; font-weight:600; }
      .dash-table td:nth-child(2), .dash-table td:nth-child(3), .dash-table td:nth-child(4) {
        grid-column:1 / -1; color:var(--muted); }
      .dash-table td:nth-child(2)::before { content:"到款公司："; }
      .dash-table td:nth-child(3)::before { content:"摘要："; }
      .dash-table td:nth-child(4)::before { content:"所属部门："; }
      .dash-table td:last-child { grid-column:2; grid-row:1; padding:0; text-align:right; }
      .dash-claim-details { color:var(--muted); }
      .dash-date-filter { width:100%; display:grid; grid-template-columns:1fr 1fr; margin-left:0; }
      .dash-date-filter .date-field { min-width:0; }
      .dash-date-filter .date-actions { grid-column:1 / -1; }
      .identity-card { display:grid !important; grid-template-columns:56px minmax(0,1fr);
        align-items:center !important; }
      .identity-card .identity-note { grid-column:1 / -1; max-width:none !important; text-align:left !important; }
    }
"""


def identity_params(actor: Optional[dict[str, str]]) -> dict[str, str]:
    if not actor:
        return {}
    return {
        "user": actor["id"],
        "name": actor["name"],
        "department": actor["department"],
        "role": actor["role"],
    }


def page(
    title: str,
    body: str,
    active: str = "",
    subtitle: str = "",
    actor: Optional[dict[str, str]] = None,
) -> HTMLResponse:
    # 携带身份参数在页面间跳转，为后续接入飞书登录后透传用户身份做准备
    ident = identity_params(actor)
    search_href = url("/search", **ident) if ident else "/search"
    split_href = url("/split-claim", **ident) if ident else "/split-claim"
    me_href = url("/me", **ident) if ident else "/me"
    nav_items = [
        (search_href, "认领搜索", "search"),
        (split_href, "分摊认领", "split"),
        (me_href, "个人中心", "me"),
    ]
    # 后台仅对真实管理员可见；不再用 ?role=admin 给所有人埋后门入口
    if actor and actor["role"] in {"finance", "admin", "superadmin"}:
        nav_items.append((url("/admin", **ident), "财务后台", "admin"))
    if actor and actor["role"] == "superadmin":
        nav_items.append((url("/admin/system", **ident), "管理后台", "system"))
    nav = "".join(
        f'<a href="{esc(href)}" class="{"active" if key == active else ""}">{label}</a>'
        for href, label, key in nav_items
    )
    if actor and actor.get("authed"):
        user_area = (
            f'<span class="muted">{esc(actor["name"])}</span>'
            f'<a href="/logout" style="margin-left:12px">退出</a>'
        )
    elif feishu_enabled():
        user_area = '<a href="/login" class="login-btn">飞书登录</a>'
    else:
        user_area = ""
    subtitle_html = f'<p class="page-sub">{esc(subtitle)}</p>' if subtitle else ""
    return HTMLResponse(
        f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title)} · 到款认领</title>
  <style>{BASE_CSS}</style>
</head>
<body>
  <header>
    <div class="brand">
      <span class="brand-mark">¥</span>
      <span class="brand-text"><strong>到款认领</strong><small>到款认领管理系统</small></span>
    </div>
    <div style="display:flex; align-items:center; gap:18px">
      <nav>{nav}</nav>
      <div class="user-area">{user_area}</div>
    </div>
  </header>
  <main>
    <div class="page-head"><h1>{esc(title)}</h1>{subtitle_html}</div>
    {body}
  </main>
</body>
</html>"""
    )


HEADER_ALIASES = {
    "received_date": {"到款日期", "日期", "入账日期", "交易日期", "交易日", "date"},
    "received_time": {"到款时间", "时间", "交易时间", "time"},
    "payer_name": {
        "付款方名称",
        "付款方",
        "对方户名",
        "对方名称",
        "客户名称",
        "付款附言",
        "收(付)方名称",
        "收（付）方名称",
        "payer",
        "payer_name",
    },
    "amount": {
        "到款金额",
        "金额",
        "收入",
        "收入金额",
        "贷方金额",
        "贷方金额（元）",
        "贷方金额(元)",
        "贷方发生额",
        "入账金额",
        "交易金额",
        "amount",
    },
    "bank_note": {"银行备注", "摘要", "备注", "用途", "附言", "付款附言", "note", "remark"},
    "receiver_company": {
        "到款公司",
        "收款公司",
        "收款方",
        "收款方名称",
        "收款户名",
        "收款单位",
        "收款人",
        "收款人名称",
        "我方公司",
        "receiver_company",
    },
    "receiver_account": {"收款账户", "账号", "账户", "account"},
    "serial_no": {"流水号", "回单号", "凭证号", "检索号", "商户订单号", "银商订单号", "serial_no", "serial"},
}


def canonical_header(header: str) -> str:
    key = header.strip().lower()
    for field, aliases in HEADER_ALIASES.items():
        if key in {alias.lower() for alias in aliases}:
            return field
    return key


def normalize_import_item(item: dict[str, str]) -> dict[str, str]:
    received_time = item.get("received_time", "").strip()
    if received_time and not item.get("received_date"):
        if re.search(r"\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}", received_time):
            item["received_date"] = received_time
            time_match = re.search(r"(\d{1,2}[:：]\d{2}(?::\d{2})?)", received_time)
            if time_match:
                item["received_time"] = time_match.group(1).replace("：", ":")

    payer_name = item.get("payer_name", "").strip()
    bank_note = item.get("bank_note", "").strip()
    if payer_name and not bank_note:
        item["bank_note"] = payer_name
    elif bank_note and not payer_name:
        item["payer_name"] = bank_note
    return item


def read_table(text: str) -> list[dict[str, str]]:
    sample = text.strip("\ufeff\n ")
    if not sample:
        return []
    dialect = csv.excel_tab if "\t" in sample.splitlines()[0] else csv.excel
    reader = csv.reader(io.StringIO(sample), dialect=dialect)
    rows = [list(row) for row in reader if any(str(cell).strip() for cell in row)]
    if not rows:
        return []

    first = [canonical_header(cell) for cell in rows[0]]
    has_header = bool({"received_date", "payer_name", "amount"} & set(first))
    data_rows = rows[1:] if has_header else rows
    headers = first if has_header else [
        "received_date",
        "payer_name",
        "amount",
        "bank_note",
        "serial_no",
        "receiver_account",
        "receiver_company",
    ]

    result = []
    for row in data_rows:
        item = {}
        for index, value in enumerate(row):
            if index < len(headers):
                key = headers[index]
                text = value.strip()
                if key and (key not in item or not item[key]):
                    item[key] = text
        result.append(normalize_import_item(item))
    return result


def excel_value_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime_time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def rows_from_excel_matrix(matrix: list[list[Any]]) -> list[dict[str, str]]:
    rows = [
        [excel_value_text(value) for value in row]
        for row in matrix
        if any(excel_value_text(value) for value in row)
    ]
    if not rows:
        return []

    required_headers = {"received_date", "payer_name", "amount"}
    header_index = None
    headers: list[str] = []
    for index, row in enumerate(rows[:20]):
        candidate = [canonical_header(cell) for cell in row]
        if len(required_headers & set(candidate)) >= 2:
            header_index = index
            headers = candidate
            break

    if header_index is None:
        header_index = -1
        headers = [
            "received_date",
            "payer_name",
            "amount",
            "bank_note",
            "serial_no",
            "receiver_account",
            "receiver_company",
        ]

    result = []
    for row in rows[header_index + 1:]:
        item: dict[str, str] = {}
        for index, value in enumerate(row):
            if index >= len(headers) or not headers[index]:
                continue
            key = headers[index]
            text = value.strip()
            if key not in item or not item[key]:
                item[key] = text
        if any(item.values()):
            result.append(normalize_import_item(item))
    return result


def is_ooxml_workbook(path: Path) -> bool:
    if not zipfile.is_zipfile(path):
        return False
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile:
        return False
    return "[Content_Types].xml" in names and any(
        name.startswith("xl/worksheets/") for name in names
    )


def rows_from_openpyxl(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    rows: list[dict[str, str]] = []
    with path.open("rb") as file_obj:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
                rows.extend(rows_from_excel_matrix(matrix))
        finally:
            workbook.close()
    return rows


def rows_from_xlrd(path: Path) -> list[dict[str, str]]:
    import xlrd

    rows: list[dict[str, str]] = []
    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        for worksheet in workbook.sheets():
            matrix = []
            for row_index in range(worksheet.nrows):
                row = []
                for cell in worksheet.row(row_index):
                    value: Any = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
                    row.append(value)
                matrix.append(row)
            rows.extend(rows_from_excel_matrix(matrix))
    finally:
        workbook.release_resources()
    return rows


def rows_from_excel(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx" or is_ooxml_workbook(path):
            return rows_from_openpyxl(path)
        elif suffix == ".xls":
            return rows_from_xlrd(path)
    except ImportError as exc:
        raise HTTPException(status_code=500, detail=f"缺少 Excel 解析依赖：{exc}") from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Excel 文件无法读取，请确认文件未损坏且未加密：{exc}") from exc
    return []


def extract_pdf_text(path: Path) -> str:
    """读取可复制文本 PDF；图片型/扫描件通常会返回空文本。"""
    try:
        from pypdf import PdfReader
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"缺少 PDF 解析依赖 pypdf：{exc}") from exc

    try:
        reader = PdfReader(str(path))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"PDF 文件无法读取：{exc}") from exc

    pages = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n".join(pages).strip()


def first_match(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, re.MULTILINE)
        if match:
            return match.group(1).strip()
    return ""


def parse_pdf_receipts(text: str) -> list[dict[str, str]]:
    """从银行入账回单文本里抽取流水。第一版偏保守，抽不到关键字段就不入库。"""
    normalized = re.sub(r"[ \t]+", " ", text.replace("\r", "\n"))
    blocks = [b.strip() for b in re.split(r"入\s*账\s*回\s*单", normalized) if b.strip()]
    if not blocks:
        blocks = [normalized]

    rows = []
    for block in blocks:
        received_date = first_match(block, [r"交易日期[:：]\s*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日|[0-9]{4}[-/.][0-9]{1,2}[-/.][0-9]{1,2})"])
        received_time = first_match(block, [r"交易时间[:：]\s*([0-9]{1,2}[:：][0-9]{2}(?::[0-9]{2})?)"])
        payer_name = first_match(block, [r"付款账户户名[:：]\s*([^\n]+)", r"付款人[:：]\s*([^\n]+)"])
        receiver_company = first_match(block, [r"收款账户户名[:：]\s*([^\n]+)", r"收款户名[:：]\s*([^\n]+)", r"收款人[:：]\s*([^\n]+)"])
        receiver_account = first_match(block, [r"收款账号[:：]\s*([0-9A-Za-z* ]+)", r"收款账户[:：]\s*([0-9A-Za-z* ]+)"])
        serial_no = first_match(block, [r"业务编号[:：]\s*([0-9A-Za-z]+)", r"回单编号[:：]\s*([0-9A-Za-z]+)", r"流水号[:：]\s*([0-9A-Za-z]+)"])
        amount = first_match(
            block,
            [
                r"交易金额[（(]小写[)）]?[:：]\s*(?:人民币|CNY|￥|¥)?\s*([0-9,]+(?:\.[0-9]{1,2})?)",
                r"交易金额[:：]\s*(?:人民币|CNY|￥|¥)?\s*([0-9,]+(?:\.[0-9]{1,2})?)",
                r"金额[:：]\s*(?:人民币|CNY|￥|¥)?\s*([0-9,]+(?:\.[0-9]{1,2})?)",
            ],
        )
        bank_note = first_match(block, [r"(?:附言|摘要|用途|交易摘要)[:：]\s*([^\n]+)"])

        if received_date and payer_name and parse_amount(amount) > 0:
            rows.append(
                {
                    "received_date": received_date,
                    "received_time": received_time,
                    "payer_name": payer_name,
                    "amount": amount,
                    "bank_note": bank_note,
                    "receiver_company": receiver_company,
                    "receiver_account": receiver_account,
                    "serial_no": serial_no,
                }
            )
    return rows


def tencent_sign(key: bytes, msg: str) -> bytes:
    return hmac.new(key, msg.encode("utf-8"), hashlib.sha256).digest()


def tencent_cloud_request(action: str, payload: dict[str, Any]) -> dict[str, Any]:
    if not (TENCENTCLOUD_SECRET_ID and TENCENTCLOUD_SECRET_KEY):
        raise HTTPException(status_code=500, detail="未配置腾讯云 OCR 密钥：TENCENTCLOUD_SECRET_ID / TENCENTCLOUD_SECRET_KEY")

    timestamp = int(time.time())
    date = datetime.utcfromtimestamp(timestamp).strftime("%Y-%m-%d")
    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    hashed_payload = hashlib.sha256(body.encode("utf-8")).hexdigest()
    canonical_headers = (
        "content-type:application/json; charset=utf-8\n"
        f"host:{TENCENT_OCR_HOST}\n"
        f"x-tc-action:{action.lower()}\n"
    )
    signed_headers = "content-type;host;x-tc-action"
    canonical_request = "\n".join(
        ["POST", "/", "", canonical_headers, signed_headers, hashed_payload]
    )
    credential_scope = f"{date}/{TENCENT_OCR_SERVICE}/tc3_request"
    string_to_sign = "\n".join(
        [
            "TC3-HMAC-SHA256",
            str(timestamp),
            credential_scope,
            hashlib.sha256(canonical_request.encode("utf-8")).hexdigest(),
        ]
    )
    secret_date = tencent_sign(("TC3" + TENCENTCLOUD_SECRET_KEY).encode("utf-8"), date)
    secret_service = tencent_sign(secret_date, TENCENT_OCR_SERVICE)
    secret_signing = tencent_sign(secret_service, "tc3_request")
    signature = hmac.new(secret_signing, string_to_sign.encode("utf-8"), hashlib.sha256).hexdigest()
    authorization = (
        "TC3-HMAC-SHA256 "
        f"Credential={TENCENTCLOUD_SECRET_ID}/{credential_scope}, "
        f"SignedHeaders={signed_headers}, Signature={signature}"
    )
    headers = {
        "Authorization": authorization,
        "Content-Type": "application/json; charset=utf-8",
        "Host": TENCENT_OCR_HOST,
        "X-TC-Action": action,
        "X-TC-Version": TENCENT_OCR_VERSION,
        "X-TC-Timestamp": str(timestamp),
        "X-TC-Region": TENCENTCLOUD_REGION,
    }
    req = urllib.request.Request(
        f"https://{TENCENT_OCR_HOST}",
        data=body.encode("utf-8"),
        method="POST",
        headers=headers,
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        try:
            data = json.loads(exc.read().decode("utf-8"))
        except Exception:
            raise HTTPException(status_code=502, detail=f"腾讯云 OCR 请求失败：HTTP {exc.code}") from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"腾讯云 OCR 网络请求失败：{exc}") from exc

    response = data.get("Response") or {}
    if response.get("Error"):
        error = response["Error"]
        raise HTTPException(
            status_code=502,
            detail=f"腾讯云 OCR 识别失败：{error.get('Code', '')} {error.get('Message', '')}",
        )
    return response


def bank_slip_value(info_map: dict[str, str], contains: list[str]) -> str:
    for key, value in info_map.items():
        if any(word in key for word in contains) and value:
            return value.strip()
    return ""


def bank_slip_infos_to_row(infos: list[dict[str, Any]]) -> dict[str, str]:
    info_map = {
        str(item.get("Name", "")).strip(): str(item.get("Value", "")).strip()
        for item in infos
        if str(item.get("Name", "")).strip()
    }
    received_date = bank_slip_value(info_map, ["交易日期", "入账日期", "日期"])
    received_time = bank_slip_value(info_map, ["交易时间", "时间"])
    payer_name = bank_slip_value(
        info_map,
        ["付款账户户名", "付款户名", "付款方户名", "付款人户名", "付款人", "付款方"],
    )
    amount = bank_slip_value(info_map, ["交易金额", "金额小写", "金额"])
    bank_note = bank_slip_value(info_map, ["附言", "摘要", "用途", "备注"])
    receiver_company = bank_slip_value(info_map, ["收款账户户名", "收款户名", "收款方户名", "收款人户名", "收款人", "收款方"])
    receiver_account = bank_slip_value(info_map, ["收款账号", "收款账户", "收款人账号"])
    serial_no = bank_slip_value(info_map, ["流水号", "回单编号", "业务编号", "凭证号码"])
    return {
        "received_date": received_date,
        "received_time": received_time,
        "payer_name": payer_name,
        "amount": amount,
        "bank_note": bank_note,
        "receiver_company": receiver_company,
        "receiver_account": receiver_account,
        "serial_no": serial_no,
    }


def ocr_pdf_with_tencent(path: Path) -> tuple[list[dict[str, str]], str]:
    raw = path.read_bytes()
    if len(base64.b64encode(raw)) > 7 * 1024 * 1024:
        return [], "PDF 超过腾讯云 OCR ImageBase64 7MB 限制，请拆分或压缩后再上传。"

    try:
        from pypdf import PdfReader
        page_count = len(PdfReader(str(path)).pages)
    except Exception:
        page_count = 1

    image_base64 = base64.b64encode(raw).decode("ascii")
    rows = []
    errors = []
    for page_no in range(1, max(page_count, 1) + 1):
        try:
            response = tencent_cloud_request(
                "BankSlipOCR",
                {"ImageBase64": image_base64, "IsPdf": True, "PdfPageNumber": page_no},
            )
        except HTTPException as exc:
            errors.append(str(exc.detail))
            continue
        infos = response.get("BankSlipInfos") or []
        row = bank_slip_infos_to_row(infos)
        if row.get("received_date") and row.get("payer_name") and parse_amount(row.get("amount")) > 0:
            rows.append(row)
        elif infos:
            errors.append(f"第 {page_no} 页已识别文字，但缺少日期/付款方/金额等关键字段")
        else:
            errors.append(f"第 {page_no} 页未识别到银行回单字段")

    if rows:
        return rows, ""
    return [], "腾讯云 OCR 未识别到完整银行回单流水。" + ("；".join(errors[:3]) if errors else "")


def rows_from_pdf(path: Path) -> tuple[list[dict[str, str]], str]:
    text = extract_pdf_text(path)
    if text:
        rows = parse_pdf_receipts(text)
        if rows:
            return rows, ""
        if OCR_PROVIDER != "tencent":
            return [], "PDF 文本已读取，但未识别到完整流水字段。请粘贴表格文本，至少包含到款日期、付款方名称、到款金额。"

    if OCR_PROVIDER == "tencent":
        return ocr_pdf_with_tencent(path)
    return [], "PDF 未识别到可复制文本，可能是图片型/扫描件。请配置腾讯云 OCR，或先 OCR 后粘贴表格文本。"



def duplicate_exists(conn: sqlite3.Connection, item: dict[str, str]) -> bool:
    serial_no = item.get("serial_no", "").strip()
    # POS 汇总流水可能同日、同额且摘要相同，无流水号时不自动判重。
    if not serial_no:
        return False
    row = conn.execute(
        "SELECT id FROM payments WHERE serial_no = ? AND status != 'closed' LIMIT 1",
        (serial_no,),
    ).fetchone()
    return row is not None


def status_badge(status: str) -> str:
    labels = {
        "draft": "待确认入池",
        "pending": "待认领",
        "partial_claiming": "部分认领中",
        "claimed": "已认领",
        "pending_confirm": "已认领",
        "rejected": "已驳回",
        "canceled": "已取消",
        "closed": "已关闭",
    }
    return f'<span class="status {esc(status)}">{esc(labels.get(status, status))}</span>'


def department_select(
    name: str,
    selected: str = "",
    required: bool = False,
    allow_blank: bool = True,
    class_name: str = "",
) -> str:
    required_attr = " required" if required else ""
    class_attr = f' class="{esc(class_name)}"' if class_name else ""
    selected = selected if selected in DEPARTMENTS else ""
    options = []
    if allow_blank:
        blank_selected = " selected" if not selected else ""
        options.append(f'<option value=""{blank_selected}>请选择部门</option>')
    for department in DEPARTMENTS:
        option_selected = " selected" if department == selected else ""
        options.append(f'<option value="{esc(department)}"{option_selected}>{esc(department)}</option>')
    return f'<select name="{esc(name)}"{class_attr}{required_attr}>{"".join(options)}</select>'


def team_select(name: str, department: str, selected: str = "", required: bool = False) -> str:
    teams = list(CATALOG.get(department, {}))
    selected = selected if selected in teams else ""
    placeholder = "请选择中心/小组" if teams else "请先选择部门"
    req = " required" if required else ""
    options = [f'<option value="">{placeholder}</option>'] + [
        f'<option value="{esc(t)}"{" selected" if t == selected else ""}>{esc(t)}</option>' for t in teams
    ]
    return f'<select name="{esc(name)}" class="cs-team"{req}>{"".join(options)}</select>'


def project_select(name: str, department: str, team: str, selected: str = "", required: bool = False) -> str:
    projects = CATALOG.get(department, {}).get(team, [])
    selected = selected if selected in projects else ""
    placeholder = "输入项目关键词" if projects else "请先选择中心/小组"
    req = " required" if required else ""
    list_id = f"project-options-{uuid.uuid4().hex}"
    options = "".join(f'<option value="{esc(p)}"></option>' for p in projects)
    return (
        f'<input name="{esc(name)}" class="cs-project" value="{esc(selected)}" '
        f'list="{list_id}" placeholder="{esc(placeholder)}" autocomplete="off"{req}>'
        f'<datalist id="{list_id}">{options}</datalist>'
    )


def require_department(department: str) -> str:
    department = department.strip()
    if department not in DEPARTMENTS:
        raise HTTPException(status_code=400, detail="请选择标准部门")
    return department


@app.get("/", response_class=HTMLResponse)
def home() -> RedirectResponse:
    return RedirectResponse("/search")


@app.get("/login")
def feishu_login(request: Request, next: str = "/search") -> RedirectResponse:
    rate_limit(request, "login")
    if not feishu_enabled():
        raise HTTPException(status_code=503, detail="尚未配置飞书登录")
    state = secrets.token_urlsafe(16)
    authorize = FEISHU_AUTHORIZE_URL + "?" + urlencode(
        {
            "app_id": FEISHU_APP_ID,
            "redirect_uri": FEISHU_REDIRECT_URI,
            "scope": FEISHU_SCOPE,
            "state": state,
        }
    )
    resp = RedirectResponse(authorize, status_code=302)
    resp.set_cookie("oauth_state", state, max_age=600, httponly=True, samesite="lax", secure=COOKIE_SECURE)
    resp.set_cookie("oauth_next", safe_next_url(next), max_age=600, httponly=True, samesite="lax", secure=COOKIE_SECURE)
    return resp


@app.get("/oauth/callback")
def feishu_callback(request: Request, code: str = "", state: str = "") -> RedirectResponse:
    if not feishu_enabled():
        raise HTTPException(status_code=503, detail="尚未配置飞书登录")
    saved_state = request.cookies.get("oauth_state", "")
    next_url = request.cookies.get("oauth_next", "/search")
    if not code or not state or not hmac.compare_digest(state, saved_state):
        raise HTTPException(status_code=400, detail="登录校验失败，请重新登录")
    token = feishu_exchange_token(code)
    if not token:
        raise HTTPException(status_code=400, detail="换取飞书令牌失败，请重试")
    info = feishu_get_userinfo(token)
    open_id = info.get("open_id")
    if not open_id:
        raise HTTPException(status_code=400, detail="获取飞书用户信息失败")
    name = info.get("name") or "飞书用户"
    # 会话只存身份；角色由 compute_role 每次请求实时判定（超管白名单 + 数据库管理员表）
    session = make_session({"id": open_id, "name": name, "src": "feishu"})
    resp = RedirectResponse(safe_next_url(next_url), status_code=302)
    resp.set_cookie(SESSION_COOKIE, session, max_age=SESSION_MAX_AGE, httponly=True, samesite="lax", secure=COOKIE_SECURE)
    resp.delete_cookie("oauth_state")
    resp.delete_cookie("oauth_next")
    with get_conn() as conn:
        # 记录所有登录过的人，供超管在后台勾选管理员
        conn.execute(
            """
            INSERT INTO app_users (open_id, name, created_at, last_login)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(open_id) DO UPDATE SET name = excluded.name, last_login = excluded.last_login
            """,
            (open_id, name, now_text(), now_text()),
        )
        audit(conn, {"id": open_id, "name": name, "role": compute_role(open_id)}, "feishu_login", None, {"name": name}, request)
    return resp


@app.get("/logout")
def logout() -> RedirectResponse:
    resp = RedirectResponse("/search", status_code=302)
    resp.delete_cookie(SESSION_COOKIE)
    return resp


@app.get("/attachments/{filename}")
def view_attachment(request: Request, filename: str) -> FileResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    if "/" in filename or "\\" in filename or filename != Path(filename).name:
        raise HTTPException(status_code=400, detail="附件路径不合法")
    path = UPLOAD_DIR / filename
    if not path.is_file():
        raise HTTPException(status_code=404, detail="附件不存在")
    return FileResponse(path, filename=filename)


CASCADE_JS = """
function fillSelect(sel, items, placeholder) {
  if (!sel) return;
  sel.innerHTML = '';
  var opt = document.createElement('option');
  opt.value = '';
  opt.textContent = placeholder;
  sel.appendChild(opt);
  for (var i = 0; i < items.length; i++) {
    var o = document.createElement('option');
    o.value = items[i];
    o.textContent = items[i];
    sel.appendChild(o);
  }
}
function fillProjectInput(input, items, placeholder) {
  if (!input) return;
  input.value = '';
  input.placeholder = placeholder;
  input.setCustomValidity('');
  var listId = input.getAttribute('list');
  var list = listId ? document.getElementById(listId) : null;
  if (!list) return;
  list.innerHTML = '';
  for (var i = 0; i < items.length; i++) {
    var o = document.createElement('option');
    o.value = items[i];
    list.appendChild(o);
  }
}
function projectInputHasKnownValue(input) {
  var value = (input.value || '').trim();
  if (!value) return !input.required;
  var listId = input.getAttribute('list');
  var list = listId ? document.getElementById(listId) : null;
  if (!list) return true;
  return Array.prototype.some.call(list.options, function (opt) {
    return opt.value === value;
  });
}
function validateProjectInput(input) {
  if (!input || !input.classList.contains('cs-project')) return true;
  if (projectInputHasKnownValue(input)) {
    input.setCustomValidity('');
    return true;
  }
  input.setCustomValidity('请选择已有项目');
  return false;
}
document.addEventListener('change', function (e) {
  var scope = e.target.closest('.cascade-scope') || e.target.closest('form');
  if (!scope) return;
  if (e.target.classList.contains('cs-dept')) {
    fillSelect(scope.querySelector('.cs-team'), Object.keys(CATALOG[e.target.value] || {}), '请选择中心/小组');
    fillProjectInput(scope.querySelector('.cs-project'), [], '请先选择中心/小组');
  } else if (e.target.classList.contains('cs-team')) {
    var deptInput = scope.querySelector('.cs-dept');
    var dept = deptInput ? deptInput.value : '';
    fillProjectInput(scope.querySelector('.cs-project'), (CATALOG[dept] || {})[e.target.value] || [], '输入项目关键词');
  }
});
document.addEventListener('input', function (e) {
  if (e.target.classList.contains('cs-project')) {
    validateProjectInput(e.target);
  }
});
document.addEventListener('submit', function (e) {
  var inputs = Array.prototype.slice.call(e.target.querySelectorAll('.cs-project'));
  var invalid = inputs.find(function (input) { return !validateProjectInput(input); });
  if (invalid) {
    invalid.reportValidity();
    e.preventDefault();
  }
});
function createCascadeSelect(name, className, placeholder, items) {
  var select = document.createElement('select');
  select.name = name;
  select.className = className;
  var blank = document.createElement('option');
  blank.value = '';
  blank.textContent = placeholder;
  select.appendChild(blank);
  for (var i = 0; i < items.length; i++) {
    var option = document.createElement('option');
    option.value = items[i];
    option.textContent = items[i];
    select.appendChild(option);
  }
  return select;
}
function createProjectInput() {
  var id = 'project-options-added-' + Date.now() + '-' + Math.random().toString(16).slice(2);
  var input = document.createElement('input');
  input.name = 'projects';
  input.className = 'cs-project';
  input.setAttribute('list', id);
  input.placeholder = '请先选择中心/小组';
  input.autocomplete = 'off';
  var list = document.createElement('datalist');
  list.id = id;
  return { input: input, list: list };
}
function refreshSplitRowNumbers(tbody) {
  Array.prototype.forEach.call(tbody.querySelectorAll('.split-row-number'), function (cell, index) {
    cell.textContent = index + 1;
  });
}
function addSplitClaimRow(button) {
  var form = button.closest('form');
  var tbody = form ? form.querySelector('.split-form-table tbody') : null;
  if (!tbody) return;
  var row = document.createElement('tr');
  row.className = 'cascade-scope';

  var numberCell = document.createElement('td');
  numberCell.className = 'split-row-number';
  row.appendChild(numberCell);

  var deptCell = document.createElement('td');
  deptCell.appendChild(createCascadeSelect('departments', 'cs-dept', '请选择部门', Object.keys(CATALOG || {})));
  row.appendChild(deptCell);

  var teamCell = document.createElement('td');
  teamCell.appendChild(createCascadeSelect('teams', 'cs-team', '请先选择部门', []));
  row.appendChild(teamCell);

  var projectCell = document.createElement('td');
  var project = createProjectInput();
  projectCell.appendChild(project.input);
  projectCell.appendChild(project.list);
  row.appendChild(projectCell);

  var amountCell = document.createElement('td');
  var amountInput = document.createElement('input');
  amountInput.name = 'amounts';
  amountInput.placeholder = '0.00 / 退款填负数';
  amountCell.appendChild(amountInput);
  row.appendChild(amountCell);

  var noteCell = document.createElement('td');
  var noteInput = document.createElement('input');
  noteInput.name = 'notes';
  noteInput.placeholder = '如：2本杂志 / 3个笔记本 / 2个参会名额';
  noteCell.appendChild(noteInput);
  row.appendChild(noteCell);

  tbody.appendChild(row);
  refreshSplitRowNumbers(tbody);
}
document.addEventListener('click', function (e) {
  var button = e.target.closest('.split-add-row');
  if (!button) return;
  addSplitClaimRow(button);
});
"""

BULK_ADMIN_JS = """
function updateBulkCloseState() {
  var boxes = visibleBulkPaymentBoxes();
  var checked = boxes.filter(function (box) { return box.checked; });
  var count = document.getElementById('bulk-selected-count');
  var button = document.getElementById('bulk-close-button');
  var all = document.getElementById('bulk-select-all');
  if (count) count.textContent = checked.length;
  if (button) button.disabled = checked.length === 0;
  if (all) {
    all.checked = boxes.length > 0 && checked.length === boxes.length;
    all.indeterminate = checked.length > 0 && checked.length < boxes.length;
  }
}
function visibleBulkPaymentBoxes() {
  return Array.prototype.slice.call(document.querySelectorAll('.bulk-payment-checkbox:not(:disabled)')).filter(function (box) {
    var row = box.closest('tr');
    return !row || row.style.display !== 'none';
  });
}
document.addEventListener('change', function (e) {
  if (e.target.id === 'bulk-select-all') {
    var boxes = visibleBulkPaymentBoxes();
    boxes.forEach(function (box) { box.checked = e.target.checked; });
    updateBulkCloseState();
  } else if (e.target.classList.contains('bulk-payment-checkbox')) {
    updateBulkCloseState();
  }
});
document.addEventListener('click', function (e) {
  var link = e.target.closest('.sort-link');
  if (!link || !link.dataset.tableUrl) return;
  e.preventDefault();
  var card = document.getElementById('payment-pool-card');
  if (!card) {
    window.location.href = link.href;
    return;
  }
  card.setAttribute('aria-busy', 'true');
  fetch(link.dataset.tableUrl, {headers: {'X-Requested-With': 'fetch'}})
    .then(function (response) {
      if (!response.ok) throw new Error('table refresh failed');
      return response.text();
    })
    .then(function (html) {
      card.outerHTML = html;
      history.replaceState(null, '', link.href);
      initProgressiveTables();
    })
    .catch(function () {
      window.location.href = link.href;
    });
});
function confirmBulkClose() {
  var count = document.querySelectorAll('.bulk-payment-checkbox:checked').length;
  if (count === 0) return false;
  return confirm('确定关闭选中的 ' + count + ' 条流水吗？关闭后 7 天内仍会显示，之后默认隐藏。');
}
"""

PROGRESSIVE_TABLE_JS = """
function applyProgressiveTable(button) {
  var group = button.dataset.progressiveGroup;
  var rows = Array.prototype.slice.call(document.querySelectorAll('[data-progressive-group="' + group + '"]'));
  var visibleCount = parseInt(button.dataset.visibleCount || '0', 10);
  rows.forEach(function (row, index) {
    row.style.display = index < visibleCount ? '' : 'none';
  });
  var remaining = Math.max(rows.length - visibleCount, 0);
  button.style.display = remaining > 0 ? '' : 'none';
  button.textContent = remaining > 0 ? '显示更多（剩余 ' + remaining + ' 条）' : '已全部显示';
}
function initProgressiveTables() {
  document.querySelectorAll('[data-progressive-more]').forEach(function (button) {
    if (!button.dataset.visibleCount) {
      button.dataset.visibleCount = button.dataset.initialCount || button.dataset.progressiveStep || '10';
    }
    applyProgressiveTable(button);
  });
  if (typeof updateBulkCloseState === 'function') updateBulkCloseState();
}
document.addEventListener('DOMContentLoaded', initProgressiveTables);
document.addEventListener('click', function (e) {
  var button = e.target.closest('[data-progressive-more]');
  if (!button) return;
  var step = parseInt(button.dataset.progressiveStep || '10', 10);
  var visibleCount = parseInt(button.dataset.visibleCount || button.dataset.initialCount || step, 10);
  button.dataset.visibleCount = visibleCount + step;
  applyProgressiveTable(button);
  if (typeof updateBulkCloseState === 'function') updateBulkCloseState();
});
"""

COPY_TODAY_TEXT_JS = """
function copyTextWithFallback(text, status) {
  function copied() {
    if (status) status.textContent = '已复制纯文本。';
  }
  function fallback() {
    var area = document.createElement('textarea');
    area.value = text;
    area.style.position = 'fixed';
    area.style.left = '-9999px';
    document.body.appendChild(area);
    area.focus();
    area.select();
    try {
      document.execCommand('copy');
      copied();
    } catch (err) {
      if (status) status.textContent = '复制失败，请手动下载 CSV。';
    }
    document.body.removeChild(area);
  }
  if (navigator.clipboard && navigator.clipboard.writeText) {
    navigator.clipboard.writeText(text).then(copied).catch(fallback);
  } else {
    fallback();
  }
}
document.addEventListener('DOMContentLoaded', function () {
  var button = document.getElementById('copy-today-plain-text');
  var status = document.getElementById('copy-today-plain-text-status');
  if (!button) return;
  function exportRangeQuery() {
    var startInput = document.getElementById('export-start-date');
    var endInput = document.getElementById('export-end-date');
    var startDate = startInput && startInput.value ? startInput.value : '';
    var endDate = endInput && endInput.value ? endInput.value : startDate;
    if (startDate && !endDate) endDate = startDate;
    if (!startDate && endDate) startDate = endDate;
    var params = [];
    if (startDate) params.push('start_date=' + encodeURIComponent(startDate));
    if (endDate) params.push('end_date=' + encodeURIComponent(endDate));
    return params.length ? '?' + params.join('&') : '';
  }
  button.addEventListener('click', function () {
    var url = '/admin/export/today-text' + exportRangeQuery();
    button.disabled = true;
    if (status) status.textContent = '正在整理...';
    fetch(url, {headers: {'X-Requested-With': 'fetch'}})
      .then(function (response) {
        if (!response.ok) throw new Error('copy text export failed');
        return response.text();
      })
      .then(function (text) {
        copyTextWithFallback(text, status);
      })
      .catch(function () {
        if (status) status.textContent = '整理失败，请刷新后重试。';
      })
      .finally(function () {
        button.disabled = false;
      });
  });
  var csvLink = document.getElementById('export-date-csv');
  var startInput = document.getElementById('export-start-date');
  var endInput = document.getElementById('export-end-date');
  function syncExportLinks() {
    if (!csvLink) return;
    csvLink.href = '/admin/export/today' + exportRangeQuery();
  }
  if (startInput) startInput.addEventListener('change', syncExportLinks);
  if (endInput) endInput.addEventListener('change', syncExportLinks);
  syncExportLinks();
});
"""

BATCH_CLAIM_JS = """
function updateBatchClaimState() {
  var boxes = Array.prototype.slice.call(document.querySelectorAll('.batch-claim-checkbox:not(:disabled)'));
  var checked = boxes.filter(function (box) { return box.checked; });
  var checkedIds = {};
  checked.forEach(function (box) { checkedIds[box.value] = true; });
  var uniqueCount = Object.keys(checkedIds).length;
  var count = document.getElementById('batch-claim-selected-count');
  var button = document.getElementById('batch-claim-button');
  var all = document.getElementById('batch-claim-select-all');
  if (count) count.textContent = uniqueCount;
  if (button) button.disabled = uniqueCount === 0;
  if (all) {
    all.checked = boxes.length > 0 && checked.length === boxes.length;
    all.indeterminate = checked.length > 0 && checked.length < boxes.length;
  }
}
document.addEventListener('change', function (e) {
  if (e.target.id === 'batch-claim-select-all') {
    var boxes = document.querySelectorAll('.batch-claim-checkbox:not(:disabled)');
    boxes.forEach(function (box) { box.checked = e.target.checked; });
    updateBatchClaimState();
  } else if (e.target.classList.contains('batch-claim-checkbox')) {
    updateBatchClaimState();
  }
});
document.addEventListener('DOMContentLoaded', updateBatchClaimState);
function confirmBatchClaim() {
  var checked = Array.prototype.slice.call(document.querySelectorAll('.batch-claim-checkbox:checked'));
  var checkedIds = {};
  checked.forEach(function (box) { checkedIds[box.value] = true; });
  var count = Object.keys(checkedIds).length;
  if (count === 0) return false;
  return confirm('确定批量认领选中的 ' + count + ' 笔到款吗？系统会按每笔剩余可认领金额分别生成认领记录。');
}
"""


DIAGNOSTIC_LOG_JS = """
(function () {
  var recentErrors = [];
  var recentActions = [];
  function rememberAction(text) {
    if (!text) return;
    recentActions.push(new Date().toISOString() + ' ' + text.trim().slice(0, 80));
    if (recentActions.length > 8) recentActions.shift();
  }
  function pushError(text) {
    recentErrors.push(new Date().toISOString() + ' ' + text);
    if (recentErrors.length > 5) recentErrors.shift();
  }
  document.addEventListener('click', function (e) {
    var target = e.target.closest('button,a,summary,input,select');
    if (!target) return;
    var label = target.innerText || target.value || target.getAttribute('aria-label') || target.name || target.tagName;
    rememberAction(target.tagName.toLowerCase() + ': ' + label);
  }, true);
  window.addEventListener('error', function (e) {
    pushError((e.message || '脚本错误') + ' @ ' + (e.filename || 'unknown') + ':' + (e.lineno || 0) + ':' + (e.colno || 0));
  });
  window.addEventListener('unhandledrejection', function (e) {
    var reason = e.reason && (e.reason.message || String(e.reason));
    pushError('未处理 Promise: ' + (reason || 'unknown'));
  });
  function browserInfo() {
    return [
      '页面地址: ' + window.location.href,
      '浏览器: ' + navigator.userAgent,
      '语言: ' + navigator.language,
      '屏幕: ' + window.screen.width + 'x' + window.screen.height,
      '视口: ' + window.innerWidth + 'x' + window.innerHeight,
      '页面标题: ' + document.title,
      '滚动位置: ' + Math.round(window.scrollY || 0),
      '时区: ' + (Intl.DateTimeFormat().resolvedOptions().timeZone || 'unknown'),
      '本机时间: ' + new Date().toISOString(),
      '最近操作: ' + (recentActions.length ? recentActions.join('\\n') : '无'),
      '最近前端错误: ' + (recentErrors.length ? recentErrors.join('\\n') : '无')
    ].join('\\n');
  }
  function buildDiagnosticLog() {
    var base = document.getElementById('diagnostic-log-base');
    return (base ? base.value : '【问题排查日志】') + '\\n\\n--- 浏览器信息 ---\\n' + browserInfo();
  }
  document.addEventListener('DOMContentLoaded', function () {
    var button = document.getElementById('copy-diagnostic-log');
    var status = document.getElementById('copy-diagnostic-status');
    if (!button) return;
    button.addEventListener('click', function () {
      var logText = buildDiagnosticLog();
      function copied() {
        if (status) status.textContent = '已复制，可以发给管理员排查。';
      }
      function fallback() {
        var area = document.createElement('textarea');
        area.value = logText;
        area.style.position = 'fixed';
        area.style.left = '-9999px';
        document.body.appendChild(area);
        area.focus();
        area.select();
        try {
          document.execCommand('copy');
          copied();
        } catch (err) {
          if (status) status.textContent = '复制失败，请刷新后重试。';
        }
        document.body.removeChild(area);
      }
      if (navigator.clipboard && navigator.clipboard.writeText) {
        navigator.clipboard.writeText(logText).then(copied).catch(fallback);
      } else {
        fallback();
      }
    });
  });
})();
"""


def catalog_script() -> str:
    catalog_json = (
        json.dumps(CATALOG, ensure_ascii=False)
        .replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
    )
    return (
        "<script>var CATALOG = "
        + catalog_json
        + ";"
        + CASCADE_JS
        + "</script>"
    )


def admin_script() -> str:
    return catalog_script() + "<script>" + BULK_ADMIN_JS + PROGRESSIVE_TABLE_JS + COPY_TODAY_TEXT_JS + "</script>"


def search_script() -> str:
    return catalog_script() + "<script>" + BATCH_CLAIM_JS + "</script>"


def personal_script() -> str:
    return catalog_script() + "<script>" + PROGRESSIVE_TABLE_JS + DIAGNOSTIC_LOG_JS + "</script>"


def claim_totals(conn: sqlite3.Connection, payment_id: int) -> dict[str, int]:
    row = conn.execute(
        """
        SELECT
            COALESCE(SUM(CASE WHEN status IN ('pending', 'accepted') THEN amount_cents ELSE 0 END), 0) AS active,
            COALESCE(SUM(CASE WHEN status = 'accepted' THEN amount_cents ELSE 0 END), 0) AS accepted,
            COALESCE(SUM(CASE WHEN status = 'pending' THEN amount_cents ELSE 0 END), 0) AS pending
        FROM claims
        WHERE payment_id = ?
        """,
        (payment_id,),
    ).fetchone()
    return {"active": row["active"], "accepted": row["accepted"], "pending": row["pending"]}


def claim_summary_html(conn: sqlite3.Connection, row: sqlite3.Row, show_amount: bool = True) -> str:
    totals = claim_totals(conn, row["id"])
    if totals["active"] <= 0:
        return ""
    remaining = max((row["amount_cents"] or 0) - totals["active"], 0)
    if not show_amount:
        return '<div class="muted">已有部分认领</div>'
    return (
        f'<div class="muted">已认领 ¥ {money(totals["active"])}'
        f' · 剩余 ¥ {money(remaining)}</div>'
    )


def batch_claim_checkbox_html(conn: sqlite3.Connection, row: sqlite3.Row) -> str:
    if row["status"] not in {"pending", "partial_claiming"}:
        return ""
    totals = claim_totals(conn, row["id"])
    remaining = max((row["amount_cents"] or 0) - totals["active"], 0)
    if remaining <= 0:
        return ""
    return (
        f'<input class="batch-claim-checkbox" form="batch-claim-form" type="checkbox" '
        f'name="payment_ids" value="{row["id"]}" aria-label="选择到款 #{row["id"]}">'
    )


def batch_claim_panel_html(actor: dict[str, str]) -> str:
    my_dept = actor.get("department") if actor.get("department") in DEPARTMENTS else ""
    my_team = actor.get("team", "")
    return f"""
    <div class="panel cascade-scope">
      <form id="batch-claim-form" method="post" action="/claim/batch" onsubmit="return confirmBatchClaim()">
        <div class="row">
          <div>
            <label>选择</label>
            <label class="muted" style="display:flex; align-items:center; gap:8px; margin:9px 0 0; white-space:nowrap">
              <input id="batch-claim-select-all" type="checkbox" style="width:auto; margin:0; box-shadow:none">
              全选当前页
            </label>
          </div>
          <div style="flex:1; min-width:180px">
            <label>批量认领部门</label>
            {department_select("department", my_dept, required=True, class_name="cs-dept")}
          </div>
          <div style="flex:1; min-width:180px">
            <label>中心 / 小组</label>
            {team_select("team", my_dept, my_team, required=True)}
          </div>
          <div style="flex:1.2; min-width:220px">
            <label>项目</label>
            {project_select("customer_project", my_dept, my_team, required=True)}
          </div>
          <div style="flex:1; min-width:180px">
            <label>统一备注</label>
            <input name="note" placeholder="可选，会写入每笔认领">
          </div>
          <div>
            <button id="batch-claim-button" type="submit" disabled>批量认领选中款项</button>
          </div>
        </div>
        <p class="hint">已选 <strong id="batch-claim-selected-count">0</strong> 笔。默认按每笔剩余可认领金额提交；不同归属或拆分金额请用“分摊认领”。</p>
      </form>
    </div>
    """


def claim_form_html(row: sqlite3.Row, actor: dict[str, str]) -> str:
    my_dept = actor.get("department") if actor.get("department") in DEPARTMENTS else ""
    my_team = actor.get("team", "")
    dept_options = '<option value="">请选择部门</option>' + "".join(
        f'<option value="{esc(d)}"{" selected" if d == my_dept else ""}>{esc(d)}</option>' for d in DEPARTMENTS
    )
    return f"""
    <details class="fold">
      <summary>认领这笔款</summary>
      <div class="fold-body">
        <form method="post" action="/claim/{row['id']}">
          <input type="hidden" name="user" value="{esc(actor['id'])}">
          <input type="hidden" name="name" value="{esc(actor['name'])}">
          <input type="hidden" name="role" value="{esc(actor['role'])}">
          <div class="field"><label>认领部门</label><select name="department" class="cs-dept" required>{dept_options}</select></div>
          <div class="field"><label>中心 / 小组</label>{team_select("team", my_dept, my_team, required=True)}</div>
          <div class="field"><label>项目</label>{project_select("customer_project", my_dept, my_team, required=True)}</div>
          <div class="field"><label>认领金额</label><input name="claim_amount" placeholder="留空表示整笔/部分认领请填写金额"></div>
          <div class="field"><label>备注说明</label><input name="note" placeholder="可选"></div>
          <button type="submit">提交认领</button>
        </form>
      </div>
    </details>
    """


def submit_batch_claims(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    payment_ids: list[int],
    department: str,
    team: str,
    customer_project: str,
    note: str = "",
    request: Optional[Request] = None,
) -> dict[str, Any]:
    department = require_department(department)
    team = team.strip()
    customer_project = customer_project.strip()
    note = note.strip()
    if team not in CATALOG.get(department, {}):
        raise HTTPException(status_code=400, detail="请选择该部门下的中心/小组")
    if customer_project not in CATALOG[department][team]:
        raise HTTPException(status_code=400, detail="请选择该中心/小组下的项目")

    unique_ids = list(dict.fromkeys(int(payment_id) for payment_id in payment_ids if int(payment_id) > 0))
    if not unique_ids:
        raise HTTPException(status_code=400, detail="请先勾选需要批量认领的到款")

    created_claim_ids: list[int] = []
    claimed_payment_ids: list[int] = []
    skipped: list[dict[str, Any]] = []
    total_amount_cents = 0
    for payment_id in unique_ids:
        row = conn.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
        if not row:
            skipped.append({"payment_id": payment_id, "reason": "not_found"})
            continue
        if row["status"] not in {"pending", "partial_claiming"}:
            skipped.append({"payment_id": payment_id, "reason": "status", "status": row["status"]})
            continue
        totals = claim_totals(conn, payment_id)
        remaining_amount = max(row["amount_cents"] - totals["active"], 0)
        if remaining_amount <= 0:
            skipped.append({"payment_id": payment_id, "reason": "no_remaining"})
            continue

        claim_created_at = now_text()
        cur = conn.execute(
            """
            INSERT INTO claims
                (payment_id, department, team, amount_cents, actor_id, actor_name, customer_project, contract_invoice, note, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, '', ?, 'accepted', ?)
            """,
            (
                payment_id,
                department,
                team,
                remaining_amount,
                actor["id"],
                actor["name"],
                customer_project,
                note,
                claim_created_at,
            ),
        )
        created_claim_ids.append(cur.lastrowid)
        claimed_payment_ids.append(payment_id)
        total_amount_cents += remaining_amount
        if row["status"] == "pending" and totals["active"] == 0 and remaining_amount == row["amount_cents"]:
            conn.execute(
                """
                UPDATE payments
                SET claimed_department = ?,
                    claimed_team = ?,
                    claimed_by = ?,
                    claimed_by_name = ?,
                    claimed_at = ?,
                    customer_project = ?,
                    claim_note = ?
                WHERE id = ?
                """,
                (
                    department,
                    team,
                    actor["id"],
                    actor["name"],
                    claim_created_at,
                    customer_project,
                    note,
                    payment_id,
                ),
            )
        refresh_payment_claim_status(conn, payment_id)

    if not created_claim_ids:
        raise HTTPException(status_code=409, detail="选中的到款都无法批量认领，请刷新后重试")

    detail = {
        "payment_ids": claimed_payment_ids,
        "claim_ids": created_claim_ids,
        "department": department,
        "team": team,
        "customer_project": customer_project,
        "count": len(created_claim_ids),
        "amount_cents": total_amount_cents,
        "skipped": skipped,
    }
    audit(conn, actor, "batch_claim_submit", None, detail, request)
    return detail


def split_claim_line_rows(count: int = 6) -> str:
    rows = []
    for index in range(count):
        rows.append(
            f"""
            <tr class="cascade-scope">
              <td class="split-row-number">{index + 1}</td>
              <td>{department_select("departments", required=False, class_name="cs-dept")}</td>
              <td>{team_select("teams", "", required=False)}</td>
              <td>{project_select("projects", "", "", required=False)}</td>
              <td><input name="amounts" placeholder="0.00 / 退款填负数"></td>
              <td><input name="notes" placeholder="如：2本杂志 / 3个笔记本 / 2个参会名额"></td>
            </tr>
            """
        )
    return "".join(rows)


def split_claim_form_html(row: sqlite3.Row) -> str:
    return f"""
    <details class="fold">
      <summary>填写分摊明细</summary>
      <div class="fold-body">
        <form method="post" action="/split-claim/{row['id']}">
          <div class="grid" style="grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); margin-bottom:14px">
            <div class="stat" style="--dot:#2456d6"><span class="stat-label">到款金额</span><strong>¥ {money(row["amount_cents"])}</strong></div>
            <div class="stat" style="--dot:#16a34a"><span class="stat-label">当前状态</span><strong style="font-size:18px">{status_badge(row["status"])}</strong></div>
          </div>
          <div class="table-wrap split-form-table" style="box-shadow:none; margin-bottom:12px">
            <table>
              <thead><tr><th>#</th><th>部门</th><th>中心 / 小组</th><th>项目</th><th>分摊金额</th><th>备注</th></tr></thead>
              <tbody>{split_claim_line_rows()}</tbody>
            </table>
          </div>
          <button type="button" class="secondary split-add-row" style="margin:0 0 12px">＋ 添加分摊行</button>
          <p class="hint" style="margin:0 0 12px">只填写需要分摊的行，退款项目请填负数；分摊净额合计必须大于 0，且不能超过该笔款剩余可认领金额。</p>
          <button type="submit">提交分摊认领</button>
        </form>
      </div>
    </details>
    """


def split_claim_candidates(conn: sqlite3.Connection, q: str) -> list[sqlite3.Row]:
    compact = q.strip()
    if not compact:
        return []
    id_match = re.fullmatch(r"#?\s*(\d+)", compact)
    if id_match:
        row = conn.execute(
            """
            SELECT *
            FROM payments
            WHERE id = ?
              AND status IN ('pending', 'partial_claiming', 'claimed', 'pending_confirm')
            """,
            (int(id_match.group(1)),),
        ).fetchone()
        return [row] if row else []
    return run_search(conn, q)


def submit_split_claims(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    payment_id: int,
    departments: list[str],
    teams: list[str],
    projects: list[str],
    amounts: list[str],
    notes: list[str],
    request: Optional[Request] = None,
) -> dict[str, Any]:
    payment = conn.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
    if not payment:
        raise HTTPException(status_code=404, detail="到款记录不存在")
    if payment["status"] not in {"pending", "partial_claiming", "claimed", "pending_confirm"}:
        raise HTTPException(status_code=409, detail="这笔款当前状态不能分摊认领")

    line_count = max(len(departments), len(teams), len(projects), len(amounts), len(notes))
    lines: list[dict[str, Any]] = []
    for index in range(line_count):
        department = departments[index].strip() if index < len(departments) else ""
        team = teams[index].strip() if index < len(teams) else ""
        project = projects[index].strip() if index < len(projects) else ""
        amount_text = amounts[index].strip() if index < len(amounts) else ""
        note = notes[index].strip() if index < len(notes) else ""
        if not any([department, team, project, amount_text, note]):
            continue
        if not department or not team or not project or not amount_text:
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 行分摊明细不完整")
        department = require_department(department)
        if team not in CATALOG.get(department, {}):
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 行中心/小组不属于所选部门")
        if project not in CATALOG[department][team]:
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 行项目不属于所选中心/小组")
        amount_cents = parse_amount(amount_text)
        if amount_cents == 0:
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 行分摊金额不能为 0，退款请填负数")
        lines.append(
            {
                "department": department,
                "team": team,
                "project": project,
                "amount_cents": amount_cents,
                "note": note,
            }
        )

    if not lines:
        raise HTTPException(status_code=400, detail="请至少填写一条分摊明细")
    totals = claim_totals(conn, payment_id)
    new_total = sum(line["amount_cents"] for line in lines)
    if new_total <= 0:
        raise HTTPException(status_code=400, detail="分摊金额合计必须大于 0")
    if totals["active"] + new_total > payment["amount_cents"]:
        raise HTTPException(status_code=409, detail="分摊金额合计超过该笔款剩余可认领金额")

    created_claim_ids = []
    for line in lines:
        cur = conn.execute(
            """
            INSERT INTO claims
                (payment_id, department, team, amount_cents, actor_id, actor_name, customer_project, contract_invoice, note, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, '', ?, 'accepted', ?)
            """,
            (
                payment_id,
                line["department"],
                line["team"],
                line["amount_cents"],
                actor["id"],
                actor["name"],
                line["project"],
                line["note"],
                now_text(),
            ),
        )
        created_claim_ids.append(cur.lastrowid)

    refresh_payment_claim_status(conn, payment_id)
    payment_after = conn.execute("SELECT status FROM payments WHERE id = ?", (payment_id,)).fetchone()
    detail = {
        "payment_id": payment_id,
        "claim_ids": created_claim_ids,
        "count": len(lines),
        "amount_cents": new_total,
        "payment_status": payment_after["status"] if payment_after else "",
    }
    audit(conn, actor, "split_claim_submit", payment_id, detail, request)
    return detail


@app.get("/search", response_class=HTMLResponse)
def search_page(request: Request, q: str = "", pending_date: str = "") -> HTMLResponse:
    rate_limit(request, "search")
    actor = actor_from_request(request)
    if not actor.get("authed"):
        if feishu_enabled():
            login_html = '<a href="/login?next=/search" class="login-btn">飞书登录</a>'
            prompt = "请先通过飞书登录。登录后，组织内同事才能搜索到款并提交认领。"
        else:
            login_html = ""
            prompt = "系统尚未配置飞书登录，请联系管理员配置后再使用认领搜索。"
        body = f"""
        <div class="panel">
          <div class="callout info" style="margin-bottom:0">
            <strong>认领搜索仅对组织内同事开放</strong><br>
            {esc(prompt)}
          </div>
          {f'<div style="margin-top:14px">{login_html}</div>' if login_html else ''}
        </div>
        """
        return page("到款认领搜索", body, active="search", subtitle="登录后可搜索到款并提交认领", actor=actor)
    q = q.strip()
    pending_date = parse_date(pending_date) if pending_date.strip() else ""
    message = ""
    results: list[sqlite3.Row] = []

    if q:
        ok, reason = validate_search_query(q)
        with get_conn() as conn:
            if not ok:
                message = reason
                audit(conn, actor, "search_blocked", None, {"query": q, "reason": reason}, request)
            else:
                results = run_search(conn, q)
                audit(conn, actor, "search", None, {"query": q, "result_count": len(results)}, request)

    result_html = ""
    if message:
        result_html = f'<div class="callout warn"><strong>搜索被限制：</strong>{esc(message)}</div>'
    elif q and not results:
        result_html = '<div class="callout info">没有找到匹配记录。请换一个更具体的客户名、金额或备注关键词。</div>'
    elif results:
        rows = []
        with get_conn() as conn:
            for row in results:
                claim_hint = ""
                if row["status"] in {"claimed", "pending_confirm"}:
                    claim_hint = f'<div class="muted">{esc(row["claimed_department"] or "")} · {esc(row["claimed_by_name"] or "")}</div>'
                elif row["status"] == "partial_claiming":
                    claim_hint = claim_summary_html(conn, row)
                rows.append(
                    f"""
                    <tr>
                      <td class="select-cell">{batch_claim_checkbox_html(conn, row)}</td>
                      <td class="nowrap">{esc(row["received_date"])}</td>
                      <td class="num">¥ {money(row["amount_cents"])}</td>
                      <td>{esc(receiver_company_label(row["receiver_company"]))}</td>
                      <td><strong>{esc(row["payer_name"])}</strong></td>
                      <td>{esc(row["bank_note"])}</td>
                      <td>{status_badge(row["status"])}{claim_hint}</td>
                      <td class="actions">{claim_form_html(row, actor)}</td>
                    </tr>
                    """
                )
        result_html = f"""
        <div class="table-wrap">
        <table>
          <thead><tr><th class="select-cell"></th><th>日期</th><th>金额</th><th>到款公司</th><th>付款方</th><th>银行备注</th><th>状态</th><th style="width:300px">认领</th></tr></thead>
          <tbody>{''.join(rows)}</tbody>
        </table>
        </div>
        """

    with get_conn() as conn:
        if pending_date:
            pending_rows = conn.execute(
                """
                SELECT * FROM payments
                WHERE status IN ('pending', 'partial_claiming') AND received_date = ?
                ORDER BY id DESC
                LIMIT 50
                """,
                (pending_date,),
            ).fetchall()
            audit(conn, actor, "pending_filter", None, {"date": pending_date, "result_count": len(pending_rows)}, request)
        else:
            pending_rows = conn.execute(
                """
                SELECT * FROM payments
                WHERE status IN ('pending', 'partial_claiming')
                ORDER BY received_date DESC, id DESC
                LIMIT 50
                """
            ).fetchall()

    clear_href = url(
        "/search", q=q, user=actor["id"], name=actor["name"],
        department=actor["department"], role=actor["role"],
    )
    filter_form = f"""
    <form method="get" action="/search" class="row" style="margin:0 0 12px">
      <input type="hidden" name="q" value="{esc(q)}">
      <input type="hidden" name="user" value="{esc(actor['id'])}">
      <input type="hidden" name="name" value="{esc(actor['name'])}">
      <input type="hidden" name="department" value="{esc(actor['department'])}">
      <input type="hidden" name="role" value="{esc(actor['role'])}">
      <div><label>按到款日期筛选</label><input type="date" name="pending_date" value="{esc(pending_date)}"></div>
      <div><button type="submit" class="secondary">筛选</button></div>
      {f'<div><a href="{esc(clear_href)}" style="display:inline-block;padding:9px 0">清除筛选</a></div>' if pending_date else ''}
    </form>
    """

    if pending_rows:
        with get_conn() as conn:
            rows = [
                f"""
                <tr>
                  <td class="select-cell">{batch_claim_checkbox_html(conn, row)}</td>
                  <td class="nowrap">#{esc(row["id"])}</td>
                  <td class="nowrap">{esc(row["received_date"])}</td>
                  <td class="num">¥ {money(row["amount_cents"])}</td>
                  <td>{esc(receiver_company_label(row["receiver_company"]))}</td>
                  <td><strong>{esc(row["payer_name"])}</strong></td>
                  <td>{esc(row["bank_note"])}</td>
                  <td>{status_badge(row["status"])}{claim_summary_html(conn, row)}</td>
                  <td class="actions">{claim_form_html(row, actor)}</td>
                </tr>
                """
                for row in pending_rows
            ]
        table_html = f"""
        <div class="table-wrap">
        <table>
          <thead><tr><th class="select-cell"></th><th>ID</th><th>日期</th><th>金额</th><th>到款公司</th><th>付款方</th><th>银行备注</th><th>状态</th><th style="width:300px">认领</th></tr></thead>
          <tbody>{''.join(rows)}</tbody>
        </table>
        </div>
        """
    else:
        table_html = (
            f'<div class="callout info">{esc(pending_date)} 没有待认领的款项，换个日期或清除筛选试试。</div>'
            if pending_date
            else '<div class="callout info">当前没有待认领的款项。</div>'
        )

    pending_html = f"""
    <h2>待认领列表</h2>
    {filter_form}
    {table_html}
    """

    body = f"""
    <div class="panel">
      <form method="get" action="/search">
        <div class="row">
          <div style="flex:1; min-width:250px">
            <label>关键词</label>
            <input name="q" value="{esc(q)}" placeholder="客户名、付款方、精确金额或备注，可组合搜索" autofocus>
          </div>
          <input type="hidden" name="user" value="{esc(actor['id'])}">
          <input type="hidden" name="name" value="{esc(actor['name'])}">
          <input type="hidden" name="department" value="{esc(actor['department'])}">
          <input type="hidden" name="role" value="{esc(actor['role'])}">
          <div><button type="submit">搜索</button></div>
        </div>
      </form>
      <p class="hint">空搜索、过宽关键词、单独日期搜索会被限制；可输入客户名、金额或备注组合搜索。</p>
    </div>
    {batch_claim_panel_html(actor)}
    {result_html}
    {pending_html}
    {search_script()}
    """
    return page("到款认领搜索", body, active="search", subtitle="输入客户名、金额或备注，找到属于你部门的到款并提交认领", actor=actor)


@app.get("/split-claim", response_class=HTMLResponse)
def split_claim_page(request: Request, q: str = "") -> HTMLResponse:
    rate_limit(request, "search")
    actor = actor_from_request(request)
    if not actor.get("authed"):
        if feishu_enabled():
            login_html = '<a href="/login?next=/split-claim" class="login-btn">飞书登录</a>'
            prompt = "请先通过飞书登录。登录后才能提交分摊认领。"
        else:
            login_html = ""
            prompt = "系统尚未配置飞书登录，请联系管理员配置后再使用分摊认领。"
        body = f"""
        <div class="panel">
          <div class="callout info" style="margin-bottom:0">
            <strong>分摊认领仅对组织内同事开放</strong><br>
            {esc(prompt)}
          </div>
          {f'<div style="margin-top:14px">{login_html}</div>' if login_html else ''}
        </div>
        """
        return page("分摊认领", body, active="split", subtitle="一笔到款拆分给多个部门、中心和项目", actor=actor)

    q = q.strip()
    message = ""
    rows: list[sqlite3.Row] = []
    if q:
        id_query = bool(re.fullmatch(r"#?\s*\d+", q))
        ok, reason = (True, "") if id_query else validate_search_query(q)
        with get_conn() as conn:
            if not ok:
                message = reason
                audit(conn, actor, "split_search_blocked", None, {"query": q, "reason": reason}, request)
            else:
                rows = split_claim_candidates(conn, q)
                audit(conn, actor, "split_search", None, {"query": q, "result_count": len(rows)}, request)

    result_html = ""
    if message:
        result_html = f'<div class="callout warn"><strong>搜索被限制：</strong>{esc(message)}</div>'
    elif q and not rows:
        result_html = '<div class="callout info">没有找到可分摊认领的到款。请换客户名、金额、备注或到款 ID 试试。</div>'
    elif rows:
        table_rows = []
        with get_conn() as conn:
            for row in rows:
                table_rows.append(
                    f"""
                    <tr>
                      <td class="nowrap">#{row['id']}<br>{esc(row["received_date"])}</td>
                      <td class="num">¥ {money(row["amount_cents"])}</td>
                      <td>{esc(receiver_company_label(row["receiver_company"]))}</td>
                      <td><strong>{esc(row["payer_name"])}</strong><div class="muted">{esc(row["bank_note"])}</div></td>
                      <td>{status_badge(row["status"])}{claim_summary_html(conn, row)}</td>
                    </tr>
                    <tr>
                      <td colspan="5" class="split-form-cell">{split_claim_form_html(row)}</td>
                    </tr>
                    """
                )
        result_html = f"""
        <div class="table-wrap">
        <table>
          <thead><tr><th>到款</th><th>金额</th><th>到款公司</th><th>付款方 / 备注</th><th>状态</th></tr></thead>
          <tbody>{''.join(table_rows)}</tbody>
        </table>
        </div>
        """

    body = f"""
    <div class="panel">
      <form method="get" action="/split-claim">
        <div class="row">
          <div style="flex:1; min-width:250px">
            <label>查找到款</label>
            <input name="q" value="{esc(q)}" placeholder="客户名、金额、备注，或直接输入到款 ID（如 #17）" autofocus>
          </div>
          <div><button type="submit">搜索</button></div>
        </div>
      </form>
      <p class="hint">适合一笔到款对应多个产品或多个部门归属的场景。每一行分摊会作为已认领明细记录。</p>
    </div>
    {result_html}
    {catalog_script()}
    """
    return page("分摊认领", body, active="split", subtitle="一笔到款拆分给多个部门、中心和项目", actor=actor)


@app.post("/split-claim/{payment_id}")
def submit_split_claim(
    request: Request,
    payment_id: int,
    departments: list[str] = Form([]),
    teams: list[str] = Form([]),
    projects: list[str] = Form([]),
    amounts: list[str] = Form([]),
    notes: list[str] = Form([]),
) -> RedirectResponse:
    actor = actor_from_request(request)
    if not actor.get("authed"):
        raise HTTPException(status_code=403, detail="请先用飞书登录后再提交分摊认领")
    with get_conn() as conn:
        submit_split_claims(conn, actor, payment_id, departments, teams, projects, amounts, notes, request)
    return RedirectResponse(f"/split-claim?q=%23{payment_id}", status_code=303)


@app.post("/claim/batch")
def submit_batch_claim_route(
    request: Request,
    payment_ids: Optional[list[int]] = Form(None),
    department: str = Form(...),
    team: str = Form(""),
    customer_project: str = Form(...),
    note: str = Form(""),
) -> RedirectResponse:
    actor = actor_from_request(request)
    if not actor.get("authed"):
        raise HTTPException(status_code=403, detail="请先用飞书登录后再提交批量认领")
    with get_conn() as conn:
        submit_batch_claims(
            conn,
            actor,
            payment_ids or [],
            department,
            team,
            customer_project,
            note,
            request,
        )
    return RedirectResponse("/search", status_code=303)


def validate_search_query(q: str) -> tuple[bool, str]:
    normalized = q.strip()
    if not normalized:
        return False, "关键词不能为空。"
    numeric = parse_amount(normalized) > 0
    compact = re.sub(r"\s+", "", normalized)
    if len(compact) < 2 and not numeric:
        return False, "关键词太短，请至少输入 2 个有效字符。"
    if compact in BROAD_TERMS:
        return False, "关键词过宽，请输入更具体的客户名、金额或备注。"
    if re.fullmatch(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", compact):
        return False, "普通用户不能只按日期搜索，请增加客户名、金额或备注。"
    return True, ""


def run_search(conn: sqlite3.Connection, q: str) -> list[sqlite3.Row]:
    terms = [term for term in re.split(r"\s+", q.strip()) if term]
    clauses = []
    params: list[Any] = []
    amount_cents = parse_amount(q)
    if amount_cents > 0 and len(terms) == 1:
        clauses.append("amount_cents = ?")
        params.append(amount_cents)
    else:
        for term in terms:
            amount = parse_amount(term)
            if amount > 0 and re.fullmatch(r"[¥￥]?\d[\d,]*(\.\d{1,2})?", term):
                clauses.append("amount_cents = ?")
                params.append(amount)
            else:
                like = f"%{term}%"
                clauses.append(
                    "(payer_name LIKE ? OR bank_note LIKE ? OR received_date LIKE ? OR serial_no LIKE ?)"
                )
                params.extend([like, like, like, like])

    where = " AND ".join(clauses)
    return conn.execute(
        f"""
        SELECT *
        FROM payments
        WHERE status IN ('pending', 'partial_claiming', 'claimed', 'pending_confirm')
          AND {where}
        ORDER BY received_date DESC, id DESC
        """,
        params,
    ).fetchall()


@app.post("/claim/{payment_id}")
def submit_claim(
    request: Request,
    payment_id: int,
    user: str = Form("demo-user"),
    name: str = Form("演示用户"),
    role: str = Form("claimant"),
    department: str = Form(...),
    team: str = Form(""),
    customer_project: str = Form(...),
    contract_invoice: str = Form(""),
    claim_amount: str = Form(""),
    note: str = Form(""),
) -> RedirectResponse:
    department = require_department(department)
    team = team.strip()
    customer_project = customer_project.strip()
    if team not in CATALOG.get(department, {}):
        raise HTTPException(status_code=400, detail="请选择该部门下的中心/小组")
    if customer_project not in CATALOG[department][team]:
        raise HTTPException(status_code=400, detail="请选择该中心/小组下的项目")
    # 已登录则认领归到会话身份，不信任表单 user/name（防冒名认领）；未登录走 demo 表单身份
    session = read_session(request.cookies.get(SESSION_COOKIE, ""))
    if session:
        actor = {**actor_from_request(request), "department": department}
    else:
        if REQUIRE_LOGIN_FOR_CLAIM:
            raise HTTPException(status_code=403, detail="请先用飞书登录后再提交认领")
        actor = actor_from_form(user, name, role, department)
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        if row["status"] not in {"pending", "partial_claiming", "claimed", "pending_confirm"}:
            raise HTTPException(status_code=409, detail="这笔款当前状态不能认领")

        totals = claim_totals(conn, payment_id)
        remaining_amount = max(row["amount_cents"] - totals["active"], 0)
        requested_amount = parse_amount(claim_amount)
        if requested_amount <= 0:
            requested_amount = remaining_amount if totals["active"] > 0 else row["amount_cents"]
        if requested_amount <= 0:
            raise HTTPException(status_code=400, detail="认领金额必须大于 0")
        if requested_amount > row["amount_cents"]:
            raise HTTPException(status_code=400, detail="认领金额不能超过到款金额")

        if totals["active"] + requested_amount > row["amount_cents"]:
            raise HTTPException(status_code=409, detail="认领金额超过该笔款剩余可认领金额")

        full_single_claim = (
            row["status"] == "pending"
            and totals["active"] == 0
            and requested_amount == row["amount_cents"]
        )
        conflict = row["status"] in {"claimed", "pending_confirm"} and (
            (row["claimed_department"] or "") != department or (row["claimed_by"] or "") != actor["id"]
        )
        claim_status = "accepted"
        conn.execute(
            """
            INSERT INTO claims
                (payment_id, department, team, amount_cents, actor_id, actor_name, customer_project, contract_invoice, note, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payment_id,
                department,
                team.strip(),
                requested_amount,
                actor["id"],
                actor["name"],
                customer_project,
                contract_invoice,
                note,
                claim_status,
                now_text(),
            ),
        )

        if not full_single_claim or conflict:
            refresh_payment_claim_status(conn, payment_id)
            action = "claim_partial_submit"
        else:
            conn.execute(
                """
                UPDATE payments
                SET status = 'claimed',
                    claimed_department = ?,
                    claimed_team = ?,
                    claimed_by = ?,
                    claimed_by_name = ?,
                    claimed_at = ?,
                    customer_project = ?,
                    contract_invoice = ?,
                    claim_note = ?
                WHERE id = ?
                """,
                (
                    department,
                    team.strip(),
                    actor["id"],
                    actor["name"],
                    now_text(),
                    customer_project,
                    contract_invoice,
                    note,
                    payment_id,
                ),
            )
            action = "claim_submit"
        audit(
            conn,
            actor,
            action,
            payment_id,
            {
                "department": department,
                "team": team.strip(),
                "customer_project": customer_project,
                "amount_cents": requested_amount,
            },
            request,
        )

    return RedirectResponse(url("/search", user=user, name=name, department=department), status_code=303)


ROLE_LABELS = {
    "claimant": "普通用户",
    "finance": "管理员",
    "admin": "管理员",
    "general_manager": "事业部总经理",
    "superadmin": "超级管理员",
}


def managed_role_options(current: str) -> str:
    current = normalize_managed_role(current)
    items = [
        ("claimant", "普通用户"),
        ("admin", "管理员"),
        ("general_manager", "事业部总经理"),
    ]
    return "".join(
        f'<option value="{value}" {"selected" if value == current else ""}>{label}</option>'
        for value, label in items
    )


def effective_app_user_role(user: sqlite3.Row) -> str:
    managed_role = normalize_managed_role(user["managed_role"] if "managed_role" in user.keys() else "")
    if managed_role != "claimant":
        return managed_role
    return "admin" if user["is_admin"] else "claimant"


def role_badge(role: str) -> str:
    css = {
        "superadmin": "closed",
        "admin": "claimed",
        "general_manager": "partial_claiming",
        "claimant": "",
    }.get(role, "")
    return f'<span class="status {esc(css)}">{esc(ROLE_LABELS.get(role, role))}</span>'


DASHBOARD_ACTIVE_CLAIM_STATUSES = {"pending", "accepted"}
DASHBOARD_VISIBLE_PAYMENT_STATUSES = {
    "draft",
    "pending",
    "partial_claiming",
    "claimed",
    "pending_confirm",
}


def dashboard_period_ranges(today: Optional[date] = None) -> list[tuple[str, date, date]]:
    today = today or date.today()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    return [
        ("每日", today, today),
        ("每周", week_start, today),
        ("每月", month_start, today),
    ]


def dashboard_scope_label(actor: dict[str, str]) -> str:
    if actor.get("dashboard_scope_label"):
        return actor["dashboard_scope_label"]
    if actor["role"] in {"finance", "admin", "superadmin"}:
        return "全公司"
    dept = actor.get("department", "")
    return dept if dept in DEPARTMENTS else "未设置部门"


def dashboard_entries(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    start: date,
    end: date,
) -> list[dict[str, Any]]:
    start_text = start.strftime("%Y-%m-%d")
    end_text = end.strftime("%Y-%m-%d")
    visible_statuses = tuple(DASHBOARD_VISIBLE_PAYMENT_STATUSES)
    active_claim_statuses = tuple(DASHBOARD_ACTIVE_CLAIM_STATUSES)

    if actor["role"] in {"finance", "admin", "superadmin"}:
        rows = conn.execute(
            f"""
            SELECT p.id, p.payer_name, p.bank_note, p.receiver_company, p.amount_cents, p.claimed_department,
                   c.department AS claim_department, c.team AS claim_team,
                   c.customer_project AS claim_project, c.actor_name AS claim_actor_name,
                   c.amount_cents AS claim_amount
            FROM payments p
            LEFT JOIN claims c
              ON c.payment_id = p.id
             AND c.status IN ({",".join("?" for _ in active_claim_statuses)})
            WHERE p.received_date BETWEEN ? AND ?
              AND p.status IN ({",".join("?" for _ in visible_statuses)})
            ORDER BY p.id
            """,
            [*active_claim_statuses, start_text, end_text, *visible_statuses],
        ).fetchall()
        grouped: dict[int, dict[str, Any]] = {}
        for row in rows:
            payment = grouped.setdefault(
                row["id"],
                {
                    "payer_name": row["payer_name"] or "未填写付款方",
                    "bank_note": row["bank_note"] or "",
                    "receiver_company": row["receiver_company"] or "",
                    "amount_cents": row["amount_cents"] or 0,
                    "claimed_department": row["claimed_department"] or "",
                    "claims": [],
                },
            )
            if row["claim_department"]:
                payment["claims"].append(
                    {
                        "department": row["claim_department"],
                        "team": row["claim_team"] or "",
                        "customer_project": row["claim_project"] or "",
                        "actor_name": row["claim_actor_name"] or "",
                        "amount_cents": row["claim_amount"] or 0,
                    }
                )

        entries: list[dict[str, Any]] = []
        for payment in grouped.values():
            active_sum = 0
            for claim in payment["claims"]:
                amount = claim["amount_cents"]
                if amount == 0:
                    continue
                active_sum += amount
                entries.append(
                    {
                        "payer_name": payment["payer_name"],
                        "bank_note": payment["bank_note"],
                        "receiver_company": payment["receiver_company"],
                        "department": claim["department"] or "未认领",
                        "team": claim["team"],
                        "customer_project": claim["customer_project"],
                        "actor_name": claim["actor_name"],
                        "amount_cents": amount,
                        "is_claim": True,
                    }
                )
            remaining = max(payment["amount_cents"] - active_sum, 0)
            if remaining > 0:
                department = payment["claimed_department"] if not payment["claims"] else ""
                entries.append(
                    {
                        "payer_name": payment["payer_name"],
                        "bank_note": payment["bank_note"],
                        "receiver_company": payment["receiver_company"],
                        "department": department or "未认领",
                        "team": "",
                        "customer_project": "",
                        "actor_name": "",
                        "amount_cents": remaining,
                        "is_claim": False,
                    }
                )
        return entries

    raw_scopes = actor.get("dashboard_scopes")
    scopes = raw_scopes if isinstance(raw_scopes, list) else []
    if not scopes:
        department = actor.get("department", "")
        if department in DEPARTMENTS:
            scopes = [{"department": department, "team": actor.get("team", "")}]
    clauses = []
    scope_params: list[Any] = []
    seen_scope_keys: set[tuple[str, str]] = set()
    for scope in scopes:
        department = str(scope.get("department", "")).strip() if isinstance(scope, dict) else ""
        team = str(scope.get("team", "")).strip() if isinstance(scope, dict) else ""
        if department not in DEPARTMENTS:
            continue
        if team and team not in CATALOG.get(department, {}):
            continue
        scope_key = (department, team)
        if scope_key in seen_scope_keys:
            continue
        seen_scope_keys.add(scope_key)
        if team:
            clauses.append("(c.department = ? AND c.team = ?)")
            scope_params.extend([department, team])
        else:
            clauses.append("c.department = ?")
            scope_params.append(department)
    if not clauses:
        return []
    scope_where = " OR ".join(clauses)
    params: list[Any] = [*active_claim_statuses, start_text, end_text, *visible_statuses, *scope_params]

    rows = conn.execute(
        f"""
        SELECT p.payer_name, p.bank_note, p.receiver_company, c.department, c.team,
               c.customer_project, c.actor_name, c.amount_cents
        FROM claims c
        JOIN payments p ON p.id = c.payment_id
        WHERE c.status IN ({",".join("?" for _ in active_claim_statuses)})
          AND p.received_date BETWEEN ? AND ?
          AND p.status IN ({",".join("?" for _ in visible_statuses)})
          AND ({scope_where})
        ORDER BY c.id DESC
        """,
        params,
    ).fetchall()
    return [
        {
            "payer_name": row["payer_name"] or "未填写付款方",
            "bank_note": row["bank_note"] or "",
            "receiver_company": row["receiver_company"] or "",
            "department": row["department"] or "未认领",
            "team": row["team"] or "",
            "customer_project": row["customer_project"] or "",
            "actor_name": row["actor_name"] or "",
            "amount_cents": row["amount_cents"] or 0,
            "is_claim": True,
        }
        for row in rows
        if (row["amount_cents"] or 0) != 0
    ]


def summarize_dashboard_entries(entries: list[dict[str, Any]], limit: int = 50) -> dict[str, Any]:
    customers: dict[str, int] = {}
    departments: dict[str, int] = {}
    rows: list[dict[str, Any]] = []
    total = 0
    for entry in entries:
        amount = int(entry.get("amount_cents") or 0)
        if amount == 0:
            continue
        total += amount
        payer_name = str(entry.get("payer_name") or "未填写付款方")
        bank_note = str(entry.get("bank_note") or "")
        receiver_company = str(entry.get("receiver_company") or "")
        department = str(entry.get("department") or "未认领")
        team = str(entry.get("team") or "")
        customer_project = str(entry.get("customer_project") or "")
        actor_name = str(entry.get("actor_name") or "")
        customers[payer_name] = customers.get(payer_name, 0) + amount
        departments[department] = departments.get(department, 0) + amount
        rows.append(
            {
                "payer_name": payer_name,
                "bank_note": bank_note,
                "receiver_company": receiver_company,
                "department": department,
                "team": team,
                "customer_project": customer_project,
                "actor_name": actor_name,
                "amount_cents": amount,
                "is_claim": bool(entry.get("is_claim")),
            }
        )

    def top_items(values: dict[str, int]) -> list[tuple[str, int]]:
        return sorted(values.items(), key=lambda item: (-item[1], item[0]))[:limit]

    return {
        "total_cents": total,
        "customers": top_items(customers),
        "departments": top_items(departments),
        "rows": sorted(
            rows,
            key=lambda item: (-item["amount_cents"], item["payer_name"], item["department"]),
        )[:limit],
    }


def personal_dashboard_data(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    today: Optional[date] = None,
    custom_start: Optional[date] = None,
    custom_end: Optional[date] = None,
) -> list[dict[str, Any]]:
    result = []
    periods = (
        [("筛选区间", custom_start, custom_end)]
        if custom_start is not None and custom_end is not None
        else dashboard_period_ranges(today)
    )
    for label, start, end in periods:
        entries = dashboard_entries(conn, actor, start, end)
        result.append(
            {
                "label": label,
                "start": start,
                "end": end,
                **summarize_dashboard_entries(entries),
            }
        )
    return result


def dashboard_custom_range(start_text: str = "", end_text: str = "") -> Optional[tuple[date, date]]:
    start_text = (start_text or "").strip()
    end_text = (end_text or "").strip()
    if not start_text and not end_text:
        return None
    if not start_text:
        start_text = end_text
    if not end_text:
        end_text = start_text
    try:
        start = date.fromisoformat(start_text)
        end = date.fromisoformat(end_text)
    except ValueError:
        raise HTTPException(status_code=400, detail="看板日期格式不正确")
    if end < start:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")
    return start, end


def render_dashboard_rows(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return '<div class="dash-table-wrap"><div class="muted" style="padding:12px 0">暂无款项数据</div></div>'
    def claim_details(row: dict[str, Any]) -> str:
        if not row.get("is_claim"):
            return ""
        parts = [
            str(row.get("department") or "").strip(),
            str(row.get("team") or "").strip(),
            str(row.get("customer_project") or "").strip(),
            str(row.get("actor_name") or "").strip(),
        ]
        detail = " · ".join(esc(part) for part in parts if part)
        if not detail:
            return ""
        return f"""
        <details class="dash-claim-details">
          <summary>查看认领信息</summary>
          <div>{detail} · ¥ {money(row["amount_cents"])}</div>
        </details>
        """

    body = "".join(
        f"""
        <tr>
          <td>{esc(row["payer_name"])}</td>
          <td>{esc(receiver_company_label(row.get("receiver_company")))}</td>
          <td>{esc(row.get("bank_note") or "")}{claim_details(row)}</td>
          <td>{esc(row["department"])}</td>
          <td class="num">¥ {money(row["amount_cents"])}</td>
        </tr>
        """
        for row in rows
    )
    return f"""
    <div class="dash-table-wrap">
      <table class="dash-table">
        <colgroup>
          <col style="width:28%">
          <col style="width:20%">
          <col style="width:26%">
          <col style="width:14%">
          <col style="width:12%">
        </colgroup>
        <thead><tr><th>付款客户</th><th>到款公司</th><th>摘要</th><th>款项所属部门</th><th>金额</th></tr></thead>
        <tbody>{body}</tbody>
      </table>
    </div>
    """


def render_dashboard_scope_selector(
    choices: list[dict[str, Any]],
    selected_key: str,
    show: bool = True,
    start_date: str = "",
    end_date: str = "",
) -> str:
    if not show:
        return ""
    options = "".join(
        f'<option value="{esc(choice["key"])}"{" selected" if choice["key"] == selected_key else ""}>{esc(choice["label"])}</option>'
        for choice in choices
    )
    date_inputs = ""
    if start_date and end_date:
        date_inputs = (
            f'<input type="hidden" name="start_date" value="{esc(start_date)}">'
            f'<input type="hidden" name="end_date" value="{esc(end_date)}">'
        )
    return f"""
    <form method="get" action="/me" class="row" style="align-items:end; margin:-2px 0 14px">
      {date_inputs}
      <div style="min-width:220px">
        <label>当前查看范围</label>
        <select name="scope" onchange="this.form.submit()">{options}</select>
      </div>
      <div><button class="secondary" type="submit">切换</button></div>
    </form>
    """


def dashboard_scope_selector_visible(choices: list[dict[str, Any]]) -> bool:
    # choices[0] 是“全部角色”；至少还要有两个实际范围，切换才有意义。
    return len(choices) > 2


def render_dashboard_date_filter(selected_scope: str, start_date: str = "", end_date: str = "") -> str:
    today_text = date.today().isoformat()
    start_value = start_date or today_text
    end_value = end_date or today_text
    reset_action = (
        f'<a class="secondary-link" href="{esc(url("/me", scope=selected_scope))}">恢复默认</a>'
        if start_date and end_date
        else ""
    )
    return f"""
    <form method="get" action="/me" class="dash-date-filter">
      <input type="hidden" name="scope" value="{esc(selected_scope)}">
      <div class="date-field"><label>开始日期</label><input type="date" name="start_date" value="{esc(start_value)}" required></div>
      <div class="date-field"><label>结束日期</label><input type="date" name="end_date" value="{esc(end_value)}" required></div>
      <div class="date-actions"><button class="secondary" type="submit">查询</button>{reset_action}</div>
    </form>
    """


def render_personal_dashboard(
    actor: dict[str, str],
    dashboard: list[dict[str, Any]],
    scope_selector: str = "",
    date_filter: str = "",
) -> str:
    cards = []
    for item in dashboard:
        date_range = item["start"].strftime("%Y-%m-%d")
        if item["start"] != item["end"]:
            date_range += " 至 " + item["end"].strftime("%Y-%m-%d")
        cards.append(
            f"""
            <div class="panel dash-panel">
              <div class="dash-head">
                <div>
                  <strong>{esc(item["label"])}</strong>
                  <div class="muted">{esc(date_range)}</div>
                </div>
                <div class="dash-amount">¥ {money(item["total_cents"])}</div>
              </div>
              {render_dashboard_rows(item["rows"])}
            </div>
            """
        )
    return f"""
    <div class="dash-toolbar">
      <div>
        <h2>数据看板</h2>
        <p class="hint">当前范围：{esc(dashboard_scope_label(actor))}</p>
      </div>
      {date_filter}
    </div>
    {scope_selector}
    <div class="dash-grid">{''.join(cards)}</div>
    """


def profile_setup_modal_html(actor: dict[str, str], cur_dept: str, cur_team: str) -> str:
    if not actor.get("authed") or (cur_dept and cur_team):
        return ""
    dept_opts = '<option value="">请选择部门</option>' + "".join(
        f'<option value="{esc(d)}"{" selected" if d == cur_dept else ""}>{esc(d)}</option>'
        for d in DEPARTMENTS
    )
    return f"""
    <div class="modal-backdrop" role="dialog" aria-modal="true" aria-labelledby="profile-setup-title">
      <div class="modal">
        <h2 id="profile-setup-title">首次使用，请先设置部门和中心</h2>
        <p class="hint">部门和中心会用于到款认领和数据看板。保存后如需调整，请联系超级管理员。</p>
        <form method="post" action="/me/profile">
          <div class="field"><label>部门</label><select name="department" class="cs-dept" required>{dept_opts}</select></div>
          <div class="field"><label>中心 / 小组</label>{team_select("team", cur_dept, cur_team, required=True)}</div>
          <button type="submit">保存并进入</button>
        </form>
      </div>
    </div>
    """


def diagnostic_log_html(actor: dict[str, str], cur_dept: str, cur_team: str) -> str:
    role_label = ROLE_LABELS.get(actor["role"], actor["role"])
    auth_state = "飞书登录" if actor.get("authed") else "演示身份"
    base_lines = [
        "【问题排查日志】",
        f"生成时间: {now_text()}",
        f"应用地址: {APP_BASE_URL or '未配置'}",
        f"用户: {actor.get('name', '')}",
        f"用户ID: {actor.get('id', '')}",
        f"登录方式: {auth_state}",
        f"身份: {role_label}",
        f"部门: {cur_dept or actor.get('department', '') or '未设置'}",
        f"中心/小组: {cur_team or '未设置'}",
    ]
    return f"""
    <h2>问题排查日志</h2>
    <div class="panel">
      <p class="hint" style="margin:0 0 12px">遇到 bug 或报错时，点击复制日志发给管理员。日志会自动生成，不包含密码、密钥或会话凭证。</p>
      <input id="diagnostic-log-base" type="hidden" value="{esc(chr(10).join(base_lines))}">
      <div class="diagnostic-actions">
        <button id="copy-diagnostic-log" class="secondary" type="button">复制日志</button>
        <span id="copy-diagnostic-status" class="copy-status" aria-live="polite"></span>
      </div>
    </div>
    """


@app.get("/me", response_class=HTMLResponse)
def personal_center(
    request: Request,
    scope: str = "all",
    start_date: str = "",
    end_date: str = "",
) -> HTMLResponse:
    actor = actor_from_request(request)
    custom_range = dashboard_custom_range(start_date, end_date)
    selected_start = custom_range[0].isoformat() if custom_range else ""
    selected_end = custom_range[1].isoformat() if custom_range else ""
    with get_conn() as conn:
        extra_scopes = get_user_scopes(conn, actor["id"]) if actor.get("authed") else []
        scope_choices = dashboard_scope_choices(actor, extra_scopes)
        selected_scope = next((choice for choice in scope_choices if choice["key"] == scope), scope_choices[0])
        show_scope_selector = dashboard_scope_selector_visible(scope_choices)
        if not show_scope_selector and len(scope_choices) > 1:
            selected_scope = scope_choices[1]
        dashboard_actor = {
            **actor,
            "dashboard_scopes": selected_scope["scopes"],
            "dashboard_scope_label": selected_scope["label"],
        }
        if actor["role"] in {"finance", "admin", "superadmin"}:
            dashboard_actor = actor
            selected_scope = scope_choices[0]
        dashboard = personal_dashboard_data(
            conn,
            dashboard_actor,
            custom_start=custom_range[0] if custom_range else None,
            custom_end=custom_range[1] if custom_range else None,
        )
        my_claims = conn.execute(
            """
            SELECT c.id AS c_id, c.department AS c_dept, c.team AS c_team,
                   c.customer_project AS c_proj, c.amount_cents AS c_amount,
                   c.created_at AS c_at, c.status AS c_status,
                   p.received_date, p.payer_name, p.receiver_company, p.bank_note, p.status AS p_status,
                   p.claimed_by, p.claimed_by_name
            FROM claims c
            JOIN payments p ON p.id = c.payment_id
            WHERE c.actor_id = ?
            ORDER BY c.id DESC
            """,
            (actor["id"],),
        ).fetchall()
        accepted = sum(1 for r in my_claims if r["c_status"] == "accepted")
        waiting = sum(1 for r in my_claims if r["c_status"] == "pending")

    role_label = ROLE_LABELS.get(actor["role"], actor["role"])
    initial = esc(actor["name"][:1]) if actor["name"] else "我"
    cur_dept = actor["department"] if actor["department"] in DEPARTMENTS else ""
    cur_team = actor.get("team", "")
    id_note = (
        "已通过飞书登录"
        if actor.get("authed")
        else "当前为演示身份（URL 参数），正式使用请走飞书登录。"
    )
    dept_display = f"{esc(cur_dept)} · {esc(cur_team)}" if cur_dept else "未设置部门"
    identity_card = f"""
    <div class="panel identity-card" style="display:flex; gap:18px; align-items:center">
      <div style="width:56px; height:56px; border-radius:14px; flex:none;
        background:linear-gradient(135deg,#2456d6,#4f7df7); color:#fff;
        display:flex; align-items:center; justify-content:center; font-size:24px; font-weight:600">{initial}</div>
      <div style="flex:1">
        <div style="font-size:18px; font-weight:600">{esc(actor['name'])}</div>
        <div class="muted" style="margin-top:4px">
          {dept_display} · {esc(role_label)}
        </div>
      </div>
      <div class="muted identity-note" style="text-align:right; font-size:12px; max-width:230px">{id_note}</div>
    </div>
    <div class="grid" style="grid-template-columns:repeat(auto-fit,minmax(150px,1fr))">
      <div class="stat" style="--dot:#16a34a"><span class="stat-label">已认领</span><strong>{accepted}</strong></div>
      <div class="stat" style="--dot:#dc2626"><span class="stat-label">待处理</span><strong>{waiting}</strong></div>
      <div class="stat" style="--dot:#2456d6"><span class="stat-label">认领记录总数</span><strong>{len(my_claims)}</strong></div>
    </div>
    """

    profile_modal = profile_setup_modal_html(actor, cur_dept, cur_team)
    scope_selector = render_dashboard_scope_selector(
        scope_choices,
        selected_scope["key"],
        show=show_scope_selector and actor["role"] not in {"finance", "admin", "superadmin"},
        start_date=selected_start,
        end_date=selected_end,
    )
    date_filter = render_dashboard_date_filter(selected_scope["key"], selected_start, selected_end)

    if my_claims:
        rows = []
        for index, r in enumerate(my_claims):
            if r["c_status"] == "accepted":
                state = '<span class="status claimed">已认领</span>'
            elif r["c_status"] == "pending":
                state = status_badge("pending_confirm")
            elif r["c_status"] == "canceled":
                state = status_badge("canceled")
            elif r["c_status"] == "rejected":
                state = status_badge("rejected")
            elif r["p_status"] == "rejected":
                state = status_badge("rejected")
            elif r["claimed_by"] and r["claimed_by"] != actor["id"]:
                state = f'<span class="status rejected">已归他人</span><div class="muted">{esc(r["claimed_by_name"] or "")}</div>'
            else:
                state = status_badge(r["p_status"])
            cancel_action = ""
            if actor.get("authed") and r["c_status"] in {"pending", "accepted"}:
                cancel_action = f"""
                <form method="post" action="/me/claims/{r['c_id']}/cancel" style="margin-top:10px"
                      onsubmit="return confirm('确定取消这条认领吗？取消后这笔款会回到待认领列表。')">
                  <button class="danger" type="submit">取消认领</button>
                </form>
                """
            row_style = ' style="display:none"' if index >= 20 else ""
            rows.append(
                f"""
                <tr data-progressive-group="my-claims"{row_style}>
                  <td class="nowrap">{esc(r["received_date"])}</td>
                  <td>
                    <strong>{esc(r["payer_name"])}</strong>
                    <div class="muted">到款公司：{esc(receiver_company_label(r["receiver_company"]))}</div>
                    <div class="muted">{esc(r["bank_note"])}</div>
                  </td>
                  <td>{esc(r["c_dept"])}<div class="muted">{esc(r["c_team"])} · {esc(r["c_proj"])} · ¥ {money(r["c_amount"])}</div></td>
                  <td class="nowrap muted">{esc(r["c_at"])}</td>
                  <td>{state}{cancel_action}</td>
                </tr>
                """
            )
        more_button = (
            '<div style="padding:12px 0 0; text-align:center"><button type="button" class="secondary" '
            'data-progressive-more data-progressive-group="my-claims" data-progressive-step="20" '
            'data-initial-count="20" data-visible-count="20">显示更多</button></div>'
            if len(my_claims) > 20
            else ""
        )
        claims_html = f"""
        <div class="table-wrap">
        <table>
          <thead><tr><th>到款日期</th><th>付款方 / 到款公司 / 备注</th><th>部门 / 中心 / 项目</th><th>提交时间</th><th>当前状态</th></tr></thead>
          <tbody>{''.join(rows)}</tbody>
        </table>
        {more_button}
        </div>
        """
    else:
        claims_html = '<div class="callout info">你还没有认领过任何到款。去「认领搜索」找到属于你的款项并提交认领吧。</div>'

    body = f"""
    {identity_card}
    {profile_modal}
    {render_personal_dashboard(dashboard_actor, dashboard, scope_selector, date_filter)}
    <h2>我的认领</h2>
    {claims_html}
    {diagnostic_log_html(actor, cur_dept, cur_team)}
    {personal_script()}
    """
    return page("个人中心", body, active="me", subtitle="查看你的身份信息和认领记录", actor=actor)


@app.post("/me/claims/{claim_id}/cancel")
def cancel_my_claim_route(request: Request, claim_id: int) -> RedirectResponse:
    actor = actor_from_request(request)
    if not actor.get("authed"):
        raise HTTPException(status_code=403, detail="请先用飞书登录")
    with get_conn() as conn:
        cancel_my_claim(conn, actor, claim_id, request)
    return RedirectResponse("/me", status_code=303)


@app.post("/me/profile")
def save_my_profile(
    request: Request,
    department: str = Form(...),
    team: str = Form(""),
) -> RedirectResponse:
    session = read_session(request.cookies.get(SESSION_COOKIE, ""))
    if not session:
        raise HTTPException(status_code=403, detail="请先用飞书登录")
    department = require_department(department)
    team = team.strip()
    if team not in CATALOG.get(department, {}):
        raise HTTPException(status_code=400, detail="请选择该部门下的中心/小组")
    actor = {"id": session["id"], "name": session.get("name", ""), "role": session.get("role", "claimant")}
    with get_conn() as conn:
        if not can_self_set_profile(conn, actor["id"]):
            raise HTTPException(status_code=403, detail="部门和中心已设置，如需调整请联系超级管理员")
        conn.execute(
            """
            INSERT INTO user_profiles (open_id, name, department, team, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(open_id) DO UPDATE SET
                name = excluded.name, department = excluded.department,
                team = excluded.team, updated_at = excluded.updated_at
            """,
            (actor["id"], actor["name"], department, team, now_text()),
        )
        audit(conn, actor, "set_profile", None, {"department": department, "team": team}, request)
    return RedirectResponse("/me", status_code=303)


@app.post("/admin/profiles/{open_id}")
def admin_set_profile(
    request: Request,
    open_id: str,
    department: str = Form(...),
    team: str = Form(""),
) -> RedirectResponse:
    # 敏感操作：只有超级管理员能替成员调整部门/中心。
    actor = actor_from_request(request)
    require_superadmin(actor)
    department = require_department(department)
    team = team.strip()
    if team not in CATALOG.get(department, {}):
        raise HTTPException(status_code=400, detail="请选择该部门下的中心/小组")
    with get_conn() as conn:
        cur = conn.execute(
            "UPDATE user_profiles SET department = ?, team = ?, updated_at = ? WHERE open_id = ?",
            (department, team, now_text(), open_id),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="成员不存在")
        audit(conn, actor, "admin_set_profile", None, {"open_id": open_id, "department": department, "team": team}, request)
    return RedirectResponse("/admin/system", status_code=303)


@app.post("/admin/scopes/{open_id}/add")
def admin_add_user_scope(
    request: Request,
    open_id: str,
    department: str = Form(...),
    team: str = Form(""),
    label: str = Form(""),
) -> RedirectResponse:
    actor = actor_from_request(request)
    require_superadmin(actor)
    department = require_department(department)
    team = team.strip()
    label = label.strip()
    if team and team not in CATALOG.get(department, {}):
        raise HTTPException(status_code=400, detail="请选择该部门下的中心/小组")
    with get_conn() as conn:
        profile = conn.execute("SELECT open_id FROM user_profiles WHERE open_id = ?", (open_id,)).fetchone()
        if not profile:
            raise HTTPException(status_code=404, detail="成员不存在")
        existing = conn.execute(
            """
            SELECT id FROM user_scopes
            WHERE open_id = ? AND department = ? AND COALESCE(team, '') = ? AND active = 1
            """,
            (open_id, department, team),
        ).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="该成员已有这个参与范围")
        conn.execute(
            """
            INSERT INTO user_scopes
                (open_id, department, team, label, active, created_at, created_by, created_by_name)
            VALUES (?, ?, ?, ?, 1, ?, ?, ?)
            """,
            (open_id, department, team, label, now_text(), actor["id"], actor["name"]),
        )
        audit(
            conn,
            actor,
            "add_user_scope",
            None,
            {"open_id": open_id, "department": department, "team": team, "label": label},
            request,
        )
    return RedirectResponse("/admin/system", status_code=303)


@app.post("/admin/scopes/{scope_id}/deactivate")
def admin_deactivate_user_scope(request: Request, scope_id: int) -> RedirectResponse:
    actor = actor_from_request(request)
    require_superadmin(actor)
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM user_scopes WHERE id = ?", (scope_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="参与范围不存在")
        conn.execute(
            "UPDATE user_scopes SET active = 0 WHERE id = ?",
            (scope_id,),
        )
        audit(
            conn,
            actor,
            "deactivate_user_scope",
            None,
            {"scope_id": scope_id, "open_id": row["open_id"], "department": row["department"], "team": row["team"] or ""},
            request,
        )
    return RedirectResponse("/admin/system", status_code=303)


@app.post("/admin/admins/{open_id}")
def admin_toggle_admin(
    request: Request,
    open_id: str,
    action: str = Form(...),
) -> RedirectResponse:
    # 敏感操作：用会话实时角色鉴权，不信任表单字段，杜绝越权
    actor = actor_from_request(request)
    require_superadmin(actor)
    if open_id in FEISHU_SUPERADMIN_OPEN_IDS:
        raise HTTPException(status_code=400, detail="超级管理员是根权限，不可更改")
    with get_conn() as conn:
        row = conn.execute("SELECT open_id FROM app_users WHERE open_id = ?", (open_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="成员不存在")
        is_admin = 1 if action == "grant" else 0
        conn.execute(
            "UPDATE app_users SET is_admin = ?, managed_role = ? WHERE open_id = ?",
            (is_admin, "admin" if is_admin else "claimant", open_id),
        )
        audit(conn, actor, "set_admin", None, {"open_id": open_id, "is_admin": is_admin}, request)
    return RedirectResponse("/admin/system", status_code=303)


@app.post("/admin/users/{open_id}/role")
def admin_set_user_role(
    request: Request,
    open_id: str,
    managed_role: str = Form(...),
) -> RedirectResponse:
    # 敏感操作：只有超级管理员能设置托管身份；事业部总经理只是身份标签，不获得后台权限
    actor = actor_from_request(request)
    require_superadmin(actor)
    if open_id in FEISHU_SUPERADMIN_OPEN_IDS:
        raise HTTPException(status_code=400, detail="超级管理员是根权限，不可更改")
    managed_role = normalize_managed_role(managed_role)
    with get_conn() as conn:
        row = conn.execute("SELECT open_id FROM app_users WHERE open_id = ?", (open_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="成员不存在")
        conn.execute(
            "UPDATE app_users SET managed_role = ?, is_admin = ? WHERE open_id = ?",
            (managed_role, 1 if managed_role == "admin" else 0, open_id),
        )
        audit(conn, actor, "set_user_role", None, {"open_id": open_id, "managed_role": managed_role}, request)
    return RedirectResponse("/admin/system", status_code=303)


@app.get("/admin/export/today")
def export_today_payments(request: Request) -> Response:
    rate_limit(request, "export")
    actor = actor_from_request(request)
    require_admin(actor)
    start_date, end_date = admin_export_range(request)
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT id, received_date, received_time, payer_name, amount_cents, bank_note,
                   receiver_company, receiver_account, serial_no, status, claimed_department, claimed_team,
                   customer_project, claimed_by_name, claimed_at, claim_note, finance_note,
                   source_ref, imported_at, confirmed_at
            FROM payments
            WHERE received_date BETWEEN ? AND ?
            ORDER BY received_date DESC, id DESC
            """,
            (start_date, end_date),
        ).fetchall()
        audit(
            conn,
            actor,
            "export_today_payments",
            None,
            {"start_date": start_date, "end_date": end_date, "count": len(rows)},
            request,
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "ID",
            "认领记录ID",
            "到款日期",
            "到款时间",
            "付款方",
            "到款公司",
            "到款金额",
            "认领金额",
            "银行备注",
            "收款账户",
            "流水号",
            "状态",
            "认领状态",
            "认领部门",
            "中心/小组",
            "项目",
            "认领人",
            "认领时间",
            "认领备注",
            "管理备注",
            "来源/凭证",
            "导入时间",
            "确认入池时间",
        ]
    )
    unclaimed_rows: list[tuple[sqlite3.Row, int]] = []
    for row in rows:
        with get_conn() as conn:
            claim_rows = conn.execute(
                "SELECT * FROM claims WHERE payment_id = ? AND status IN ('pending', 'accepted') ORDER BY id",
                (row["id"],),
            ).fetchall()
        active_amount_cents = sum(int(claim["amount_cents"] or 0) for claim in claim_rows)
        remaining_cents = max(int(row["amount_cents"] or 0) - active_amount_cents, 0)
        if row["status"] in {"pending", "partial_claiming"} and remaining_cents > 0:
            unclaimed_rows.append((row, remaining_cents))
        if not claim_rows:
            if row["status"] in {"pending", "partial_claiming"}:
                continue
            claim_rows = [None]
        for claim in claim_rows:
            writer.writerow(
                [
                    row["id"],
                    claim["id"] if claim else "",
                    row["received_date"],
                    row["received_time"],
                    row["payer_name"],
                    row["receiver_company"],
                    money(row["amount_cents"]),
                    money(claim["amount_cents"]) if claim else "",
                    row["bank_note"],
                    row["receiver_account"],
                    row["serial_no"],
                    row["status"],
                    claim["status"] if claim else "",
                    claim["department"] if claim else row["claimed_department"],
                    claim["team"] if claim else row["claimed_team"],
                    claim["customer_project"] if claim else row["customer_project"],
                    claim["actor_name"] if claim else row["claimed_by_name"],
                    claim["created_at"] if claim else row["claimed_at"],
                    claim["note"] if claim else row["claim_note"],
                    row["finance_note"],
                    row["source_ref"],
                    row["imported_at"],
                    row["confirmed_at"],
                ]
            )
    for row, remaining_cents in unclaimed_rows:
        writer.writerow(
            [
                row["id"],
                "",
                row["received_date"],
                row["received_time"],
                row["payer_name"],
                row["receiver_company"],
                money(row["amount_cents"]),
                money(remaining_cents),
                row["bank_note"],
                row["receiver_account"],
                row["serial_no"],
                row["status"],
                "未认领",
                "",
                "",
                "",
                "",
                "",
                "",
                row["finance_note"],
                row["source_ref"],
                row["imported_at"],
                row["confirmed_at"],
            ]
        )
    filename = f"claim_pool_{start_date}.csv" if start_date == end_date else f"claim_pool_{start_date}_to_{end_date}.csv"
    return Response(
        "\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def plain_text_date_label(date_text: str) -> str:
    try:
        day = datetime.strptime(date_text, "%Y-%m-%d")
    except ValueError:
        return date_text
    return f"{day.month}月{day.day}日"


def plain_text_range_label(start_date: str, end_date: str) -> str:
    if start_date == end_date:
        return plain_text_date_label(start_date)
    return f"{plain_text_date_label(start_date)}-{plain_text_date_label(end_date)}"


def admin_export_range(request: Request) -> tuple[str, str]:
    today = datetime.now().strftime("%Y-%m-%d")
    params = request.query_params if hasattr(request, "query_params") else {}
    raw_date = str(params.get("date", "")).strip()
    raw_start = str(params.get("start_date", "")).strip()
    raw_end = str(params.get("end_date", "")).strip()
    if raw_date and not raw_start and not raw_end:
        raw_start = raw_date
        raw_end = raw_date
    if raw_start and not raw_end:
        raw_end = raw_start
    if raw_end and not raw_start:
        raw_start = raw_end
    if not raw_start and not raw_end:
        raw_start = today
        raw_end = today

    start_date = parse_date(raw_start)
    end_date = parse_date(raw_end)
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="导出日期格式不正确")
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")
    return start_date, end_date


def build_today_claim_plain_text(conn: sqlite3.Connection, date_text: str) -> str:
    return build_claim_plain_text(conn, date_text, date_text)


def unclaimed_payer_amounts(conn: sqlite3.Connection, start_date: str, end_date: str) -> list[tuple[str, int]]:
    rows = conn.execute(
        """
        SELECT
            p.id AS payment_id,
            p.payer_name AS payer_name,
            p.amount_cents AS payment_amount_cents,
            COALESCE(SUM(CASE WHEN c.status IN ('pending', 'accepted') THEN c.amount_cents ELSE 0 END), 0)
                AS claimed_amount_cents
        FROM payments p
        LEFT JOIN claims c ON c.payment_id = p.id
        WHERE p.received_date BETWEEN ? AND ?
          AND p.status IN ('pending', 'partial_claiming')
        GROUP BY p.id
        """,
        (start_date, end_date),
    ).fetchall()
    payer_items: dict[str, int] = {}
    for row in rows:
        payer = (row["payer_name"] or "").strip() or "未填写付款方"
        remaining = int(row["payment_amount_cents"] or 0) - int(row["claimed_amount_cents"] or 0)
        if remaining > 0:
            payer_items[payer] = payer_items.get(payer, 0) + remaining
    return [(payer, amount_cents) for payer, amount_cents in payer_items.items()]


def build_claim_plain_text(conn: sqlite3.Connection, start_date: str, end_date: str) -> str:
    rows = conn.execute(
        """
        SELECT
            p.id AS payment_id,
            p.received_date AS received_date,
            p.payer_name AS payer_name,
            c.department AS department,
            c.customer_project AS customer_project,
            c.amount_cents AS amount_cents
        FROM payments p
        JOIN claims c ON c.payment_id = p.id
        WHERE p.received_date BETWEEN ? AND ?
          AND p.status != 'closed'
          AND c.status IN ('pending', 'accepted')
        ORDER BY p.received_date, p.id, c.id
        """,
        (start_date, end_date),
    ).fetchall()

    payer_items: dict[str, dict[str, Any]] = {}
    for row in rows:
        payer = (row["payer_name"] or "").strip() or "未填写付款方"
        department = (row["department"] or "").strip() or "未填写部门"
        project = (row["customer_project"] or "").strip() or "未填写项目"
        amount_cents = int(row["amount_cents"] or 0)
        if amount_cents == 0:
            continue

        payer_item = payer_items.setdefault(payer, {"total": 0, "departments": {}})
        payer_item["total"] += amount_cents
        dept_items = payer_item["departments"].setdefault(department, {})
        dept_items[project] = dept_items.get(project, 0) + amount_cents

    lines = [plain_text_range_label(start_date, end_date), ""]
    grand_total = 0
    index = 0
    for index, (payer, item) in enumerate(payer_items.items(), start=1):
        grand_total += item["total"]
        department_parts = []
        for department, projects in item["departments"].items():
            project_parts = [f"{project}{money(amount)}元" for project, amount in projects.items()]
            department_parts.append(f"{department}  {'，'.join(project_parts)}")
        lines.append(f"{index}.{payer}\t{money(item['total'])}元（{'；'.join(department_parts)}）")
    for payer, amount_cents in unclaimed_payer_amounts(conn, start_date, end_date):
        index += 1
        lines.append(f"{index}.{payer}\t{money(amount_cents)}元（未认领）")
        grand_total += amount_cents
    total_label = "今日合计" if start_date == end_date else "区间合计"
    lines.append(f"{total_label}：{money(grand_total)}元")
    return "\n".join(lines)


@app.get("/admin/export/today-text")
def export_today_claim_plain_text(request: Request) -> Response:
    rate_limit(request, "export")
    actor = actor_from_request(request)
    require_admin(actor)
    start_date, end_date = admin_export_range(request)
    with get_conn() as conn:
        text = build_claim_plain_text(conn, start_date, end_date)
        active_count = len([line for line in text.splitlines() if re.match(r"^\d+\.", line)])
        audit(
            conn,
            actor,
            "export_today_claim_plain_text",
            None,
            {"start_date": start_date, "end_date": end_date, "count": active_count},
            request,
        )
    return Response(text, media_type="text/plain; charset=utf-8")


def batch_status_badge(status: str) -> str:
    labels = {
        "draft": ("待确认", "draft"),
        "confirmed": ("已入池", "claimed"),
        "canceled": ("已取消", "rejected"),
    }
    label, css = labels.get(status, (status, ""))
    return f'<span class="status {esc(css)}">{esc(label)}</span>'


def render_batch_actions(row: sqlite3.Row, actor: dict[str, str]) -> str:
    if row["status"] != "draft":
        return '<span class="muted">无需操作</span>'
    return f"""
    <div class="row" style="gap:8px; align-items:center">
      <form method="post" action="/admin/batches/{row['id']}/confirm">
        {finance_hidden(actor)}
        <button type="submit">确认入池</button>
      </form>
      <form method="post" action="/admin/batches/{row['id']}/cancel" onsubmit="return confirm('确定取消这个导入批次吗？待确认流水会被移除。')">
        {finance_hidden(actor)}
        <button class="danger" type="submit">取消</button>
      </form>
    </div>
    """


def admin_notice_html(notice: str, imported: int = 0, skipped: int = 0) -> str:
    if notice == "import_success":
        return (
            '<div class="callout success">'
            f'<strong>导入成功</strong><br>已导入 {esc(imported)} 条流水，跳过 {esc(skipped)} 条重复或无效记录。'
            '</div>'
        )
    if notice == "catalog_updated":
        return '<div class="callout success"><strong>目录更新成功</strong></div>'
    return ""


ADMIN_PAYMENT_SORTS = {"id", "date", "amount", "payer", "status"}


def normalize_sort_dir(direction: str) -> str:
    return "asc" if (direction or "").lower() == "asc" else "desc"


def admin_payment_order_clause(sort: str, direction: str) -> tuple[str, str, str]:
    sort = sort if sort in ADMIN_PAYMENT_SORTS else "id"
    direction = normalize_sort_dir(direction)
    sql_dir = direction.upper()
    status_case = (
        "CASE status "
        "WHEN 'draft' THEN 1 "
        "WHEN 'pending' THEN 2 "
        "WHEN 'partial_claiming' THEN 3 "
        "WHEN 'pending_confirm' THEN 4 "
        "WHEN 'claimed' THEN 5 "
        "WHEN 'rejected' THEN 6 "
        "WHEN 'closed' THEN 7 "
        "ELSE 99 END"
    )
    clauses = {
        "id": f"id {sql_dir}",
        "date": f"received_date {sql_dir}, received_time {sql_dir}, id {sql_dir}",
        "amount": f"amount_cents {sql_dir}, id {sql_dir}",
        "payer": f"payer_name COLLATE NOCASE {sql_dir}, id {sql_dir}",
        "status": f"{status_case} {sql_dir}, id {sql_dir}",
    }
    return clauses[sort], sort, direction


def admin_sort_th(label: str, key: str, current_sort: str, current_dir: str, class_name: str = "") -> str:
    is_active = key == current_sort
    next_dir = "asc"
    arrow = ""
    if is_active:
        next_dir = "desc" if current_dir == "asc" else "asc"
        arrow = f'<span class="sort-arrow">{"↑" if current_dir == "asc" else "↓"}</span>'
    href = "/admin?" + urlencode({"sort": key, "dir": next_dir})
    table_url = "/admin/payments/table?" + urlencode({"sort": key, "dir": next_dir})
    active_class = " active" if is_active else ""
    class_attr = f' class="{esc(class_name)}"' if class_name else ""
    return (
        f'<th{class_attr}><a class="sort-link{active_class}" href="{esc(href)}" '
        f'data-table-url="{esc(table_url)}">{esc(label)}{arrow}</a></th>'
    )


def admin_payment_rows(
    conn: sqlite3.Connection,
    sort: str,
    direction: str,
) -> tuple[list[sqlite3.Row], str, str]:
    payment_order, sort, direction = admin_payment_order_clause(sort, direction)
    closed_cutoff = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    rows = conn.execute(
        f"""
        SELECT * FROM payments
        WHERE status != 'closed'
           OR COALESCE(closed_at, '') = ''
           OR closed_at > ?
        ORDER BY {payment_order}
        """,
        (closed_cutoff,),
    ).fetchall()
    return rows, sort, direction


def admin_claim_search_rows(
    conn: sqlite3.Connection,
    q: str,
    actor: dict[str, str],
    request: Optional[Request] = None,
) -> tuple[list[sqlite3.Row], str]:
    q = q.strip()
    if not q:
        return [], ""
    ok, reason = validate_search_query(q)
    if not ok:
        audit(conn, actor, "admin_search_blocked", None, {"query": q, "reason": reason}, request)
        return [], reason
    rows = run_search(conn, q)
    audit(conn, actor, "admin_search", None, {"query": q, "result_count": len(rows)}, request)
    return rows, ""


def render_payment_pool_html(
    payments: list[sqlite3.Row],
    actor: dict[str, str],
    sort: str,
    direction: str,
) -> str:
    def progressive_payment_attrs(index: int) -> str:
        hidden = ' style="display:none"' if index >= 20 else ""
        return f' data-progressive-group="payment-pool"{hidden}'

    payment_rows = "".join(
        render_admin_payment_row(
            row,
            actor,
            row_attrs=progressive_payment_attrs(index),
        )
        for index, row in enumerate(payments)
    )
    more_button = (
        '<div style="padding:12px 0 0; text-align:center"><button type="button" class="secondary" '
        'data-progressive-more data-progressive-group="payment-pool" data-progressive-step="20" '
        'data-initial-count="20" data-visible-count="20">显示更多</button></div>'
        if len(payments) > 20
        else ""
    )
    return f"""
    <div id="payment-pool-card" class="table-wrap">
    <div class="bulk-bar">
      <form id="bulk-close-form" method="post" action="/admin/payments/bulk-close" onsubmit="return confirmBulkClose()">
        {finance_hidden(actor)}
        <button id="bulk-close-button" class="danger" type="submit" disabled>批量关闭</button>
        <span class="bulk-count">已选 <strong id="bulk-selected-count">0</strong> 条</span>
      </form>
      <span class="muted">关闭不会删除流水、附件或认领记录。</span>
    </div>
    <table class="admin-payment-table">
      <thead><tr>
        <th class="select-cell col-select"><input id="bulk-select-all" type="checkbox" aria-label="全选当前页"></th>
        {admin_sort_th("ID", "id", sort, direction, "col-id")}
        {admin_sort_th("日期", "date", sort, direction, "col-date")}
        {admin_sort_th("金额", "amount", sort, direction, "col-amount")}
        <th class="col-receiver">到款公司</th>
        {admin_sort_th("付款方 / 摘要", "payer", sort, direction)}
        {admin_sort_th("状态 / 认领", "status", sort, direction, "col-status")}
        <th class="col-actions">管理操作</th>
      </tr></thead>
      <tbody>{payment_rows or '<tr><td colspan="8" class="empty">暂无记录</td></tr>'}</tbody>
    </table>
    {more_button}
    </div>
    """


def render_admin_claim_search_panel(
    q: str,
    results: list[sqlite3.Row],
    message: str,
    actor: dict[str, str],
) -> str:
    q = q.strip()
    clear_link = (
        f'<a href="/admin" style="display:inline-block;padding:9px 0">清除搜索</a>'
        if q else ""
    )
    result_html = ""
    if message:
        result_html = f'<div class="callout warn" style="margin-top:14px"><strong>搜索被限制：</strong>{esc(message)}</div>'
    elif q and not results:
        result_html = '<div class="callout info" style="margin-top:14px">没有找到匹配记录。请换一个更具体的客户名、金额或备注关键词。</div>'
    elif results:
        rows = "".join(render_admin_payment_row(row, actor, selectable=False) for row in results)
        result_html = f"""
        <div class="table-wrap" style="box-shadow:none; margin-top:14px">
        <table class="admin-payment-table">
          <thead><tr>
            <th class="select-cell col-select"></th>
            <th class="col-id">ID</th>
            <th class="col-date">日期</th>
            <th class="col-amount">金额</th>
            <th class="col-receiver">到款公司</th>
            <th>付款方 / 摘要</th>
            <th class="col-status">状态 / 认领</th>
            <th class="col-actions">管理操作</th>
          </tr></thead>
          <tbody>{rows}</tbody>
        </table>
        </div>
        """
    return f"""
    <h2>到款认领搜索</h2>
    <div class="panel">
      <form method="get" action="/admin">
        {finance_hidden(actor)}
        <label>关键词</label>
        <div class="row" style="align-items:end">
          <div style="flex:1; min-width:260px">
            <input name="admin_q" value="{esc(q)}" placeholder="客户名、付款方、精确金额或备注，可组合搜索">
          </div>
          <div><button type="submit">搜索</button></div>
          {f'<div>{clear_link}</div>' if clear_link else ''}
        </div>
        <p class="hint">空搜索、过宽关键词、单独日期搜索会被限制；可输入客户名、金额或备注组合搜索。</p>
      </form>
      {result_html}
    </div>
    """


@app.get("/admin/payments/table", response_class=HTMLResponse)
def admin_payments_table(
    request: Request,
    sort: str = "id",
    dir: str = "desc",
) -> HTMLResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        payments, sort, dir = admin_payment_rows(conn, sort, dir)
    return HTMLResponse(render_payment_pool_html(payments, actor, sort, dir))


@app.get("/admin", response_class=HTMLResponse)
def admin_page(
    request: Request,
    notice: str = "",
    imported: int = 0,
    skipped: int = 0,
    sort: str = "id",
    dir: str = "desc",
    admin_q: str = "",
) -> HTMLResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    admin_q = admin_q.strip()
    with get_conn() as conn:
        stats = conn.execute(
            """
            SELECT status, COUNT(*) count, COALESCE(SUM(amount_cents), 0) amount
            FROM payments
            GROUP BY status
            """
        ).fetchall()
        batches = conn.execute("SELECT * FROM import_batches ORDER BY id DESC").fetchall()
        payments, sort, dir = admin_payment_rows(conn, sort, dir)
        admin_search_results, admin_search_message = admin_claim_search_rows(conn, admin_q, actor, request)

    stat_map = {row["status"]: row for row in stats}
    stat_html = "".join(
        f'<div class="stat" style="--dot:{dot}">'
        f'<span class="stat-label">{esc(label)}</span>'
        f'<strong>{stat_map.get(status, {"count": 0})["count"] if status in stat_map else 0}</strong>'
        f'<span class="stat-amount">¥ {money(stat_map.get(status, {"amount": 0})["amount"] if status in stat_map else 0)}</span>'
        f"</div>"
        for status, label, dot in [
            ("draft", "待确认", "#64748b"),
            ("pending", "待认领", "#d97706"),
            ("partial_claiming", "部分认领中", "#2456d6"),
            ("claimed", "已认领", "#16a34a"),
            ("rejected", "已驳回", "#94a3b8"),
            ("closed", "已关闭", "#6366f1"),
        ]
    )
    default_export_date = datetime.now().strftime("%Y-%m-%d")

    def progressive_batch_attrs(index: int) -> str:
        hidden = ' style="display:none"' if index >= 10 else ""
        return f' data-progressive-group="admin-batches"{hidden}'

    batch_rows = "".join(
        f"""
        <tr{progressive_batch_attrs(index)}>
          <td class="nowrap">#{row['id']}</td>
          <td>{esc(row['source_name'])}</td>
          <td class="nowrap">{esc(row['created_at'])}</td>
          <td class="nowrap">{esc(row['raw_count'])} / {esc(row['imported_count'])} / {esc(row['skipped_count'])}</td>
          <td>{batch_status_badge(row['status'])}</td>
          <td>{render_batch_actions(row, actor)}</td>
        </tr>
        """
        for index, row in enumerate(batches)
    )
    batch_more_button = (
        '<div style="padding:12px 0 0; text-align:center"><button type="button" class="secondary" '
        'data-progressive-more data-progressive-group="admin-batches" data-progressive-step="10" '
        'data-initial-count="10" data-visible-count="10">显示更多</button></div>'
        if len(batches) > 10
        else ""
    )

    body = f"""
    {admin_notice_html(notice, imported, skipped)}
    <div class="grid admin-stat-grid">{stat_html}</div>

    <div class="panel">
      <div class="row" style="justify-content:space-between; align-items:center">
        <div>
          <div style="font-weight:600">按日期导出</div>
          <p class="hint" style="margin:4px 0 0">选择到款日期范围后导出对应区间的数据；CSV 可直接用 Excel 打开，纯文本用于发到款认领摘要。</p>
        </div>
        <div class="row" style="gap:8px; align-items:center; justify-content:flex-end">
          <label class="muted" for="export-start-date" style="display:flex;align-items:center;gap:6px">开始日期
            <input id="export-start-date" type="date" value="{esc(default_export_date)}" style="width:150px">
          </label>
          <span class="muted">至</span>
          <label class="muted" for="export-end-date" style="display:flex;align-items:center;gap:6px">结束日期
            <input id="export-end-date" type="date" value="{esc(default_export_date)}" style="width:150px">
          </label>
          <a id="export-date-csv" class="login-btn" href="/admin/export/today?start_date={esc(default_export_date)}&amp;end_date={esc(default_export_date)}">下载 CSV</a>
          <button id="copy-today-plain-text" type="button" class="secondary">复制纯文本</button>
          <span id="copy-today-plain-text-status" class="muted"></span>
        </div>
      </div>
    </div>

    <h2>导入流水</h2>
    <div class="panel">
      <form method="post" action="/admin/import" enctype="multipart/form-data">
        {finance_hidden(actor)}
        <div class="row">
          <div style="flex:1.4; min-width:220px">
            <label>来源名称</label>
            <input name="source_name" value="{datetime.now().strftime('%Y-%m-%d')} 银行流水">
          </div>
          <div style="flex:1; min-width:220px">
            <label>原始凭证附件</label>
            <input type="file" name="attachment" accept=".pdf,.xls,.xlsx">
          </div>
        </div>
        <p class="hint">可上传 PDF、.xls 或 .xlsx 文件；Excel 会自动读取所有非空工作表。</p>
        <button type="submit">导入为待确认</button>
      </form>
    </div>

    {render_admin_claim_search_panel(admin_q, admin_search_results, admin_search_message, actor)}

    <h2>最近导入批次</h2>
    <div class="table-wrap">
    <table>
      <thead><tr><th>ID</th><th>来源</th><th>创建时间</th><th>原始 / 导入 / 跳过</th><th>状态</th><th>操作</th></tr></thead>
      <tbody>{batch_rows or '<tr><td colspan="6" class="empty">暂无批次</td></tr>'}</tbody>
    </table>
    {batch_more_button}
    </div>

    <h2>全量认领池</h2>
    {render_payment_pool_html(payments, actor, sort, dir)}

    <h2>项目管理</h2>
    <div class="panel">
      <div style="font-weight:600; margin-bottom:12px">组织架构</div>
      <div class="row" style="align-items:end">
        <form method="post" action="/admin/catalog/departments" class="row" style="align-items:end; flex:1; min-width:300px">
          {finance_hidden(actor)}
          <div style="min-width:220px; flex:1"><label>新部门名称</label><input name="department" maxlength="100" required></div>
          <div><button type="submit">添加部门</button></div>
        </form>
        <form method="post" action="/admin/catalog/teams" class="row" style="align-items:end; flex:1.5; min-width:420px">
          {finance_hidden(actor)}
          <div style="min-width:180px; flex:1"><label>所属部门</label>{department_select("department", required=True, class_name="cs-dept")}</div>
          <div style="min-width:220px; flex:1.2"><label>新中心 / 小组名称</label><input name="team" maxlength="100" required></div>
          <div><button type="submit">添加中心/小组</button></div>
        </form>
      </div>
      <div style="font-weight:600; border-top:1px solid var(--line); margin-top:18px; padding-top:18px">项目</div>
      <form method="post" action="/admin/catalog/projects" class="row" style="align-items:end">
        {finance_hidden(actor)}
        <input type="hidden" name="action" value="add">
        <div style="min-width:180px; flex:1"><label>部门</label>{department_select("department", required=True, class_name="cs-dept")}</div>
        <div style="min-width:180px; flex:1"><label>中心 / 小组</label>{team_select("team", "", required=True)}</div>
        <div style="min-width:220px; flex:1.4"><label>新项目名称</label><input name="project" maxlength="100" required></div>
        <div><button type="submit">添加项目</button></div>
      </form>
      <form method="post" action="/admin/catalog/projects" class="row" style="align-items:end; border-top:1px solid var(--line); margin-top:16px; padding-top:16px" onsubmit="return confirm('确定删除这个项目吗？历史认领记录会保留，但后续无法再选择该项目。')">
        {finance_hidden(actor)}
        <input type="hidden" name="action" value="delete">
        <div style="min-width:180px; flex:1"><label>部门</label>{department_select("department", required=True, class_name="cs-dept")}</div>
        <div style="min-width:180px; flex:1"><label>中心 / 小组</label>{team_select("team", "", required=True)}</div>
        <div style="min-width:220px; flex:1.4"><label>项目</label>{project_select("project", "", "", required=True)}</div>
        <div><button class="danger" type="submit">删除项目</button></div>
      </form>
    </div>

    {admin_script()}
    """
    return page("财务后台", body, active="admin", subtitle="导入流水、管理项目、处理认领", actor=actor)


@app.get("/admin/system", response_class=HTMLResponse)
def admin_system_page(request: Request) -> HTMLResponse:
    actor = actor_from_request(request)
    require_superadmin(actor)
    with get_conn() as conn:
        logs = conn.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 30").fetchall()
        profiles = conn.execute(
            "SELECT open_id, name, department, team, updated_at FROM user_profiles ORDER BY updated_at DESC"
        ).fetchall()
        scope_rows = conn.execute(
            "SELECT * FROM user_scopes WHERE active = 1 ORDER BY open_id, id DESC"
        ).fetchall()
        app_user_rows = conn.execute(
            """
            SELECT open_id, name, managed_role, is_admin, last_login
            FROM app_users
            ORDER BY
                CASE
                    WHEN managed_role = 'admin' OR is_admin = 1 THEN 0
                    WHEN managed_role = 'general_manager' THEN 1
                    ELSE 2
                END,
                last_login DESC
            """
        ).fetchall()

    user_scopes_by_open_id: dict[str, list[sqlite3.Row]] = {}
    for scope_row in scope_rows:
        user_scopes_by_open_id.setdefault(scope_row["open_id"], []).append(scope_row)

    profile_parts = []
    for p in profiles:
        user_scope_rows = user_scopes_by_open_id.get(p["open_id"], [])
        scope_badges = "".join(
            f"""
            <div class="muted" style="margin-top:6px">
              参与范围：{esc(scope_display_label(row["department"], row["team"] or "", row["label"] or ""))}
              <form method="post" action="/admin/scopes/{row['id']}/deactivate" style="display:inline" onsubmit="return confirm('确定停用这个参与范围吗？')">
                {finance_hidden(actor)}
                <button class="secondary" type="submit" style="padding:3px 8px; margin-left:6px; font-size:12px">停用</button>
              </form>
            </div>
            """
            for row in user_scope_rows
        )
        profile_parts.append(
            f"""
        <tr>
          <td><strong>{esc(p["name"] or "")}</strong><br><span class="code">{esc(p["open_id"])}</span></td>
          <td class="actions">
            <form method="post" action="/admin/profiles/{esc(p['open_id'])}" class="row" style="align-items:end">
              {finance_hidden(actor)}
              <div style="min-width:170px; flex:1"><label>部门</label>{department_select("department", p["department"] or "", required=True, class_name="cs-dept")}</div>
              <div style="min-width:170px; flex:1"><label>中心 / 小组</label>{team_select("team", p["department"] or "", p["team"] or "", required=True)}</div>
              <div><button class="secondary" type="submit">保存</button></div>
            </form>
            {scope_badges}
            <form method="post" action="/admin/scopes/{esc(p['open_id'])}/add" class="row" style="align-items:end; margin-top:10px; border-top:1px dashed var(--line); padding-top:10px">
              {finance_hidden(actor)}
              <div style="min-width:170px; flex:1"><label>新增参与部门</label>{department_select("department", required=True, class_name="cs-dept")}</div>
              <div style="min-width:170px; flex:1"><label>中心 / 小组（可选）</label>{team_select("team", "", required=False)}</div>
              <div style="min-width:180px; flex:1"><label>显示名称（可选）</label><input name="label" placeholder="如：年会报名组"></div>
              <div><button class="secondary" type="submit">添加参与范围</button></div>
            </form>
          </td>
          <td class="nowrap muted">{esc(p["updated_at"])}</td>
        </tr>
        """
        )
    profile_rows = "".join(profile_parts)

    admin_user_rows = "".join(
        (
            f"""
        <tr>
          <td><strong>{esc(u["name"] or "")}</strong><br><span class="code">{esc(u["open_id"])}</span></td>
          <td>{role_badge("superadmin") if u["open_id"] in FEISHU_SUPERADMIN_OPEN_IDS else role_badge(effective_app_user_role(u))}</td>
          <td class="nowrap muted">{esc(u["last_login"] or "")}</td>
          <td>{
              '<span class="muted">根权限，不可更改</span>'
              if u["open_id"] in FEISHU_SUPERADMIN_OPEN_IDS
              else f'''<form method="post" action="/admin/users/{esc(u['open_id'])}/role" class="row" style="align-items:end">
                {finance_hidden(actor)}
                <div style="min-width:160px"><label>身份</label><select name="managed_role">{managed_role_options(effective_app_user_role(u))}</select></div>
                <button class="secondary" type="submit">保存身份</button>
              </form>'''
          }</td>
        </tr>
        """
        )
        for u in app_user_rows
    )
    log_rows = "".join(
        f"<tr><td class='nowrap'>{esc(row['at'])}</td><td class='nowrap'>{esc(row['actor_name'])}</td><td class='nowrap'>{esc(row['action'])}</td><td>{esc(row['payment_id'])}</td><td><span class='code'>{esc(row['detail_json'])}</span></td></tr>"
        for row in logs
    )

    body = f"""
    <h2>成员部门</h2>
    <p class="hint" style="margin:-4px 0 12px">登录过并设置过部门的成员都在这里。有人选错了，超级管理员可直接改。</p>
    <div class="table-wrap">
    <table>
      <thead><tr><th>成员</th><th style="width:520px">部门 / 中心</th><th>更新时间</th></tr></thead>
      <tbody>{profile_rows or '<tr><td colspan="3" class="empty">还没有成员设置过部门</td></tr>'}</tbody>
    </table>
    </div>

    <h2>身份管理</h2>
    <p class="hint" style="margin:-4px 0 12px">只有超级管理员能看到这里。事业部总经理只是身份标签，不会获得财务后台权限。</p>
    <div class="table-wrap">
    <table>
      <thead><tr><th>成员</th><th>当前身份</th><th>最近登录</th><th>操作</th></tr></thead>
      <tbody>{admin_user_rows or '<tr><td colspan="4" class="empty">还没有人登录过</td></tr>'}</tbody>
    </table>
    </div>

    <h2>最近操作日志</h2>
    <div class="table-wrap">
    <table>
      <thead><tr><th>时间</th><th>用户</th><th>动作</th><th>记录</th><th>详情</th></tr></thead>
      <tbody>{log_rows or '<tr><td colspan="5" class="empty">暂无日志</td></tr>'}</tbody>
    </table>
    </div>
    {admin_script()}
    """
    return page("管理后台", body, active="system", subtitle="管理成员部门、身份与操作日志", actor=actor)


def normalize_catalog_name(value: str, label: str) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    if not value or len(value) > 100:
        raise HTTPException(status_code=400, detail=f"{label}需为 1-100 个字符")
    return value


def save_catalog_change(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    change_type: str,
    department: str,
    team: str,
    project: str,
    active: int,
    audit_action: str,
    request: Optional[Request] = None,
) -> None:
    conn.execute(
        """
        INSERT INTO catalog_project_changes
            (change_type, department, team, project, active, updated_at, updated_by, updated_by_name)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(department, team, project) DO UPDATE SET
            change_type = excluded.change_type,
            active = excluded.active,
            updated_at = excluded.updated_at,
            updated_by = excluded.updated_by,
            updated_by_name = excluded.updated_by_name
        """,
        (change_type, department, team, project, active, now_text(), actor["id"], actor["name"]),
    )
    audit(
        conn,
        actor,
        audit_action,
        None,
        {"department": department, "team": team, "project": project},
        request,
    )


@app.post("/admin/catalog/departments")
def admin_catalog_department(
    request: Request,
    department: str = Form(...),
) -> RedirectResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    department = normalize_catalog_name(department, "部门名称")
    if department in CATALOG:
        raise HTTPException(status_code=409, detail="该部门已存在")
    with get_conn() as conn:
        save_catalog_change(
            conn,
            actor,
            "department",
            department,
            "",
            "",
            1,
            "add_catalog_department",
            request,
        )
    refresh_catalog()
    return RedirectResponse("/admin?notice=catalog_updated", status_code=303)


@app.post("/admin/catalog/teams")
def admin_catalog_team(
    request: Request,
    department: str = Form(...),
    team: str = Form(...),
) -> RedirectResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    department = require_department(department)
    team = normalize_catalog_name(team, "中心/小组名称")
    if team in CATALOG.get(department, {}):
        raise HTTPException(status_code=409, detail="该中心/小组已存在")
    with get_conn() as conn:
        save_catalog_change(
            conn,
            actor,
            "team",
            department,
            team,
            "",
            1,
            "add_catalog_team",
            request,
        )
    refresh_catalog()
    return RedirectResponse("/admin?notice=catalog_updated", status_code=303)


@app.post("/admin/catalog/projects")
def admin_catalog_project(
    request: Request,
    action: str = Form(...),
    department: str = Form(...),
    team: str = Form(...),
    project: str = Form(...),
) -> RedirectResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    department = require_department(department)
    team = team.strip()
    project = re.sub(r"\s+", " ", project).strip()
    if team not in CATALOG.get(department, {}):
        raise HTTPException(status_code=400, detail="请选择该部门下的现有中心/小组")
    if not project or len(project) > 100:
        raise HTTPException(status_code=400, detail="项目名称需为 1-100 个字符")
    visible_projects = CATALOG.get(department, {}).get(team, [])
    if action == "add":
        if project in visible_projects:
            raise HTTPException(status_code=409, detail="该项目已存在")
        active = 1
        audit_action = "add_catalog_project"
    elif action == "delete":
        if project not in visible_projects:
            raise HTTPException(status_code=404, detail="项目不存在或已删除")
        if len(visible_projects) <= 1:
            raise HTTPException(status_code=409, detail="不能删除该中心/小组的最后一个项目")
        active = 0
        audit_action = "delete_catalog_project"
    else:
        raise HTTPException(status_code=400, detail="项目操作不合法")

    with get_conn() as conn:
        save_catalog_change(
            conn,
            actor,
            "project",
            department,
            team,
            project,
            active,
            audit_action,
            request,
        )
    refresh_catalog()
    return RedirectResponse("/admin?notice=catalog_updated", status_code=303)


def finance_hidden(actor: dict[str, str]) -> str:
    return f"""
    <input type="hidden" name="user" value="{esc(actor['id'])}">
    <input type="hidden" name="name" value="{esc(actor['name'])}">
    <input type="hidden" name="role" value="{esc(actor['role'])}">
    """


def render_admin_payment_row(
    row: sqlite3.Row,
    actor: dict[str, str],
    selectable: bool = True,
    row_attrs: str = "",
) -> str:
    with get_conn() as conn:
        claim_rows = conn.execute(
            "SELECT * FROM claims WHERE payment_id = ? ORDER BY id",
            (row["id"],),
        ).fetchall()
        totals = claim_totals(conn, row["id"])
    claimed = ""
    if row["claimed_department"] or row["claimed_by_name"]:
        parts = [part for part in (row["claimed_department"], row["claimed_team"], row["claimed_by_name"]) if part]
        claimed = f'<div class="muted">{esc(" · ".join(parts))}<br>{esc(row["customer_project"])}</div>'
    if totals["active"] > 0:
        claimed += (
            f'<div class="muted">认领合计 ¥ {money(totals["active"])}'
            f' · 已认领 ¥ {money(totals["accepted"])}'
            f' · 剩余 ¥ {money(max((row["amount_cents"] or 0) - totals["active"], 0))}</div>'
        )
    finance_note = ""
    if row["finance_note"]:
        finance_note = f'<div class="muted">管理备注：{esc(row["finance_note"])}</div>'
    active_claim_rows = [claim for claim in claim_rows if claim["status"] in {"pending", "accepted"}]
    claim_details_html = admin_claim_details_html(row, claim_rows)
    claim_note_text = claim_note_summary(row, active_claim_rows)
    claim_note_html = ""
    if claim_note_text:
        claim_note_html = f'<div class="muted">备注说明：{esc(claim_note_text)}</div>'
    # 仅对已有人认领的单子显示「驳回退回」
    reject_ui = ""
    if row["status"] in {"claimed", "pending_confirm"}:
        reject_ui = f"""
        <details class="fold">
          <summary>驳回退回</summary>
          <div class="fold-body">
            <form method="post" action="/admin/payments/{row['id']}/reject">
              {finance_hidden(actor)}
              <div class="field"><label>驳回原因（会飞书通知认领人）</label><input name="reason" placeholder="例如：部门归属不对，请重新认领"></div>
              <button class="danger" type="submit">驳回并退回待认领</button>
            </form>
          </div>
        </details>
        """
    select_cell = (
        f'<td class="select-cell col-select"><input class="bulk-payment-checkbox" form="bulk-close-form" type="checkbox" name="payment_ids" value="{row["id"]}" aria-label="选择流水 #{row["id"]}" {"disabled" if row["status"] == "closed" else ""}></td>'
        if selectable
        else '<td class="select-cell col-select"></td>'
    )
    return f"""
    <tr{row_attrs}>
      {select_cell}
      <td class="nowrap col-id">#{row['id']}<br><span class="muted">批次 {esc(row['batch_id'])}</span></td>
      <td class="nowrap col-date">{esc(row['received_date'])}<br><span class="muted">{esc(row['received_time'])}</span></td>
      <td class="num col-amount">¥ {money(row['amount_cents'])}</td>
      <td class="col-receiver">{esc(receiver_company_label(row['receiver_company']))}</td>
      <td class="payment-summary">
        <strong>{esc(row['payer_name'])}</strong>
        <div>{esc(row['bank_note'])}</div>
        {claim_details_html}
        {claim_note_html}
      </td>
      <td class="col-status">{status_badge(row['status'])}{claimed}{finance_note}</td>
      <td class="actions admin-actions col-actions">
        <details class="fold">
          <summary>编辑字段</summary>
          <div class="fold-body">
            <form method="post" action="/admin/payments/{row['id']}/edit">
              {finance_hidden(actor)}
              <div class="edit-row">
                <div style="width:115px"><label>日期</label><input name="received_date" value="{esc(row['received_date'])}"></div>
                <div style="width:110px"><label>金额</label><input name="amount" value="{money(row['amount_cents']).replace(',', '')}"></div>
              </div>
              <div class="field" style="margin-top:10px"><label>到款公司</label><input name="receiver_company" value="{esc(row['receiver_company'])}"></div>
              <div class="field" style="margin-top:10px"><label>付款方</label><input name="payer_name" value="{esc(row['payer_name'])}"></div>
              <div class="field"><label>备注</label><input name="bank_note" value="{esc(row['bank_note'])}"></div>
              <button class="secondary" type="submit">保存字段</button>
            </form>
          </div>
        </details>
        <details class="fold">
          <summary>处理状态</summary>
          <div class="fold-body">
            <form method="post" action="/admin/payments/{row['id']}/resolve">
              {finance_hidden(actor)}
              <div class="field"><label>状态</label><select name="status">
                {status_options(row['status'])}
              </select></div>
              <div class="field"><label>分配部门</label>{department_select("department", row["claimed_department"], class_name="cs-dept")}</div>
              <div class="field"><label>中心 / 小组</label>{team_select("team", row["claimed_department"] or "", row["claimed_team"] or "")}</div>
              <div class="field"><label>管理备注</label><input name="finance_note" value="{esc(row['finance_note'])}"></div>
              <button type="submit">更新状态</button>
            </form>
          </div>
        </details>
        {reject_ui}
      </td>
    </tr>
    """


def status_options(current: str) -> str:
    items = [
        ("pending", "待认领"),
        ("partial_claiming", "部分认领中"),
        ("claimed", "已认领"),
        ("rejected", "已驳回"),
        ("closed", "已关闭"),
    ]
    return "".join(
        f'<option value="{value}" {"selected" if value == current else ""}>{label}</option>'
        for value, label in items
    )


@app.post("/admin/import")
def admin_import(
    request: Request,
    source_name: str = Form(...),
    table_text: str = Form(""),
    attachment: Optional[UploadFile] = File(None),
) -> RedirectResponse:
    # 敏感操作：用会话实时角色鉴权，不信任表单字段，杜绝越权
    rate_limit(request, "import")
    actor = actor_from_request(request)
    require_admin(actor)
    source_ref = save_attachment(attachment) or source_name
    rows = read_table(table_text)
    if not rows and source_ref.startswith("uploads/"):
        source_path = UPLOAD_DIR / Path(source_ref).name
        suffix = source_path.suffix.lower()
        if suffix == ".pdf":
            rows, reason = rows_from_pdf(source_path)
            if not rows:
                raise HTTPException(status_code=400, detail=reason)
        elif suffix in {".xls", ".xlsx"}:
            rows = rows_from_excel(source_path)
    if not rows:
        raise HTTPException(status_code=400, detail="未识别到任何流水。请上传 PDF / Excel 文件。")
    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO import_batches
                (source_name, created_at, created_by, raw_count, imported_count, skipped_count)
            VALUES (?, ?, ?, ?, 0, 0)
            """,
            (source_name, now_text(), actor["name"], len(rows)),
        )
        batch_id = cur.lastrowid
        imported = 0
        skipped = 0
        for item in rows:
            amount_cents = parse_amount(item.get("amount"))
            if duplicate_exists(conn, item):
                skipped += 1
                continue
            received_date = parse_date(item.get("received_date"))
            payer_name = item.get("payer_name", "").strip()
            confidence = 1.0
            missing = []
            if not received_date:
                missing.append("到款日期")
            if not payer_name:
                missing.append("付款方名称")
            if amount_cents <= 0:
                missing.append("到款金额")
            if missing:
                confidence = 0.4
            conn.execute(
                """
                INSERT INTO payments
                    (batch_id, imported_at, received_date, received_time, payer_name, amount_cents,
                     bank_note, receiver_company, receiver_account, serial_no, source_ref, confidence, status, finance_note, raw_json)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'draft', ?, ?)
                """,
                (
                    batch_id,
                    now_text(),
                    received_date,
                    item.get("received_time", "").strip(),
                    payer_name,
                    max(amount_cents, 0),
                    item.get("bank_note", "").strip(),
                    item.get("receiver_company", "").strip(),
                    item.get("receiver_account", "").strip(),
                    item.get("serial_no", "").strip(),
                    source_ref,
                    confidence,
                    f"字段缺失：{', '.join(missing)}" if missing else "",
                    json.dumps(item, ensure_ascii=False),
                ),
            )
            imported += 1
        conn.execute(
            "UPDATE import_batches SET imported_count = ?, skipped_count = ? WHERE id = ?",
            (imported, skipped, batch_id),
        )
        audit(
            conn,
            actor,
            "import_batch",
            None,
            {"batch_id": batch_id, "raw": len(rows), "imported": imported, "skipped": skipped, "source_ref": source_ref},
            request,
        )
    return RedirectResponse("/admin", status_code=303)


@app.post("/admin/batches/{batch_id}/confirm")
def confirm_batch(
    request: Request,
    batch_id: int,
) -> RedirectResponse:
    # 敏感操作：用会话实时角色鉴权，不信任表单字段，杜绝越权
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        confirm_import_batch(conn, actor, batch_id, request)
    return RedirectResponse("/admin", status_code=303)


def confirm_import_batch(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    batch_id: int,
    request: Optional[Request] = None,
) -> dict[str, Any]:
    rows = conn.execute(
        "SELECT id, amount_cents FROM payments WHERE batch_id = ? AND status = 'draft'",
        (batch_id,),
    ).fetchall()
    total_cents = sum(int(row["amount_cents"] or 0) for row in rows)
    conn.execute(
        "UPDATE payments SET status = 'pending', confirmed_at = ? WHERE batch_id = ? AND status = 'draft'",
        (now_text(), batch_id),
    )
    conn.execute("UPDATE import_batches SET status = 'confirmed' WHERE id = ?", (batch_id,))

    notified = False
    if rows:
        notified = feishu_send_chat_text(FEISHU_NOTIFY_CHAT_ID, build_batch_confirm_message())

    detail = {
        "batch_id": batch_id,
        "count": len(rows),
        "amount_cents": total_cents,
        "notified": notified,
    }
    audit(conn, actor, "confirm_batch", None, detail, request)
    return detail


@app.post("/admin/batches/{batch_id}/cancel")
def cancel_batch(
    request: Request,
    batch_id: int,
) -> RedirectResponse:
    # 敏感操作：仅允许取消尚未确认入池的 draft 批次，避免误删已开放认领的数据
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        batch = conn.execute("SELECT * FROM import_batches WHERE id = ?", (batch_id,)).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="批次不存在")
        if batch["status"] != "draft":
            raise HTTPException(status_code=409, detail="只有待确认批次可以取消")
        locked = conn.execute(
            "SELECT COUNT(*) AS count FROM payments WHERE batch_id = ? AND status != 'draft'",
            (batch_id,),
        ).fetchone()["count"]
        if locked:
            raise HTTPException(status_code=409, detail="该批次已有非待确认流水，不能取消")
        rows = conn.execute("SELECT id FROM payments WHERE batch_id = ? AND status = 'draft'", (batch_id,)).fetchall()
        conn.execute("DELETE FROM payments WHERE batch_id = ? AND status = 'draft'", (batch_id,))
        conn.execute("UPDATE import_batches SET status = 'canceled' WHERE id = ?", (batch_id,))
        audit(conn, actor, "cancel_batch", None, {"batch_id": batch_id, "count": len(rows)}, request)
    return RedirectResponse("/admin", status_code=303)


def close_payments_bulk(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    payment_ids: list[int],
    request: Optional[Request] = None,
) -> dict[str, Any]:
    unique_ids: list[int] = []
    seen: set[int] = set()
    for payment_id in payment_ids:
        if payment_id <= 0 or payment_id in seen:
            continue
        seen.add(payment_id)
        unique_ids.append(payment_id)
    if not unique_ids:
        raise HTTPException(status_code=400, detail="请先选择要关闭的流水")

    placeholders = ",".join("?" for _ in unique_ids)
    rows = conn.execute(
        f"SELECT id, status FROM payments WHERE id IN ({placeholders})",
        unique_ids,
    ).fetchall()
    status_by_id = {row["id"]: row["status"] for row in rows}
    found_ids = set(status_by_id)
    close_ids = [
        payment_id
        for payment_id in unique_ids
        if payment_id in status_by_id and status_by_id[payment_id] != "closed"
    ]
    skipped_closed = [
        payment_id
        for payment_id in unique_ids
        if status_by_id.get(payment_id) == "closed"
    ]
    missing_ids = [payment_id for payment_id in unique_ids if payment_id not in found_ids]
    closed_at = now_text()

    if close_ids:
        close_placeholders = ",".join("?" for _ in close_ids)
        conn.execute(
            f"""
            UPDATE payments
            SET status = 'closed',
                closed_at = COALESCE(NULLIF(closed_at, ''), ?)
            WHERE id IN ({close_placeholders})
              AND status != 'closed'
            """,
            [closed_at, *close_ids],
        )

    detail = {
        "requested_ids": unique_ids,
        "closed_ids": close_ids,
        "skipped_closed_ids": skipped_closed,
        "missing_ids": missing_ids,
        "count": len(close_ids),
        "skipped": len(skipped_closed) + len(missing_ids),
    }
    audit(conn, actor, "bulk_close_payments", None, detail, request)
    return detail


@app.post("/admin/payments/bulk-close")
def bulk_close_payments_route(
    request: Request,
    payment_ids: Optional[list[int]] = Form(None),
) -> RedirectResponse:
    # 敏感操作：批量关闭只改状态，不删除流水、附件或认领记录
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        close_payments_bulk(conn, actor, payment_ids or [], request)
    return RedirectResponse("/admin", status_code=303)


def refresh_payment_claim_status(conn: sqlite3.Connection, payment_id: int) -> None:
    row = conn.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
    if not row:
        return
    totals = claim_totals(conn, payment_id)
    active_claims = conn.execute(
        """
        SELECT *
        FROM claims
        WHERE payment_id = ? AND status IN ('pending', 'accepted')
        ORDER BY id
        """,
        (payment_id,),
    ).fetchall()
    if totals["active"] <= 0:
        conn.execute(
            """
            UPDATE payments
            SET status = 'pending',
                claimed_department = NULL,
                claimed_team = NULL,
                claimed_by = NULL,
                claimed_by_name = NULL,
                claimed_at = NULL,
                customer_project = NULL,
                claim_note = NULL,
                finance_note = NULL
            WHERE id = ?
            """,
            (payment_id,),
        )
    elif len(active_claims) == 1:
        claim = active_claims[0]
        claim_keys = set(claim.keys())
        claimed_at = claim["created_at"] if "created_at" in claim_keys else now_text()
        finance_note = "部分认领金额已覆盖整笔到款" if totals["active"] >= row["amount_cents"] else "存在部分认领，尚未认满"
        conn.execute(
            """
            UPDATE payments
            SET status = ?,
                claimed_department = ?,
                claimed_team = ?,
                claimed_by = ?,
                claimed_by_name = ?,
                claimed_at = ?,
                customer_project = ?,
                claim_note = ?,
                finance_note = ?
            WHERE id = ?
            """,
            (
                "claimed" if totals["active"] >= row["amount_cents"] else "partial_claiming",
                claim["department"] if "department" in claim_keys else row["claimed_department"],
                claim["team"] if "team" in claim_keys else row["claimed_team"],
                claim["actor_id"] if "actor_id" in claim_keys else row["claimed_by"],
                claim["actor_name"] if "actor_name" in claim_keys else row["claimed_by_name"],
                claimed_at,
                claim["customer_project"] if "customer_project" in claim_keys else row["customer_project"],
                claim["note"] if "note" in claim_keys else row["claim_note"],
                finance_note,
                payment_id,
            ),
        )
    else:
        conn.execute(
            """
            UPDATE payments
            SET status = ?,
                claimed_department = '多部门分摊',
                claimed_team = NULL,
                claimed_by = NULL,
                claimed_by_name = NULL,
                claimed_at = ?,
                customer_project = NULL,
                claim_note = NULL,
                finance_note = ?
            WHERE id = ?
            """,
            (
                "claimed" if totals["active"] >= row["amount_cents"] else "partial_claiming",
                now_text(),
                "部分认领金额已覆盖整笔到款" if totals["active"] >= row["amount_cents"] else "存在部分认领，尚未认满",
                payment_id,
            ),
        )


def reset_payment_to_initial_pending(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    payment_id: int,
    reason: str = "",
    action: str = "reset_payment_to_pending",
    request: Optional[Request] = None,
    detail_extra: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    payment = conn.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
    if not payment:
        raise HTTPException(status_code=404, detail="到款记录不存在")
    reason = reason.strip()
    active_claims = conn.execute(
        """
        SELECT id FROM claims
        WHERE payment_id = ? AND status IN ('pending', 'accepted')
        ORDER BY id
        """,
        (payment_id,),
    ).fetchall()
    reset_claim_ids = [int(row["id"]) for row in active_claims]
    if reset_claim_ids:
        placeholders = ",".join("?" for _ in reset_claim_ids)
        conn.execute(
            f"UPDATE claims SET status = 'rejected' WHERE id IN ({placeholders})",
            reset_claim_ids,
        )
    finance_note = f"驳回退回：{reason}" if reason else "驳回退回"
    conn.execute(
        """
        UPDATE payments
        SET status = 'pending',
            claimed_department = NULL,
            claimed_team = NULL,
            claimed_by = NULL,
            claimed_by_name = NULL,
            claimed_at = NULL,
            customer_project = NULL,
            claim_note = NULL,
            finance_note = ?,
            closed_at = NULL
        WHERE id = ?
        """,
        (finance_note, payment_id),
    )
    detail = {
        "payment_status": "pending",
        "reset_claim_ids": reset_claim_ids,
        "reason": reason,
    }
    if detail_extra:
        detail.update(detail_extra)
    audit(conn, actor, action, payment_id, detail, request)
    return detail


def repair_rejected_payments_to_pending(conn: sqlite3.Connection) -> dict[str, Any]:
    rows = conn.execute(
        """
        SELECT id, finance_note FROM payments
        WHERE status = 'rejected'
        ORDER BY id
        """
    ).fetchall()
    repaired_ids: list[int] = []
    reset_claim_ids: list[int] = []
    for row in rows:
        payment_id = int(row["id"])
        active_claims = conn.execute(
            """
            SELECT id FROM claims
            WHERE payment_id = ? AND status IN ('pending', 'accepted')
            ORDER BY id
            """,
            (payment_id,),
        ).fetchall()
        claim_ids = [int(claim["id"]) for claim in active_claims]
        if claim_ids:
            placeholders = ",".join("?" for _ in claim_ids)
            conn.execute(
                f"UPDATE claims SET status = 'rejected' WHERE id IN ({placeholders})",
                claim_ids,
            )
            reset_claim_ids.extend(claim_ids)
        finance_note = (row["finance_note"] or "").strip() or "驳回退回"
        conn.execute(
            """
            UPDATE payments
            SET status = 'pending',
                claimed_department = NULL,
                claimed_team = NULL,
                claimed_by = NULL,
                claimed_by_name = NULL,
                claimed_at = NULL,
                customer_project = NULL,
                claim_note = NULL,
                finance_note = ?,
                closed_at = NULL
            WHERE id = ?
            """,
            (finance_note, payment_id),
        )
        audit(
            conn,
            {"id": "system", "name": "系统修复", "role": "system"},
            "repair_rejected_payment_to_pending",
            payment_id,
            {"payment_status": "pending", "reset_claim_ids": claim_ids},
        )
        repaired_ids.append(payment_id)
    return {"payment_ids": repaired_ids, "reset_claim_ids": reset_claim_ids}


def repair_pending_claims_to_accepted(conn: sqlite3.Connection) -> dict[str, Any]:
    rows = conn.execute(
        """
        SELECT payment_id, GROUP_CONCAT(id) AS claim_ids, COALESCE(SUM(amount_cents), 0) AS amount_cents
        FROM claims
        WHERE status = 'pending'
        GROUP BY payment_id
        ORDER BY payment_id
        """
    ).fetchall()
    repaired_payment_ids: list[int] = []
    accepted_claim_ids: list[int] = []
    for row in rows:
        payment_id = int(row["payment_id"])
        claim_ids = [int(value) for value in str(row["claim_ids"] or "").split(",") if value]
        if not claim_ids:
            continue
        placeholders = ",".join("?" for _ in claim_ids)
        conn.execute(
            f"UPDATE claims SET status = 'accepted' WHERE id IN ({placeholders})",
            claim_ids,
        )
        refresh_payment_claim_status(conn, payment_id)
        audit(
            conn,
            {"id": "system", "name": "系统修复", "role": "system"},
            "repair_pending_claims_to_accepted",
            payment_id,
            {"claim_ids": claim_ids, "amount_cents": int(row["amount_cents"] or 0)},
        )
        repaired_payment_ids.append(payment_id)
        accepted_claim_ids.extend(claim_ids)
    return {"payment_ids": repaired_payment_ids, "claim_ids": accepted_claim_ids}


def cancel_my_claim(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    claim_id: int,
    request: Optional[Request] = None,
) -> dict[str, Any]:
    claim = conn.execute("SELECT * FROM claims WHERE id = ?", (claim_id,)).fetchone()
    if not claim:
        raise HTTPException(status_code=404, detail="认领记录不存在")
    if claim["actor_id"] != actor["id"]:
        raise HTTPException(status_code=403, detail="只能取消自己的认领")
    if claim["status"] not in {"pending", "accepted"}:
        raise HTTPException(status_code=409, detail="这条认领当前不能取消")

    payment_id = int(claim["payment_id"])
    payment_before = conn.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
    conn.execute("UPDATE claims SET status = 'canceled' WHERE id = ?", (claim_id,))
    refresh_payment_claim_status(conn, payment_id)
    payment = conn.execute("SELECT status FROM payments WHERE id = ?", (payment_id,)).fetchone()
    notify_detail = notify_admins_claim_canceled(
        conn,
        payment_before or payment,
        claim,
        actor,
        payment["status"] if payment else "",
    )
    detail = {
        "claim_id": claim_id,
        "payment_id": payment_id,
        "previous_status": claim["status"],
        "payment_status": payment["status"] if payment else "",
        "amount_cents": claim["amount_cents"],
        **notify_detail,
    }
    audit(conn, actor, "cancel_my_claim", payment_id, detail, request)
    return detail


def accept_payment_claims(
    conn: sqlite3.Connection,
    actor: dict[str, str],
    payment_id: int,
    request: Optional[Request] = None,
) -> dict[str, Any]:
    payment = conn.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
    if not payment:
        raise HTTPException(status_code=404, detail="到款记录不存在")
    pending_claims = conn.execute(
        "SELECT * FROM claims WHERE payment_id = ? AND status = 'pending' ORDER BY id",
        (payment_id,),
    ).fetchall()
    if not pending_claims:
        raise HTTPException(status_code=409, detail="这笔款没有待处理认领")
    pending_amount = sum(int(claim["amount_cents"] or 0) for claim in pending_claims)
    accepted = claim_totals(conn, payment_id)["accepted"]
    if accepted + pending_amount > payment["amount_cents"]:
        raise HTTPException(status_code=409, detail="处理后金额会超过到款金额")
    claim_ids = [claim["id"] for claim in pending_claims]
    placeholders = ",".join("?" for _ in claim_ids)
    conn.execute(
        f"UPDATE claims SET status = 'accepted' WHERE id IN ({placeholders})",
        claim_ids,
    )
    refresh_payment_claim_status(conn, payment_id)
    detail = {
        "claim_ids": claim_ids,
        "count": len(claim_ids),
        "amount_cents": pending_amount,
    }
    audit(conn, actor, "accept_payment_claims", payment_id, detail, request)
    return detail


@app.post("/admin/payments/{payment_id}/confirm-claims")
def accept_payment_claims_route(
    request: Request,
    payment_id: int,
) -> RedirectResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        accept_payment_claims(conn, actor, payment_id, request)
    return RedirectResponse("/admin", status_code=303)


@app.post("/admin/claims/{claim_id}/accept")
def accept_claim(
    request: Request,
    claim_id: int,
) -> RedirectResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        claim = conn.execute("SELECT * FROM claims WHERE id = ?", (claim_id,)).fetchone()
        if not claim:
            raise HTTPException(status_code=404, detail="认领记录不存在")
        if claim["status"] != "pending":
            raise HTTPException(status_code=409, detail="只有待处理认领可以处理")
        payment = conn.execute("SELECT * FROM payments WHERE id = ?", (claim["payment_id"],)).fetchone()
        if not payment:
            raise HTTPException(status_code=404, detail="到款记录不存在")
        accepted = claim_totals(conn, claim["payment_id"])["accepted"]
        if accepted + claim["amount_cents"] > payment["amount_cents"]:
            raise HTTPException(status_code=409, detail="处理后金额会超过到款金额")
        conn.execute("UPDATE claims SET status = 'accepted' WHERE id = ?", (claim_id,))
        refresh_payment_claim_status(conn, claim["payment_id"])
        audit(
            conn,
            actor,
            "accept_claim",
            claim["payment_id"],
            {"claim_id": claim_id, "amount_cents": claim["amount_cents"]},
            request,
        )
    return RedirectResponse("/admin", status_code=303)


@app.post("/admin/claims/{claim_id}/reject")
def reject_claim(
    request: Request,
    claim_id: int,
) -> RedirectResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        claim = conn.execute("SELECT * FROM claims WHERE id = ?", (claim_id,)).fetchone()
        if not claim:
            raise HTTPException(status_code=404, detail="认领记录不存在")
        if claim["status"] != "pending":
            raise HTTPException(status_code=409, detail="只有待处理认领可以驳回")
        payment = conn.execute("SELECT * FROM payments WHERE id = ?", (claim["payment_id"],)).fetchone()
        notified = False
        if payment:
            notified = feishu_send_text(claim["actor_id"], build_claim_reject_message(payment, claim))
        conn.execute("UPDATE claims SET status = 'rejected' WHERE id = ?", (claim_id,))
        refresh_payment_claim_status(conn, claim["payment_id"])
        audit(
            conn,
            actor,
            "reject_claim",
            claim["payment_id"],
            {"claim_id": claim_id, "amount_cents": claim["amount_cents"], "notified": notified},
            request,
        )
    return RedirectResponse("/admin", status_code=303)


@app.post("/admin/payments/{payment_id}/edit")
def edit_payment(
    request: Request,
    payment_id: int,
    received_date: str = Form(""),
    amount: str = Form(""),
    receiver_company: str = Form(""),
    payer_name: str = Form(""),
    bank_note: str = Form(""),
) -> RedirectResponse:
    # 敏感操作：用会话实时角色鉴权，不信任表单字段，杜绝越权
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE payments
            SET received_date = ?, amount_cents = ?, receiver_company = ?, payer_name = ?, bank_note = ?
            WHERE id = ?
            """,
            (
                parse_date(received_date),
                parse_amount(amount),
                receiver_company.strip(),
                payer_name.strip(),
                bank_note.strip(),
                payment_id,
            ),
        )
        audit(
            conn,
            actor,
            "edit_payment",
            payment_id,
            {"received_date": received_date, "amount": amount, "receiver_company": receiver_company.strip()},
            request,
        )
    return RedirectResponse("/admin", status_code=303)


@app.post("/admin/payments/{payment_id}/resolve")
def resolve_payment(
    request: Request,
    payment_id: int,
    status: str = Form(...),
    department: str = Form(""),
    team: str = Form(""),
    finance_note: str = Form(""),
) -> RedirectResponse:
    # 敏感操作：用会话实时角色鉴权，不信任表单字段，杜绝越权
    actor = actor_from_request(request)
    require_admin(actor)
    if status not in {"pending", "partial_claiming", "claimed", "pending_confirm", "rejected", "closed"}:
        raise HTTPException(status_code=400, detail="状态不合法")
    department = department.strip()
    team = team.strip()
    if department and department not in DEPARTMENTS:
        raise HTTPException(status_code=400, detail="请选择标准部门")
    if status == "claimed" and not department:
        raise HTTPException(status_code=400, detail="标记已认领时必须选择部门")
    with get_conn() as conn:
        if team:
            row = conn.execute(
                "SELECT claimed_department FROM payments WHERE id = ?", (payment_id,)
            ).fetchone()
            effective_dept = department or (row["claimed_department"] if row else "")
            if team not in CATALOG.get(effective_dept or "", {}):
                raise HTTPException(status_code=400, detail="请选择该部门下的中心/小组")
        if status == "pending":
            conn.execute(
                """
                UPDATE payments
                SET status = 'pending',
                    claimed_department = NULL,
                    claimed_team = NULL,
                    claimed_by = NULL,
                    claimed_by_name = NULL,
                    claimed_at = NULL,
                    closed_at = NULL,
                    finance_note = ?
                WHERE id = ?
                """,
                (finance_note, payment_id),
            )
        elif status == "rejected":
            reset_payment_to_initial_pending(
                conn,
                actor,
                payment_id,
                finance_note,
                action="resolve_payment_reject",
                request=request,
            )
            return RedirectResponse("/admin", status_code=303)
        else:
            conn.execute(
                """
                UPDATE payments
                SET status = ?,
                    claimed_department = COALESCE(NULLIF(?, ''), claimed_department),
                    claimed_team = COALESCE(NULLIF(?, ''), claimed_team),
                    closed_at = CASE
                        WHEN ? = 'closed' THEN COALESCE(NULLIF(closed_at, ''), ?)
                        ELSE NULL
                    END,
                    finance_note = ?
                WHERE id = ?
                """,
                (status, department, team, status, now_text(), finance_note, payment_id),
            )
        audit(conn, actor, "resolve_payment", payment_id, {"status": status, "department": department, "team": team}, request)
    return RedirectResponse("/admin", status_code=303)


@app.post("/admin/payments/{payment_id}/reject")
def reject_payment(
    request: Request,
    payment_id: int,
    reason: str = Form(""),
) -> RedirectResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    reason = reason.strip()
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        claimed_by = row["claimed_by"]
        # 先通知原认领人（飞书单聊），失败不影响退回
        notified = False
        if claimed_by:
            notified = feishu_send_text(claimed_by, build_payment_reject_message(row, reason))
        detail = reset_payment_to_initial_pending(
            conn,
            actor,
            payment_id,
            reason,
            action="reject_payment",
            request=request,
            detail_extra={"prev_claimed_by": claimed_by, "notified": notified},
        )
    return RedirectResponse("/admin", status_code=303)


@app.get("/admin/payments/{payment_id}/logs", response_class=HTMLResponse)
def payment_logs(request: Request, payment_id: int) -> HTMLResponse:
    actor = actor_from_request(request)
    require_admin(actor)
    with get_conn() as conn:
        logs = conn.execute("SELECT * FROM audit_logs WHERE payment_id = ? ORDER BY id DESC", (payment_id,)).fetchall()
    rows = "".join(
        f"<tr><td class='nowrap'>{esc(row['at'])}</td><td class='nowrap'>{esc(row['actor_name'])}</td><td class='nowrap'>{esc(row['action'])}</td><td><span class='code'>{esc(row['detail_json'])}</span></td></tr>"
        for row in logs
    )
    return page(
        f"记录 #{payment_id} 操作日志",
        f"<div class='table-wrap'><table><thead><tr><th>时间</th><th>用户</th><th>动作</th><th>详情</th></tr></thead><tbody>{rows or '<tr><td colspan=4 class=empty>暂无日志</td></tr>'}</tbody></table></div>",
        active="admin",
        actor=actor,
    )
