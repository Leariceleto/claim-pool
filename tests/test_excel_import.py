import importlib
import json
import sqlite3
import sys
import tempfile
import types
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook


def install_fastapi_stub() -> None:
    try:
        import fastapi  # noqa: F401
        return
    except ModuleNotFoundError:
        pass

    fastapi_module = types.ModuleType("fastapi")
    responses_module = types.ModuleType("fastapi.responses")

    class HTTPException(Exception):
        def __init__(self, status_code: int, detail: str):
            super().__init__(detail)
            self.status_code = status_code
            self.detail = detail

    class FastAPI:
        def __init__(self, *args, **kwargs):
            pass

        def get(self, *args, **kwargs):
            return lambda func: func

        def post(self, *args, **kwargs):
            return lambda func: func

        def middleware(self, *args, **kwargs):
            return lambda func: func

        def on_event(self, *args, **kwargs):
            return lambda func: func

    def form_or_file(default=None, *args, **kwargs):
        return default

    class Request:
        pass

    class UploadFile:
        pass

    class Response:
        pass

    class HTMLResponse(Response):
        pass

    class RedirectResponse(Response):
        pass

    class FileResponse(Response):
        pass

    fastapi_module.FastAPI = FastAPI
    fastapi_module.File = form_or_file
    fastapi_module.Form = form_or_file
    fastapi_module.HTTPException = HTTPException
    fastapi_module.Request = Request
    fastapi_module.UploadFile = UploadFile
    responses_module.FileResponse = FileResponse
    responses_module.HTMLResponse = HTMLResponse
    responses_module.RedirectResponse = RedirectResponse
    responses_module.Response = Response
    sys.modules["fastapi"] = fastapi_module
    sys.modules["fastapi.responses"] = responses_module


class ExcelImportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        install_fastapi_stub()
        cls.app = importlib.import_module("app")

    def test_parse_date_supports_compact_yyyymmdd(self) -> None:
        self.assertEqual(self.app.parse_date("20260707"), "2026-07-07")
        self.assertEqual(self.app.parse_date("2026-07-07"), "2026-07-07")

    def test_ooxml_workbook_with_xls_suffix_imports(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["到款日期", "付款方名称", "到款金额", "银行备注", "流水号"])
        sheet.append(["2026-06-25", "北京测试教育科技有限公司", 1200, "银行导出", "MX001"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "bank_export.xls"
            workbook.save(path)
            rows = self.app.rows_from_excel(path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["received_date"], "2026-06-25")
        self.assertEqual(rows[0]["payer_name"], "北京测试教育科技有限公司")
        self.assertEqual(rows[0]["amount"], "1200")
        self.assertEqual(rows[0]["serial_no"], "MX001")

    def test_bank_transaction_headers_import(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["交易时间", "终端号", "交易类型", "卡号", "交易金额", "流水号", "银商订单号", "付款附言"])
        sheet.append([
            "2026-06-24 14:13:31",
            "50351488",
            "消费",
            "000000******0000",
            "999",
            "304143",
            "26062457713041431031801750",
            "上海市虹口区特教指导中心 骨干教师",
        ])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "bank_export.xls"
            workbook.save(path)
            rows = self.app.rows_from_excel(path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["received_date"], "2026-06-24 14:13:31")
        self.assertEqual(rows[0]["received_time"], "14:13:31")
        self.assertEqual(rows[0]["payer_name"], "上海市虹口区特教指导中心 骨干教师")
        self.assertEqual(rows[0]["bank_note"], "上海市虹口区特教指导中心 骨干教师")
        self.assertEqual(rows[0]["amount"], "999")
        self.assertEqual(rows[0]["serial_no"], "304143")

    def test_income_header_imports_as_amount(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["到款日期", "付款方名称", "收入", "银行备注"])
        sheet.append(["2026-06-25", "广州测试学校", 1980, "报名费"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "income_header.xlsx"
            workbook.save(path)
            rows = self.app.rows_from_excel(path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["amount"], "1980")

    def test_receiver_company_header_imports(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["到款日期", "付款方名称", "到款公司", "到款金额", "银行备注"])
        sheet.append(["2026-06-25", "深圳测试学校", "北京蒲公英教育科技有限公司", 3980, "参会费"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "receiver_company.xlsx"
            workbook.save(path)
            rows = self.app.rows_from_excel(path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["receiver_company"], "北京蒲公英教育科技有限公司")

    def test_duplicate_exists_ignores_closed_payments(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                received_date TEXT,
                payer_name TEXT,
                amount_cents INTEGER NOT NULL,
                bank_note TEXT,
                serial_no TEXT,
                status TEXT NOT NULL
            );
            """
        )
        item_with_serial = {
            "received_date": "2026-07-01",
            "payer_name": "北京测试学校",
            "amount": "1000",
            "bank_note": "报名费",
            "serial_no": "SERIAL-001",
        }
        conn.execute(
            """
            INSERT INTO payments (id, received_date, payer_name, amount_cents, bank_note, serial_no, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (1, "2026-07-01", "北京测试学校", 100000, "报名费", "SERIAL-001", "closed"),
        )

        self.assertFalse(self.app.duplicate_exists(conn, item_with_serial, 100000))
        conn.execute("UPDATE payments SET status = 'pending' WHERE id = 1")
        self.assertTrue(self.app.duplicate_exists(conn, item_with_serial, 100000))

        conn.execute("DELETE FROM payments")
        item_without_serial = {
            "received_date": "2026-07-01",
            "payer_name": "北京测试学校",
            "amount": "1000",
            "bank_note": "报名费",
            "serial_no": "",
        }
        conn.execute(
            """
            INSERT INTO payments (id, received_date, payer_name, amount_cents, bank_note, serial_no, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (2, "2026-07-01", "北京测试学校", 100000, "报名费", "", "closed"),
        )

        self.assertFalse(self.app.duplicate_exists(conn, item_without_serial, 100000))
        conn.execute("UPDATE payments SET status = 'pending' WHERE id = 2")
        self.assertTrue(self.app.duplicate_exists(conn, item_without_serial, 100000))

    def test_bulk_close_payments_skips_closed_and_logs(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                status TEXT NOT NULL,
                closed_at TEXT
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.executemany(
            "INSERT INTO payments (id, status, closed_at) VALUES (?, ?, ?)",
            [
                (1, "pending", None),
                (2, "closed", "2026-06-20 10:00:00"),
                (3, "draft", ""),
            ],
        )
        actor = {"id": "finance", "name": "财务管理员", "role": "finance"}

        detail = self.app.close_payments_bulk(conn, actor, [1, 2, 2, 3, 999])

        rows = conn.execute("SELECT id, status, closed_at FROM payments ORDER BY id").fetchall()
        self.assertEqual([row["status"] for row in rows], ["closed", "closed", "closed"])
        self.assertTrue(rows[0]["closed_at"])
        self.assertEqual(rows[1]["closed_at"], "2026-06-20 10:00:00")
        self.assertTrue(rows[2]["closed_at"])
        self.assertEqual(detail["closed_ids"], [1, 3])
        self.assertEqual(detail["skipped_closed_ids"], [2])
        self.assertEqual(detail["missing_ids"], [999])

        audit_row = conn.execute("SELECT action, detail_json FROM audit_logs").fetchone()
        self.assertEqual(audit_row["action"], "bulk_close_payments")
        audit_detail = json.loads(audit_row["detail_json"])
        self.assertEqual(audit_detail["count"], 2)

    def test_confirm_import_batch_sends_group_notification(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                batch_id INTEGER NOT NULL,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                confirmed_at TEXT
            );
            CREATE TABLE import_batches (
                id INTEGER PRIMARY KEY,
                status TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.execute("INSERT INTO import_batches (id, status) VALUES (?, ?)", (7, "draft"))
        conn.executemany(
            "INSERT INTO payments (id, batch_id, amount_cents, status) VALUES (?, ?, ?, ?)",
            [(1, 7, 12000, "draft"), (2, 7, 34567, "draft"), (3, 7, 999, "pending")],
        )
        sent: list[tuple[str, str]] = []
        old_chat_id = self.app.FEISHU_NOTIFY_CHAT_ID
        old_send = self.app.feishu_send_chat_text
        self.app.FEISHU_NOTIFY_CHAT_ID = "oc_test_chat"
        self.app.feishu_send_chat_text = lambda chat_id, text: sent.append((chat_id, text)) or True
        try:
            detail = self.app.confirm_import_batch(
                conn,
                {"id": "finance", "name": "财务管理员", "role": "admin"},
                7,
            )
        finally:
            self.app.FEISHU_NOTIFY_CHAT_ID = old_chat_id
            self.app.feishu_send_chat_text = old_send

        statuses = conn.execute("SELECT id, status FROM payments ORDER BY id").fetchall()
        batch = conn.execute("SELECT status FROM import_batches WHERE id = 7").fetchone()
        audit_row = conn.execute("SELECT action, detail_json FROM audit_logs").fetchone()

        self.assertEqual(detail["count"], 2)
        self.assertEqual(detail["amount_cents"], 46567)
        self.assertTrue(detail["notified"])
        self.assertEqual([(row["id"], row["status"]) for row in statuses], [(1, "pending"), (2, "pending"), (3, "pending")])
        self.assertEqual(batch["status"], "confirmed")
        self.assertEqual(sent[0][0], "oc_test_chat")
        self.assertIn("共 2 笔，合计 ¥ 465.67", sent[0][1])
        self.assertEqual(audit_row["action"], "confirm_batch")
        self.assertTrue(json.loads(audit_row["detail_json"])["notified"])

    def test_confirm_import_batch_continues_when_group_notification_is_unconfigured(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                batch_id INTEGER NOT NULL,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                confirmed_at TEXT
            );
            CREATE TABLE import_batches (
                id INTEGER PRIMARY KEY,
                status TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.execute("INSERT INTO import_batches (id, status) VALUES (?, ?)", (8, "draft"))
        conn.execute(
            "INSERT INTO payments (id, batch_id, amount_cents, status) VALUES (?, ?, ?, ?)",
            (1, 8, 10000, "draft"),
        )
        old_chat_id = self.app.FEISHU_NOTIFY_CHAT_ID
        self.app.FEISHU_NOTIFY_CHAT_ID = ""
        try:
            detail = self.app.confirm_import_batch(
                conn,
                {"id": "finance", "name": "财务管理员", "role": "admin"},
                8,
            )
        finally:
            self.app.FEISHU_NOTIFY_CHAT_ID = old_chat_id

        payment = conn.execute("SELECT status FROM payments WHERE id = 1").fetchone()
        self.assertEqual(payment["status"], "pending")
        self.assertFalse(detail["notified"])

    def test_reject_notification_messages_include_reason_without_links(self) -> None:
        payment_message = self.app.build_payment_reject_message(
            {
                "payer_name": "测试客户",
                "received_date": "2026-07-03",
                "claimed_department": "年会事业部",
                "claimed_team": "创新中心",
                "customer_project": "测试项目",
            },
            "部门归属不对",
        )
        claim_message = self.app.build_claim_reject_message(
            {"payer_name": "测试客户", "received_date": "2026-07-03"},
            {
                "amount_cents": 12345,
                "department": "年会事业部",
                "team": "创新中心",
                "customer_project": "测试项目",
            },
        )

        self.assertIn("驳回原因：部门归属不对", payment_message)
        self.assertIn("驳回原因：未填写", claim_message)
        self.assertNotIn("重新认领", payment_message)
        self.assertNotIn("http", payment_message)
        self.assertNotIn("重新认领", claim_message)
        self.assertNotIn("http", claim_message)

    def test_accept_payment_claims_confirms_pending_claims_and_refreshes_payment(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY,
                payment_id INTEGER NOT NULL,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.execute(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note)
            VALUES (?, ?, ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)
            """,
            (1, 100668, "pending_confirm"),
        )
        conn.executemany(
            "INSERT INTO claims (id, payment_id, amount_cents, status) VALUES (?, ?, ?, ?)",
            [(1, 1, 100668, "pending")],
        )
        actor = {"id": "finance", "name": "财务管理员", "role": "admin"}

        detail = self.app.accept_payment_claims(conn, actor, 1)

        payment = conn.execute("SELECT status, claimed_by_name, finance_note FROM payments WHERE id = 1").fetchone()
        claim = conn.execute("SELECT status FROM claims WHERE id = 1").fetchone()
        audit_row = conn.execute("SELECT action, detail_json FROM audit_logs").fetchone()

        self.assertEqual(detail["count"], 1)
        self.assertEqual(claim["status"], "accepted")
        self.assertEqual(payment["status"], "claimed")
        self.assertIsNone(payment["claimed_by_name"])
        self.assertEqual(payment["finance_note"], "部分认领金额已覆盖整笔到款")
        self.assertEqual(audit_row["action"], "accept_payment_claims")

    def test_cancel_my_claim_returns_payment_to_pending(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY,
                payment_id INTEGER NOT NULL,
                amount_cents INTEGER NOT NULL,
                actor_id TEXT NOT NULL,
                status TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.execute(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                1,
                776000,
                "claimed",
                "中国教育创新年会事业部",
                "创新年会",
                "user-a",
                "测试用户",
                "2026-07-01 11:09:00",
                "报名费",
                "",
                "",
            ),
        )
        conn.execute(
            "INSERT INTO claims (id, payment_id, amount_cents, actor_id, status) VALUES (?, ?, ?, ?, ?)",
            (1, 1, 776000, "user-a", "accepted"),
        )
        actor = {"id": "user-a", "name": "测试用户", "role": "claimant"}

        detail = self.app.cancel_my_claim(conn, actor, 1)

        payment = conn.execute("SELECT status, claimed_by, claimed_department FROM payments WHERE id = 1").fetchone()
        claim = conn.execute("SELECT status FROM claims WHERE id = 1").fetchone()
        audit_row = conn.execute("SELECT action, detail_json FROM audit_logs").fetchone()

        self.assertEqual(detail["payment_status"], "pending")
        self.assertEqual(claim["status"], "canceled")
        self.assertEqual(payment["status"], "pending")
        self.assertIsNone(payment["claimed_by"])
        self.assertIsNone(payment["claimed_department"])
        self.assertEqual(audit_row["action"], "cancel_my_claim")

    def test_cancel_my_claim_rejects_other_users_claim(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY,
                payment_id INTEGER NOT NULL,
                amount_cents INTEGER NOT NULL,
                actor_id TEXT NOT NULL,
                status TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.execute(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note)
            VALUES (?, ?, ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)
            """,
            (1, 10000, "pending_confirm"),
        )
        conn.execute(
            "INSERT INTO claims (id, payment_id, amount_cents, actor_id, status) VALUES (?, ?, ?, ?, ?)",
            (1, 1, 10000, "user-a", "pending"),
        )

        with self.assertRaises(self.app.HTTPException) as ctx:
            self.app.cancel_my_claim(conn, {"id": "user-b", "name": "其他用户", "role": "claimant"}, 1)

        self.assertEqual(ctx.exception.status_code, 403)
        claim = conn.execute("SELECT status FROM claims WHERE id = 1").fetchone()
        self.assertEqual(claim["status"], "pending")

    def test_admin_reject_resets_payment_to_initial_pending_state(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT,
                closed_at TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY,
                payment_id INTEGER NOT NULL,
                amount_cents INTEGER NOT NULL,
                actor_id TEXT NOT NULL,
                status TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.execute(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                1,
                199800,
                "claimed",
                "培训事业部",
                "教师培训",
                "user-a",
                "认领同事",
                "2026-07-03 10:00:00",
                "教师培训网络课程费",
                "原认领备注",
                "",
            ),
        )
        conn.execute(
            "INSERT INTO claims (id, payment_id, amount_cents, actor_id, status) VALUES (?, ?, ?, ?, ?)",
            (1, 1, 199800, "user-a", "accepted"),
        )

        detail = self.app.reset_payment_to_initial_pending(
            conn,
            {"id": "finance", "name": "财务管理员", "role": "admin"},
            1,
            "部门归属不对",
            action="reject_payment",
        )

        payment = conn.execute(
            """
            SELECT status, claimed_department, claimed_team, claimed_by, claimed_by_name,
                   claimed_at, customer_project, claim_note, finance_note
            FROM payments WHERE id = 1
            """
        ).fetchone()
        claim = conn.execute("SELECT status FROM claims WHERE id = 1").fetchone()
        audit_row = conn.execute("SELECT action, detail_json FROM audit_logs").fetchone()

        self.assertEqual(detail["payment_status"], "pending")
        self.assertEqual(detail["reset_claim_ids"], [1])
        self.assertEqual(self.app.claim_totals(conn, 1), {"active": 0, "accepted": 0, "pending": 0})
        self.assertEqual(payment["status"], "pending")
        self.assertIsNone(payment["claimed_department"])
        self.assertIsNone(payment["claimed_team"])
        self.assertIsNone(payment["claimed_by"])
        self.assertIsNone(payment["claimed_by_name"])
        self.assertIsNone(payment["claimed_at"])
        self.assertIsNone(payment["customer_project"])
        self.assertIsNone(payment["claim_note"])
        self.assertEqual(payment["finance_note"], "驳回退回：部门归属不对")
        self.assertEqual(claim["status"], "rejected")
        self.assertEqual(audit_row["action"], "reject_payment")

    def test_startup_repair_moves_existing_rejected_payments_back_to_pending(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT,
                closed_at TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY,
                payment_id INTEGER NOT NULL,
                amount_cents INTEGER NOT NULL,
                actor_id TEXT NOT NULL,
                status TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.executemany(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note, closed_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    1,
                    199800,
                    "rejected",
                    "培训事业部",
                    "教师培训",
                    "user-a",
                    "认领同事",
                    "2026-07-03 10:00:00",
                    "教师培训网络课程费",
                    "原认领备注",
                    "驳回退回",
                    "",
                ),
                (2, 50000, "closed", "年会", "", "user-b", "其他同事", "", "报名费", "", "", "2026-07-03 12:00:00"),
            ],
        )
        conn.executemany(
            "INSERT INTO claims (id, payment_id, amount_cents, actor_id, status) VALUES (?, ?, ?, ?, ?)",
            [
                (1, 1, 199800, "user-a", "accepted"),
                (2, 2, 50000, "user-b", "accepted"),
            ],
        )

        detail = self.app.repair_rejected_payments_to_pending(conn)

        payments = conn.execute("SELECT id, status, claimed_by, claimed_department, closed_at FROM payments ORDER BY id").fetchall()
        claims = conn.execute("SELECT id, status FROM claims ORDER BY id").fetchall()
        audit_row = conn.execute("SELECT action, payment_id FROM audit_logs").fetchone()

        self.assertEqual(detail["payment_ids"], [1])
        self.assertEqual(detail["reset_claim_ids"], [1])
        self.assertEqual([row["status"] for row in payments], ["pending", "closed"])
        self.assertIsNone(payments[0]["claimed_by"])
        self.assertIsNone(payments[0]["claimed_department"])
        self.assertIsNone(payments[0]["closed_at"])
        self.assertEqual([row["status"] for row in claims], ["rejected", "accepted"])
        self.assertEqual(audit_row["action"], "repair_rejected_payment_to_pending")
        self.assertEqual(audit_row["payment_id"], 1)

    def test_repair_pending_claims_converts_legacy_waiting_claims_to_accepted(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY,
                payment_id INTEGER NOT NULL,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        conn.execute(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note)
            VALUES (?, ?, ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)
            """,
            (1, 199800, "pending_confirm"),
        )
        conn.executemany(
            "INSERT INTO claims (id, payment_id, amount_cents, status) VALUES (?, ?, ?, ?)",
            [(1, 1, 99900, "pending"), (2, 1, 99900, "pending")],
        )

        detail = self.app.repair_pending_claims_to_accepted(conn)

        payment = conn.execute("SELECT status, claimed_by_name, finance_note FROM payments WHERE id = 1").fetchone()
        claims = conn.execute("SELECT status FROM claims ORDER BY id").fetchall()
        audit_row = conn.execute("SELECT action, payment_id FROM audit_logs").fetchone()

        self.assertEqual(detail["payment_ids"], [1])
        self.assertEqual(detail["claim_ids"], [1, 2])
        self.assertEqual(payment["status"], "claimed")
        self.assertIsNone(payment["claimed_by_name"])
        self.assertEqual(payment["finance_note"], "部分认领金额已覆盖整笔到款")
        self.assertEqual([row["status"] for row in claims], ["accepted", "accepted"])
        self.assertEqual(audit_row["action"], "repair_pending_claims_to_accepted")
        self.assertEqual(audit_row["payment_id"], 1)

    def test_submit_split_claims_creates_accepted_lines_and_refreshes_payment(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payment_id INTEGER NOT NULL,
                department TEXT NOT NULL,
                team TEXT,
                amount_cents INTEGER NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                customer_project TEXT,
                contract_invoice TEXT,
                note TEXT,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        department = self.app.DEPARTMENTS[0]
        team = next(iter(self.app.CATALOG[department]))
        projects = self.app.CATALOG[department][team][:2]
        if len(projects) < 2:
            projects = [projects[0], projects[0]]
        conn.execute(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note)
            VALUES (?, ?, ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)
            """,
            (1, 10000, "pending"),
        )
        actor = {"id": "split-user", "name": "分摊同事", "role": "claimant"}

        detail = self.app.submit_split_claims(
            conn,
            actor,
            1,
            [department, department],
            [team, team],
            [projects[0], projects[1]],
            ["60", "40"],
            ["2本杂志", "2个参会名额"],
        )

        payment = conn.execute("SELECT status, finance_note FROM payments WHERE id = 1").fetchone()
        claim_rows = conn.execute("SELECT amount_cents, status, note FROM claims ORDER BY id").fetchall()
        audit_row = conn.execute("SELECT action, detail_json FROM audit_logs").fetchone()

        self.assertEqual(detail["count"], 2)
        self.assertEqual(detail["amount_cents"], 10000)
        self.assertEqual(detail["payment_status"], "claimed")
        self.assertEqual(payment["status"], "claimed")
        self.assertEqual(payment["finance_note"], "部分认领金额已覆盖整笔到款")
        self.assertEqual([row["amount_cents"] for row in claim_rows], [6000, 4000])
        self.assertEqual([row["status"] for row in claim_rows], ["accepted", "accepted"])
        self.assertEqual([row["note"] for row in claim_rows], ["2本杂志", "2个参会名额"])
        self.assertEqual(audit_row["action"], "split_claim_submit")

    def test_submit_batch_claims_creates_one_accepted_claim_per_payment(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payment_id INTEGER NOT NULL,
                department TEXT NOT NULL,
                team TEXT,
                amount_cents INTEGER NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                customer_project TEXT,
                contract_invoice TEXT,
                note TEXT,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        department = self.app.DEPARTMENTS[0]
        team = next(iter(self.app.CATALOG[department]))
        project = self.app.CATALOG[department][team][0]
        conn.executemany(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note)
            VALUES (?, ?, ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)
            """,
            [(1, 10000, "pending"), (2, 25000, "pending")],
        )
        actor = {"id": "batch-user", "name": "批量同事", "role": "claimant"}

        detail = self.app.submit_batch_claims(
            conn,
            actor,
            [1, 2],
            department,
            team,
            project,
            "统一采购",
        )

        payments = conn.execute(
            """
            SELECT id, status, claimed_department, claimed_team, claimed_by,
                   claimed_by_name, customer_project, claim_note
            FROM payments
            ORDER BY id
            """
        ).fetchall()
        claims = conn.execute(
            "SELECT payment_id, amount_cents, department, team, customer_project, note, status FROM claims ORDER BY id"
        ).fetchall()
        audit_row = conn.execute("SELECT action, detail_json FROM audit_logs").fetchone()

        self.assertEqual(detail["count"], 2)
        self.assertEqual(detail["amount_cents"], 35000)
        self.assertEqual(detail["skipped"], [])
        self.assertEqual([row["status"] for row in payments], ["claimed", "claimed"])
        self.assertTrue(all(row["claimed_department"] == department for row in payments))
        self.assertTrue(all(row["claimed_team"] == team for row in payments))
        self.assertTrue(all(row["claimed_by"] == "batch-user" for row in payments))
        self.assertTrue(all(row["claimed_by_name"] == "批量同事" for row in payments))
        self.assertTrue(all(row["customer_project"] == project for row in payments))
        self.assertTrue(all(row["claim_note"] == "统一采购" for row in payments))
        self.assertEqual([row["payment_id"] for row in claims], [1, 2])
        self.assertEqual([row["amount_cents"] for row in claims], [10000, 25000])
        self.assertTrue(all(row["department"] == department for row in claims))
        self.assertTrue(all(row["team"] == team for row in claims))
        self.assertTrue(all(row["customer_project"] == project for row in claims))
        self.assertEqual([row["note"] for row in claims], ["统一采购", "统一采购"])
        self.assertEqual([row["status"] for row in claims], ["accepted", "accepted"])
        self.assertEqual(audit_row["action"], "batch_claim_submit")
        self.assertEqual(json.loads(audit_row["detail_json"])["payment_ids"], [1, 2])

    def test_submit_batch_claims_uses_remaining_amount_and_skips_invalid_rows(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL,
                claimed_department TEXT,
                claimed_team TEXT,
                claimed_by TEXT,
                claimed_by_name TEXT,
                claimed_at TEXT,
                customer_project TEXT,
                claim_note TEXT,
                finance_note TEXT
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payment_id INTEGER NOT NULL,
                department TEXT NOT NULL,
                team TEXT,
                amount_cents INTEGER NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                customer_project TEXT,
                contract_invoice TEXT,
                note TEXT,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                at TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                actor_role TEXT NOT NULL,
                action TEXT NOT NULL,
                payment_id INTEGER,
                detail_json TEXT NOT NULL,
                ip TEXT
            );
            """
        )
        department = self.app.DEPARTMENTS[0]
        team = next(iter(self.app.CATALOG[department]))
        project = self.app.CATALOG[department][team][0]
        conn.executemany(
            """
            INSERT INTO payments
                (id, amount_cents, status, claimed_department, claimed_team, claimed_by,
                 claimed_by_name, claimed_at, customer_project, claim_note, finance_note)
            VALUES (?, ?, ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)
            """,
            [(1, 10000, "partial_claiming"), (2, 8000, "closed")],
        )
        conn.execute(
            """
            UPDATE payments
            SET claimed_department = ?, claimed_team = ?, claimed_by = ?, claimed_by_name = ?,
                customer_project = ?, claim_note = ?
            WHERE id = 1
            """,
            ("原部门", "原中心", "other-user", "其他同事", "原项目", "原备注"),
        )
        conn.execute(
            """
            INSERT INTO claims
                (payment_id, department, team, amount_cents, actor_id, actor_name, customer_project, contract_invoice, note, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, '', ?, 'accepted', ?)
            """,
            (1, department, team, 3000, "other-user", "其他同事", project, "已认领部分", "2026-07-01 10:00:00"),
        )
        actor = {"id": "batch-user", "name": "批量同事", "role": "claimant"}

        detail = self.app.submit_batch_claims(
            conn,
            actor,
            [1, 2, 999],
            department,
            team,
            project,
            "",
        )

        payment = conn.execute(
            "SELECT status, finance_note, claimed_department, claimed_team, claimed_by, claimed_by_name, customer_project, claim_note FROM payments WHERE id = 1"
        ).fetchone()
        claims = conn.execute("SELECT payment_id, amount_cents, actor_id, status FROM claims ORDER BY id").fetchall()

        self.assertEqual(detail["count"], 1)
        self.assertEqual(detail["amount_cents"], 7000)
        self.assertEqual(detail["payment_ids"], [1])
        self.assertEqual(
            detail["skipped"],
            [
                {"payment_id": 2, "reason": "status", "status": "closed"},
                {"payment_id": 999, "reason": "not_found"},
            ],
        )
        self.assertEqual(payment["status"], "claimed")
        self.assertEqual(payment["finance_note"], "部分认领金额已覆盖整笔到款")
        self.assertEqual(payment["claimed_department"], "多部门分摊")
        self.assertIsNone(payment["claimed_team"])
        self.assertIsNone(payment["claimed_by"])
        self.assertIsNone(payment["claimed_by_name"])
        self.assertIsNone(payment["customer_project"])
        self.assertIsNone(payment["claim_note"])
        self.assertEqual([row["amount_cents"] for row in claims], [3000, 7000])
        self.assertEqual(claims[-1]["actor_id"], "batch-user")
        self.assertEqual(claims[-1]["status"], "accepted")

    def test_search_pending_list_shows_unique_payment_ids(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "search-pending-ids.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-06 10:00:00", "finance", "confirmed"),
                    )
                    conn.executemany(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (52, 1, "2026-07-06 10:00:00", "2026-07-06 10:01:00", "2026-07-06", "10:00:00", "客户A", 10000, "备注A", "蒲公英智库", "pending"),
                            (53, 1, "2026-07-06 10:00:00", "2026-07-06 10:01:00", "2026-07-06", "10:00:00", "客户B", 20000, "备注B", "蒲公英智库", "pending"),
                        ],
                    )
                self.app.actor_from_request = lambda request: {
                    "id": "user-a",
                    "name": "认领同事",
                    "role": "claimant",
                    "department": self.app.DEPARTMENTS[0],
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.77"),
                    cookies={},
                    query_params={},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.search_page(request)
                html = response.body.decode("utf-8")
                pending_section = html[html.index("待认领列表"):]

                self.assertIn("<th>ID</th>", pending_section)
                self.assertIn('<td class="nowrap">#52</td>', pending_section)
                self.assertIn('<td class="nowrap">#53</td>', pending_section)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_run_search_returns_all_matching_rows_without_limit(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.execute(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                received_date TEXT,
                payer_name TEXT,
                amount_cents INTEGER,
                bank_note TEXT,
                serial_no TEXT,
                status TEXT
            )
            """
        )
        conn.executemany(
            """
            INSERT INTO payments (id, received_date, payer_name, amount_cents, bank_note, serial_no, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (i, "2026-07-03", f"银联POS款客户{i}", 10000 + i, "银联POS款", f"POS{i:03d}", "pending")
                for i in range(1, 13)
            ],
        )

        rows = self.app.run_search(conn, "银联POS款")

        self.assertEqual(len(rows), 12)

    def test_project_select_uses_keyword_input_with_existing_project_options(self) -> None:
        department = next(dept for dept, teams in self.app.CATALOG.items() if teams)
        team = next(team for team, projects in self.app.CATALOG[department].items() if projects)
        project = self.app.CATALOG[department][team][0]

        html = self.app.project_select("customer_project", department, team, project, required=True)

        self.assertIn('name="customer_project"', html)
        self.assertIn('class="cs-project"', html)
        self.assertIn('placeholder="输入项目关键词"', html)
        self.assertIn('required', html)
        self.assertIn("<datalist", html)
        self.assertIn(f'value="{project}"', html)
        self.assertIn(f'<option value="{project}"></option>', html)
        self.assertNotIn("<select", html)

    def test_split_claim_form_can_add_more_rows_client_side(self) -> None:
        row = {
            "id": 66,
            "amount_cents": 1016800,
            "status": "pending",
        }

        html = self.app.split_claim_form_html(row)

        self.assertEqual(html.count('name="departments"'), 6)
        self.assertEqual(html.count('class="split-row-number"'), 6)
        self.assertIn('class="secondary split-add-row"', html)
        self.assertIn("＋ 添加分摊行", html)
        self.assertIn("addSplitClaimRow", self.app.CASCADE_JS)
        self.assertIn("createProjectInput", self.app.CASCADE_JS)

    def test_admin_payment_sort_clause_is_whitelisted(self) -> None:
        clause, sort, direction = self.app.admin_payment_order_clause("amount", "asc")
        self.assertEqual(sort, "amount")
        self.assertEqual(direction, "asc")
        self.assertEqual(clause, "amount_cents ASC, id ASC")

        clause, sort, direction = self.app.admin_payment_order_clause("bad; DROP TABLE payments", "sideways")
        self.assertEqual(sort, "id")
        self.assertEqual(direction, "desc")
        self.assertEqual(clause, "id DESC")

        clause, sort, direction = self.app.admin_payment_order_clause("status", "desc")
        self.assertEqual(sort, "status")
        self.assertEqual(direction, "desc")
        self.assertIn("CASE status", clause)
        self.assertTrue(clause.endswith("DESC, id DESC"))

    def test_admin_sort_header_supports_partial_table_refresh(self) -> None:
        header = self.app.admin_sort_th("金额", "amount", "id", "desc")

        self.assertIn('href="/admin?sort=amount&amp;dir=asc"', header)
        self.assertIn('data-table-url="/admin/payments/table?sort=amount&amp;dir=asc"', header)
        self.assertIn("sort-link", header)

    def test_admin_payment_row_shows_claim_details_before_remark_description(self) -> None:
        old_db_path = self.app.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "admin-row.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-06 10:00:00", "finance", "confirmed"),
                    )
                    conn.execute(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status,
                             claimed_department, claimed_team, claimed_by, claimed_by_name, claimed_at,
                             customer_project, claim_note)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            1,
                            1,
                            "2026-07-06 10:00:00",
                            "2026-07-06 10:01:00",
                            "2026-07-06",
                            "10:00:00",
                            "测试客户",
                            178000,
                            "银行备注",
                            "蒲公英智库",
                            "pending_confirm",
                            "培训事业部",
                            "会议中心",
                            "claim-user",
                            "认领同事",
                            "2026-07-06 10:05:00",
                            "2026年班主任峰会",
                            "统一采购",
                        ),
                    )
                    conn.executemany(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, "培训事业部", "会议中心", 100000, "claim-user", "张三", "项目A", "", "统一采购", "pending", "2026-07-06 10:05:00"),
                            (1, "年会事业部", "创新中心", 78000, "claim-user-2", "李四", "项目B", "", "补充说明", "accepted", "2026-07-06 10:06:00"),
                            (1, "旧部门", "旧中心", 178000, "old-user", "旧认领人", "旧项目", "", "旧备注", "rejected", "2026-07-06 10:04:00"),
                        ],
                    )
                    conn.commit()
                    row = conn.execute("SELECT * FROM payments WHERE id = 1").fetchone()

                    html = self.app.render_admin_payment_row(row, {"id": "finance", "name": "财务", "role": "admin"})

                self.assertIn("认领明细：", html)
                self.assertIn("培训事业部 · 会议中心 · 项目A · 张三 · ¥ 1,000.00", html)
                self.assertIn("年会事业部 · 创新中心 · 项目B · 李四 · ¥ 780.00", html)
                self.assertIn("备注说明：统一采购；补充说明", html)
                self.assertNotIn("认领备注", html)
                self.assertNotIn("旧部门", html)
                self.assertNotIn("旧备注", html)
                self.assertLess(html.index("银行备注"), html.index("认领明细："))
                self.assertLess(html.index("认领明细："), html.index("备注说明：统一采购；补充说明"))
        finally:
            self.app.DB_PATH = old_db_path

    def test_admin_payment_row_keeps_single_claim_compact_without_empty_remark(self) -> None:
        old_db_path = self.app.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "admin-row-no-note.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-06 10:00:00", "finance", "confirmed"),
                    )
                    conn.execute(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status,
                             claimed_department, claimed_team, claimed_by, claimed_by_name, claimed_at,
                             customer_project, claim_note)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            1,
                            1,
                            "2026-07-06 10:00:00",
                            "2026-07-06 10:01:00",
                            "2026-07-06",
                            "10:00:00",
                            "测试客户",
                            100000,
                            "银行备注",
                            "蒲公英智库",
                            "pending_confirm",
                            "培训事业部",
                            "会议中心",
                            "claim-user",
                            "张三",
                            "2026-07-06 10:05:00",
                            "项目A",
                            "",
                        ),
                    )
                    conn.execute(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (1, "培训事业部", "会议中心", 100000, "claim-user", "张三", "项目A", "", "", "pending", "2026-07-06 10:05:00"),
                    )
                    conn.commit()
                    row = conn.execute("SELECT * FROM payments WHERE id = 1").fetchone()

                    html = self.app.render_admin_payment_row(row, {"id": "finance", "name": "财务", "role": "admin"})

                self.assertNotIn("认领明细：", html)
                self.assertIn("培训事业部 · 会议中心 · 张三", html)
                self.assertIn("项目A", html)
                self.assertNotIn("备注说明：", html)
        finally:
            self.app.DB_PATH = old_db_path

    def test_admin_page_search_panel_precedes_payment_pool_with_admin_actions(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "admin-search.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-06 10:00:00", "finance", "confirmed"),
                    )
                    conn.execute(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status,
                             claimed_department, claimed_team, claimed_by, claimed_by_name, claimed_at,
                             customer_project, claim_note)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            1,
                            1,
                            "2026-07-06 10:00:00",
                            "2026-07-06 10:01:00",
                            "2026-07-06",
                            "10:00:00",
                            "测试客户",
                            178000,
                            "银行备注",
                            "蒲公英智库",
                            "pending_confirm",
                            "培训事业部",
                            "会议中心",
                            "claim-user",
                            "认领同事",
                            "2026-07-06 10:05:00",
                            "2026年班主任峰会",
                            "统一采购",
                        ),
                    )
                    conn.execute(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (1, "培训事业部", "会议中心", 178000, "claim-user", "认领同事", "2026年班主任峰会", "", "统一采购", "pending", "2026-07-06 10:05:00"),
                    )
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=None,
                    cookies={},
                    query_params={},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.admin_page(request, admin_q="测试客户")
                html = response.body.decode("utf-8")

                self.assertLess(html.index("到款认领搜索"), html.index("全量认领池"))
                search_section = html[html.index("到款认领搜索"):html.index("全量认领池")]
                self.assertIn("测试客户", search_section)
                self.assertIn("编辑字段", search_section)
                self.assertIn("处理状态", search_section)
                self.assertNotIn("bulk-payment-checkbox", search_section)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_status_cards_omit_pending_confirm_and_fit_one_row(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "admin-stats.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-06 10:00:00", "finance", "confirmed"),
                    )
                    statuses = ["draft", "pending", "partial_claiming", "claimed", "pending_confirm", "rejected", "closed"]
                    conn.executemany(
                        """
                        INSERT INTO payments
                            (batch_id, imported_at, received_date, payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, "2026-07-06 10:00:00", "2026-07-06", f"客户{status}", 10000, "测试备注", "蒲公英智库", status)
                            for status in statuses
                        ],
                    )
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=None,
                    cookies={},
                    query_params={},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.admin_page(request)
                html = response.body.decode("utf-8")
                stat_section = html[html.index("admin-stat-grid"):html.index("按日期导出")]

                self.assertIn("admin-stat-grid", stat_section)
                self.assertNotIn("待财务确认", stat_section)
                for label in ["待确认", "待认领", "部分认领中", "已认领", "已驳回", "已关闭"]:
                    self.assertIn(label, stat_section)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_export_panel_uses_selected_date_controls(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "admin-export-panel.db"
                self.app.init_db()
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.81"),
                    cookies={},
                    query_params={},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.admin_page(request)
                html = response.body.decode("utf-8")

                self.assertIn("按日期导出", html)
                self.assertIn('id="export-start-date" type="date"', html)
                self.assertIn('id="export-end-date" type="date"', html)
                self.assertIn('id="export-date-csv"', html)
                self.assertIn("开始日期", html)
                self.assertIn("结束日期", html)
                self.assertIn("下载 CSV", html)
                self.assertIn("复制纯文本", html)
                self.assertNotIn("下载今日 CSV", html)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_page_section_order_matches_management_flow(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "admin-order.db"
                self.app.init_db()
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=None,
                    cookies={},
                    query_params={},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.admin_page(request)
                html = response.body.decode("utf-8")

                order = [
                    "导入流水",
                    "到款认领搜索",
                    "最近导入批次",
                    "全量认领池",
                    "项目管理",
                    "成员部门",
                ]
                positions = [html.index(label) for label in order]
                self.assertEqual(positions, sorted(positions))
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_page_progressively_reveals_batches_and_payment_pool(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "admin-progressive.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.executemany(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        [
                            (i, f"批次 {i}", f"2026-07-{i:02d} 10:00:00", "finance", "confirmed")
                            for i in range(1, 13)
                        ],
                    )
                    conn.executemany(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (
                                i,
                                1,
                                "2026-07-07 10:00:00",
                                "2026-07-07 10:01:00",
                                "2026-07-07",
                                "10:00:00",
                                f"测试客户{i}",
                                10000 + i,
                                "测试备注",
                                "重庆市蒲公英未来科技有限公司",
                                "pending",
                            )
                            for i in range(1, 26)
                        ],
                    )
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=None,
                    cookies={},
                    query_params={},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.admin_page(request)
                html = response.body.decode("utf-8")

                batch_section = html[html.index("最近导入批次"):html.index("全量认领池")]
                pool_section = html[html.index("全量认领池"):html.index("项目管理")]

                self.assertEqual(batch_section.count('<tr data-progressive-group="admin-batches"'), 12)
                self.assertEqual(batch_section.count('style="display:none"'), 2)
                self.assertIn('data-progressive-step="10"', batch_section)
                self.assertIn("显示更多", batch_section)
                self.assertEqual(pool_section.count('<tr data-progressive-group="payment-pool"'), 25)
                self.assertEqual(pool_section.count('style="display:none"'), 5)
                self.assertIn('data-progressive-step="20"', pool_section)
                self.assertIn("显示更多", pool_section)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_general_manager_is_identity_not_admin_permission(self) -> None:
        old_db_path = self.app.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "roles.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.executemany(
                        """
                        INSERT INTO app_users (open_id, name, managed_role, is_admin, created_at, last_login)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        [
                            ("gm-user", "事业部总经理", "general_manager", 0, "2026-06-25 10:00:00", "2026-06-25 10:00:00"),
                            ("legacy-admin", "旧管理员", "claimant", 1, "2026-06-25 10:00:00", "2026-06-25 10:00:00"),
                        ],
                    )

                self.assertEqual(self.app.compute_role("gm-user"), "general_manager")
                self.assertEqual(self.app.compute_role("legacy-admin"), "admin")
                with self.assertRaises(self.app.HTTPException):
                    self.app.require_admin({"role": "general_manager"})
                with self.assertRaises(self.app.HTTPException):
                    self.app.require_superadmin({"role": "general_manager"})
        finally:
            self.app.DB_PATH = old_db_path

    def test_personal_dashboard_respects_role_scope(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                received_date TEXT,
                payer_name TEXT,
                receiver_company TEXT,
                bank_note TEXT,
                amount_cents INTEGER NOT NULL,
                claimed_department TEXT,
                status TEXT NOT NULL
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY,
                payment_id INTEGER NOT NULL,
                department TEXT NOT NULL,
                actor_id TEXT NOT NULL,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL
            );
            """
        )
        conn.executemany(
            """
            INSERT INTO payments (id, received_date, payer_name, receiver_company, bank_note, amount_cents, claimed_department, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (1, "2026-06-25", "未认领客户", "蒲公英教育科技", "测试摘要1", 10000, "", "pending"),
                (2, "2026-06-25", "年会客户A", "蒲公英智库", "测试摘要2", 20000, "年会事业部", "claimed"),
                (3, "2026-06-25", "混合客户", "蒲公英教育科技", "测试摘要3", 30000, "", "pending_confirm"),
                (4, "2026-06-25", "关闭客户", "蒲公英教育科技", "测试摘要4", 99900, "", "closed"),
            ],
        )
        department = self.app.DEPARTMENTS[0]
        other_department = self.app.DEPARTMENTS[1] if len(self.app.DEPARTMENTS) > 1 else "其他部门"
        conn.executemany(
            "INSERT INTO claims (id, payment_id, department, actor_id, amount_cents, status) VALUES (?, ?, ?, ?, ?, ?)",
            [
                (1, 2, department, "user-a", 20000, "accepted"),
                (2, 3, department, "user-b", 12000, "pending"),
                (3, 3, other_department, "user-c", 18000, "pending"),
                (4, 4, department, "user-a", 99900, "accepted"),
            ],
        )
        today = date(2026, 6, 25)

        admin_dashboard = self.app.personal_dashboard_data(
            conn,
            {"id": "admin", "role": "admin", "department": "未设置部门"},
            today,
        )[0]
        gm_dashboard = self.app.personal_dashboard_data(
            conn,
            {"id": "gm", "role": "general_manager", "department": department},
            today,
        )[0]
        user_department_dashboard = self.app.personal_dashboard_data(
            conn,
            {"id": "user-b", "role": "claimant", "department": department},
            today,
        )[0]
        user_all_dashboard = self.app.personal_dashboard_data(
            conn,
            {
                "id": "user-b",
                "role": "claimant",
                "department": department,
                "dashboard_scope_label": "全部角色",
                "dashboard_scopes": [
                    {"department": department, "team": ""},
                    {"department": other_department, "team": ""},
                ],
            },
            today,
        )[0]

        self.assertEqual(admin_dashboard["total_cents"], 60000)
        self.assertIn(("未认领", 10000), admin_dashboard["departments"])
        self.assertIn((department, 32000), admin_dashboard["departments"])
        self.assertIn((other_department, 18000), admin_dashboard["departments"])
        self.assertEqual(gm_dashboard["total_cents"], 32000)
        self.assertEqual(gm_dashboard["departments"], [(department, 32000)])
        self.assertEqual(user_department_dashboard["total_cents"], 32000)
        self.assertEqual(user_department_dashboard["departments"], [(department, 32000)])
        self.assertIn(("混合客户", 12000), user_department_dashboard["customers"])
        self.assertIn(("年会客户A", 20000), user_department_dashboard["customers"])
        self.assertTrue(all(row["department"] == department for row in user_department_dashboard["rows"]))
        self.assertFalse(any(row["department"] == other_department for row in user_department_dashboard["rows"]))
        self.assertEqual(user_all_dashboard["total_cents"], 50000)
        self.assertIn((department, 32000), user_all_dashboard["departments"])
        self.assertIn((other_department, 18000), user_all_dashboard["departments"])

    def test_today_claim_plain_text_groups_payer_and_appends_unclaimed_summary(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY,
                received_date TEXT,
                payer_name TEXT,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL
            );
            CREATE TABLE claims (
                id INTEGER PRIMARY KEY,
                payment_id INTEGER NOT NULL,
                department TEXT NOT NULL,
                customer_project TEXT,
                amount_cents INTEGER NOT NULL,
                status TEXT NOT NULL
            );
            """
        )
        conn.executemany(
            "INSERT INTO payments (id, received_date, payer_name, amount_cents, status) VALUES (?, ?, ?, ?, ?)",
            [
                (1, "2026-07-01", "银联POS款", 55740, "claimed"),
                (2, "2026-07-01", "银联POS款", 59130, "claimed"),
                (3, "2026-07-01", "南京工业大学实验小学", 64200, "claimed"),
                (4, "2026-07-01", "未认领客户", 20000, "pending"),
                (5, "2026-07-01", "已取消客户", 99900, "closed"),
                (6, "2026-07-02", "明日客户", 10000, "claimed"),
            ],
        )
        conn.executemany(
            """
            INSERT INTO claims (id, payment_id, department, customer_project, amount_cents, status)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                (1, 1, "学生发展事业部", "咖啡收入", 40540, "accepted"),
                (2, 1, "学生发展事业部", "产品与文创", 15200, "pending"),
                (3, 2, "学生发展事业部", "阅读创新", 59130, "accepted"),
                (4, 3, "教育智能研究院", "云智库地图", 60600, "accepted"),
                (5, 3, "教育智能研究院", "家长认知与行为地图", 3600, "pending"),
                (6, 5, "学生发展事业部", "已取消项目", 99900, "canceled"),
                (7, 6, "学生发展事业部", "明日项目", 10000, "accepted"),
            ],
        )

        text = self.app.build_today_claim_plain_text(conn, "2026-07-01")

        self.assertEqual(
            text,
            "\n".join(
                [
                    "7月1日",
                    "",
                    "1.银联POS款\t1,148.70元（学生发展事业部  咖啡收入405.40元，产品与文创152.00元，阅读创新591.30元）",
                    "2.南京工业大学实验小学\t642.00元（教育智能研究院  云智库地图606.00元，家长认知与行为地图36.00元）",
                    "3.未认领\t200.00元",
                    "今日合计：1,990.70元",
                ]
            ),
        )
        self.assertNotIn("未认领客户", text)
        self.assertNotIn("已取消客户", text)
        self.assertNotIn("明日客户", text)

    def test_admin_csv_export_uses_selected_date(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "export-date.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-01 10:00:00", "finance", "confirmed"),
                    )
                    conn.executemany(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, 1, "2026-07-01 10:00:00", "2026-07-01 10:01:00", "2026-07-01", "10:00:00", "目标客户", 10000, "目标备注", "蒲公英智库", "claimed"),
                            (2, 1, "2026-07-02 10:00:00", "2026-07-02 10:01:00", "2026-07-02", "10:00:00", "其他客户", 20000, "其他备注", "蒲公英智库", "pending"),
                        ],
                    )
                    conn.execute(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (1, "培训事业部", "会议中心", 10000, "claim-user", "张三", "目标项目", "", "", "accepted", "2026-07-01 10:05:00"),
                    )
                    conn.commit()
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.82"),
                    cookies={},
                    query_params={"date": "2026-07-01"},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.export_today_payments(request)
                body = response.body.decode("utf-8-sig")

                self.assertIn("目标客户", body)
                self.assertIn("目标备注", body)
                self.assertNotIn("其他客户", body)
                self.assertEqual(response.headers["content-disposition"], 'attachment; filename="claim_pool_2026-07-01.csv"')
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_csv_export_appends_unclaimed_summary_after_claim_rows(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "export-unclaimed-summary.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-01 10:00:00", "finance", "confirmed"),
                    )
                    conn.executemany(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, 1, "2026-07-01 10:00:00", "2026-07-01 10:01:00", "2026-07-01", "10:00:00", "已认领客户", 10000, "已认领备注", "蒲公英智库", "claimed"),
                            (2, 1, "2026-07-01 10:00:00", "2026-07-01 10:01:00", "2026-07-01", "10:00:00", "未认领客户", 25000, "未认领备注", "蒲公英智库", "pending"),
                            (3, 1, "2026-07-01 10:00:00", "2026-07-01 10:01:00", "2026-07-01", "10:00:00", "部分认领客户", 50000, "部分认领备注", "蒲公英智库", "partial_claiming"),
                        ],
                    )
                    conn.executemany(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, "培训事业部", "会议中心", 10000, "claim-user", "张三", "已认领项目", "", "", "accepted", "2026-07-01 10:05:00"),
                            (3, "培训事业部", "会议中心", 30000, "claim-user", "张三", "部分认领项目", "", "", "accepted", "2026-07-01 10:06:00"),
                        ],
                    )
                    conn.commit()
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.88"),
                    cookies={},
                    query_params={"date": "2026-07-01"},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.export_today_payments(request)
                body = response.body.decode("utf-8-sig")
                data_lines = body.splitlines()[1:]

                self.assertIn("已认领客户", body)
                self.assertIn("部分认领客户", body)
                self.assertNotIn("未认领客户", body)
                self.assertNotIn("未认领备注", body)
                self.assertIn("未认领", data_lines[-1])
                self.assertIn("450.00", data_lines[-1])
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_exports_include_compact_yyyymmdd_dates_after_repair(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "export-compact-date.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-07 10:00:00", "finance", "confirmed"),
                    )
                    conn.execute(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            60,
                            1,
                            "2026-07-07 10:00:00",
                            "2026-07-07 10:01:00",
                            "20260707",
                            "",
                            "银联POS款",
                            356000,
                            "深圳市南山实验教育集团园丁学校班主任培训",
                            "重庆市蒲公英未来科技有限公司",
                            "claimed",
                        ),
                    )
                    conn.execute(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            60,
                            "培训事业部",
                            "会议中心",
                            356000,
                            "claim-user",
                            "赵军红",
                            "2026年班主任峰会",
                            "",
                            "",
                            "accepted",
                            "2026-07-07 10:05:00",
                        ),
                    )
                    conn.commit()

                self.app.init_db()
                with self.app.get_conn() as conn:
                    received_date = conn.execute("SELECT received_date FROM payments WHERE id = 60").fetchone()[0]
                self.assertEqual(received_date, "2026-07-07")

                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.87"),
                    cookies={},
                    query_params={"date": "2026-07-07"},
                    url=types.SimpleNamespace(scheme="http"),
                )

                csv_response = self.app.export_today_payments(request)
                csv_body = csv_response.body.decode("utf-8-sig")
                text_response = self.app.export_today_claim_plain_text(request)
                text_body = text_response.body.decode("utf-8")

                self.assertIn("银联POS款", csv_body)
                self.assertIn("2026-07-07", csv_body)
                self.assertIn("银联POS款", text_body)
                self.assertIn("2026年班主任峰会3,560.00元", text_body)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_csv_export_skips_rejected_claim_history(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "export-active-claims.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-01 10:00:00", "finance", "confirmed"),
                    )
                    conn.execute(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (30, 1, "2026-07-02 10:00:00", "2026-07-02 10:01:00", "2026-07-02", "10:00:00", "深圳市龙岗区龙岭初级中学", 480600, "班主任峰会+龙岭初级中学+线下+3人+培训费", "蒲公英智库", "claimed"),
                    )
                    conn.executemany(
                        """
                        INSERT INTO claims
                            (id, payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (15, 30, "培训事业部", "产品中心", 480600, "old-user", "旧认领人", "旧项目", "", "驳回前", "rejected", "2026-07-02 10:05:00"),
                            (19, 30, "培训事业部", "产品中心", 480600, "old-user", "旧认领人", "旧项目", "", "驳回前2", "rejected", "2026-07-02 10:06:00"),
                            (51, 30, "中台战略委员会", "整合服务中台", 480600, "new-user", "新认领人", "新项目", "", "最终认领", "accepted", "2026-07-02 11:00:00"),
                        ],
                    )
                    conn.commit()
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.86"),
                    cookies={},
                    query_params={"date": "2026-07-02"},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.export_today_payments(request)
                body = response.body.decode("utf-8-sig")
                rows = [line for line in body.splitlines() if "深圳市龙岗区龙岭初级中学" in line]

                self.assertEqual(len(rows), 1)
                self.assertIn(",51,", rows[0])
                self.assertIn("accepted", rows[0])
                self.assertIn("中台战略委员会", rows[0])
                self.assertIn("最终认领", rows[0])
                self.assertNotIn("rejected", body)
                self.assertNotIn("旧认领人", body)
                self.assertNotIn("驳回前", body)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_plain_text_export_uses_selected_date(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "export-text-date.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-01 10:00:00", "finance", "confirmed"),
                    )
                    conn.executemany(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, 1, "2026-07-01 10:00:00", "2026-07-01 10:01:00", "2026-07-01", "10:00:00", "目标客户", 10000, "目标备注", "蒲公英智库", "pending_confirm"),
                            (2, 1, "2026-07-02 10:00:00", "2026-07-02 10:01:00", "2026-07-02", "10:00:00", "其他客户", 20000, "其他备注", "蒲公英智库", "pending_confirm"),
                        ],
                    )
                    conn.executemany(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, "培训事业部", "会议中心", 10000, "claim-user", "张三", "项目A", "", "", "pending", "2026-07-01 10:05:00"),
                            (2, "年会事业部", "创新中心", 20000, "claim-user-2", "李四", "项目B", "", "", "pending", "2026-07-02 10:05:00"),
                        ],
                    )
                    conn.commit()
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.83"),
                    cookies={},
                    query_params={"date": "2026-07-01"},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.export_today_claim_plain_text(request)
                body = response.body.decode("utf-8")

                self.assertIn("7月1日", body)
                self.assertIn("目标客户", body)
                self.assertIn("项目A100.00元", body)
                self.assertNotIn("其他客户", body)
                self.assertNotIn("项目B", body)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_csv_export_uses_selected_date_range(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "export-date-range.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-01 10:00:00", "finance", "confirmed"),
                    )
                    conn.executemany(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, 1, "2026-07-01 10:00:00", "2026-07-01 10:01:00", "2026-07-01", "10:00:00", "区间外客户", 10000, "区间外备注", "蒲公英智库", "pending"),
                            (2, 1, "2026-07-02 10:00:00", "2026-07-02 10:01:00", "2026-07-02", "10:00:00", "区间客户A", 20000, "区间备注A", "蒲公英智库", "claimed"),
                            (3, 1, "2026-07-04 10:00:00", "2026-07-04 10:01:00", "2026-07-04", "10:00:00", "区间客户B", 30000, "区间备注B", "蒲公英智库", "claimed"),
                            (4, 1, "2026-07-05 10:00:00", "2026-07-05 10:01:00", "2026-07-05", "10:00:00", "区间后客户", 40000, "区间后备注", "蒲公英智库", "pending"),
                        ],
                    )
                    conn.executemany(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (2, "培训事业部", "会议中心", 20000, "claim-user", "张三", "区间项目A", "", "", "accepted", "2026-07-02 10:05:00"),
                            (3, "培训事业部", "会议中心", 30000, "claim-user", "张三", "区间项目B", "", "", "accepted", "2026-07-04 10:05:00"),
                        ],
                    )
                    conn.commit()
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.84"),
                    cookies={},
                    query_params={"start_date": "2026-07-02", "end_date": "2026-07-04"},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.export_today_payments(request)
                body = response.body.decode("utf-8-sig")

                self.assertIn("区间客户A", body)
                self.assertIn("区间客户B", body)
                self.assertNotIn("区间外客户", body)
                self.assertNotIn("区间后客户", body)
                self.assertEqual(
                    response.headers["content-disposition"],
                    'attachment; filename="claim_pool_2026-07-02_to_2026-07-04.csv"',
                )
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_admin_plain_text_export_uses_selected_date_range(self) -> None:
        old_db_path = self.app.DB_PATH
        old_actor_from_request = self.app.actor_from_request
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                self.app.DB_PATH = Path(tmpdir) / "export-text-range.db"
                self.app.init_db()
                with self.app.get_conn() as conn:
                    conn.execute(
                        """
                        INSERT INTO import_batches (id, source_name, created_at, created_by, status)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (1, "测试批次", "2026-07-01 10:00:00", "finance", "confirmed"),
                    )
                    conn.executemany(
                        """
                        INSERT INTO payments
                            (id, batch_id, imported_at, confirmed_at, received_date, received_time,
                             payer_name, amount_cents, bank_note, receiver_company, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, 1, "2026-07-01 10:00:00", "2026-07-01 10:01:00", "2026-07-01", "10:00:00", "区间外客户", 10000, "区间外备注", "蒲公英智库", "claimed"),
                            (2, 1, "2026-07-02 10:00:00", "2026-07-02 10:01:00", "2026-07-02", "10:00:00", "区间客户A", 20000, "区间备注A", "蒲公英智库", "claimed"),
                            (3, 1, "2026-07-04 10:00:00", "2026-07-04 10:01:00", "2026-07-04", "10:00:00", "区间客户B", 30000, "区间备注B", "蒲公英智库", "claimed"),
                            (4, 1, "2026-07-05 10:00:00", "2026-07-05 10:01:00", "2026-07-05", "10:00:00", "区间后客户", 40000, "区间后备注", "蒲公英智库", "claimed"),
                        ],
                    )
                    conn.executemany(
                        """
                        INSERT INTO claims
                            (payment_id, department, team, amount_cents, actor_id, actor_name,
                             customer_project, contract_invoice, note, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (1, "培训事业部", "会议中心", 10000, "claim-user", "张三", "区间外项目", "", "", "accepted", "2026-07-01 10:05:00"),
                            (2, "培训事业部", "会议中心", 20000, "claim-user", "张三", "项目A", "", "", "accepted", "2026-07-02 10:05:00"),
                            (3, "年会事业部", "创新中心", 30000, "claim-user-2", "李四", "项目B", "", "", "accepted", "2026-07-04 10:05:00"),
                            (4, "年会事业部", "创新中心", 40000, "claim-user-2", "李四", "区间后项目", "", "", "accepted", "2026-07-05 10:05:00"),
                        ],
                    )
                    conn.commit()
                self.app.actor_from_request = lambda request: {
                    "id": "finance",
                    "name": "财务",
                    "role": "admin",
                    "department": "财务部",
                    "team": "",
                    "authed": "1",
                }
                request = types.SimpleNamespace(
                    headers={},
                    client=types.SimpleNamespace(host="127.0.0.85"),
                    cookies={},
                    query_params={"start_date": "2026-07-02", "end_date": "2026-07-04"},
                    url=types.SimpleNamespace(scheme="http"),
                )

                response = self.app.export_today_claim_plain_text(request)
                body = response.body.decode("utf-8")

                self.assertIn("7月2日-7月4日", body)
                self.assertIn("区间客户A", body)
                self.assertIn("项目A200.00元", body)
                self.assertIn("区间客户B", body)
                self.assertIn("项目B300.00元", body)
                self.assertIn("区间合计：500.00元", body)
                self.assertNotIn("区间外客户", body)
                self.assertNotIn("区间后客户", body)
        finally:
            self.app.actor_from_request = old_actor_from_request
            self.app.DB_PATH = old_db_path

    def test_profile_setup_modal_only_for_authed_users_without_complete_profile(self) -> None:
        actor = {"id": "u1", "name": "测试用户", "role": "claimant", "authed": "1"}

        missing_html = self.app.profile_setup_modal_html(actor, "", "")
        self.assertIn("首次使用，请先设置部门和中心", missing_html)
        self.assertIn('action="/me/profile"', missing_html)

        department = self.app.DEPARTMENTS[0]
        team = next(iter(self.app.CATALOG[department]))
        complete_html = self.app.profile_setup_modal_html(actor, department, team)
        demo_html = self.app.profile_setup_modal_html({**actor, "authed": ""}, "", "")

        self.assertEqual(complete_html, "")
        self.assertEqual(demo_html, "")

    def test_diagnostic_log_html_has_copyable_non_secret_context(self) -> None:
        department = self.app.DEPARTMENTS[0]
        team = next(iter(self.app.CATALOG[department]))
        html = self.app.diagnostic_log_html(
            {
                "id": "ou_test_user",
                "name": "测试用户",
                "role": "general_manager",
                "department": department,
                "team": team,
                "authed": "1",
            },
            department,
            team,
        )

        self.assertIn("问题排查日志", html)
        self.assertIn('id="diagnostic-log-base"', html)
        self.assertIn('id="copy-diagnostic-log"', html)
        self.assertNotIn("<textarea", html)
        self.assertNotIn("请用户补充", html)
        self.assertIn("事业部总经理", html)
        self.assertIn(department, html)
        self.assertNotIn("cookie", html.lower())
        self.assertNotIn("secret", html.lower())

    def test_profile_self_setup_locks_after_department_and_team_are_complete(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE user_profiles (
                open_id TEXT PRIMARY KEY,
                name TEXT,
                department TEXT,
                team TEXT,
                updated_at TEXT
            );
            """
        )

        self.assertTrue(self.app.can_self_set_profile(conn, "new-user"))
        conn.execute(
            "INSERT INTO user_profiles (open_id, name, department, team, updated_at) VALUES (?, ?, ?, ?, ?)",
            ("partial-user", "待补全", self.app.DEPARTMENTS[0], "", "2026-06-26 10:00:00"),
        )
        self.assertTrue(self.app.can_self_set_profile(conn, "partial-user"))

        department = self.app.DEPARTMENTS[0]
        team = next(iter(self.app.CATALOG[department]))
        conn.execute(
            "INSERT INTO user_profiles (open_id, name, department, team, updated_at) VALUES (?, ?, ?, ?, ?)",
            ("complete-user", "已完成", department, team, "2026-06-26 10:00:00"),
        )
        self.assertFalse(self.app.can_self_set_profile(conn, "complete-user"))


if __name__ == "__main__":
    unittest.main()
